import sys
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFrame, QProgressBar,
    QListWidget, QListWidgetItem, QGroupBox, QGridLayout,
    QStatusBar, QMessageBox, QStackedWidget, QFileDialog,
    QCheckBox, QAbstractItemView, QSplitter, QRadioButton,
    QButtonGroup, QDateEdit, QDoubleSpinBox, QScrollArea, QSizePolicy
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt5.QtGui import QFont, QColor
import openpyxl
from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
#  DEMO VERİSİ — Araç 1 (Stok Bağlama)
# ══════════════════════════════════════════════════════════════
DEMO_MAMUL_AGACLARI = [
    ("MA-001", "Ahşap Masa Ünitesi"), ("MA-002", "Metal Raf Sistemi"),
    ("MA-003", "Sandalye Grubu"),     ("MA-004", "Dolap Modülü"),
    ("MA-005", "Montaj Aparatı"),     ("MA-006", "Çelik Kapı Kasası"),
    ("MA-007", "Alüminyum Profil Çerçeve"), ("MA-008", "Hidrolik Piston Ünitesi"),
]

DEMO_STOKLAR = [
    {"stok_kodu": "STK-001", "stok_adi": "Çelik Profil 40x40",
     "fatura_sayisi": 3, "toplam_tutar": "37.500,00 ₺",
     "ilk_fatura": "10.01.2024", "son_fatura": "15.03.2024",
     "tedarikci": "ABC Çelik San. Tic. A.Ş.", "fatura_turleri": "Alış, İthalat"},
    {"stok_kodu": "STK-014", "stok_adi": "Endüstriyel Boya (Epoksi)",
     "fatura_sayisi": 5, "toplam_tutar": "21.800,00 ₺",
     "ilk_fatura": "05.02.2024", "son_fatura": "18.03.2024",
     "tedarikci": "Renkli Kimya Ltd. Şti.", "fatura_turleri": "Alış, Masraf"},
    {"stok_kodu": "STK-022", "stok_adi": "Vida-Somun Seti M8",
     "fatura_sayisi": 2, "toplam_tutar": "6.400,00 ₺",
     "ilk_fatura": "20.01.2024", "son_fatura": "22.03.2024",
     "tedarikci": "Teknik Donanım A.Ş.", "fatura_turleri": "Alış"},
    {"stok_kodu": "STK-037", "stok_adi": "MDF Levha 18mm",
     "fatura_sayisi": 4, "toplam_tutar": "96.000,00 ₺",
     "ilk_fatura": "03.01.2024", "son_fatura": "25.03.2024",
     "tedarikci": "Orman Ürünleri San. ve Tic.", "fatura_turleri": "Alış, İthalat"},
    {"stok_kodu": "STK-045", "stok_adi": "Köpük Conta (15mm)",
     "fatura_sayisi": 3, "toplam_tutar": "6.300,00 ₺",
     "ilk_fatura": "14.02.2024", "son_fatura": "02.04.2024",
     "tedarikci": "Kauçuk San. ve Tic.", "fatura_turleri": "Alış, Masraf"},
    {"stok_kodu": "STK-053", "stok_adi": "Rulman 6205-ZZ",
     "fatura_sayisi": 2, "toplam_tutar": "36.800,00 ₺",
     "ilk_fatura": "01.03.2024", "son_fatura": "05.04.2024",
     "tedarikci": "SKF Türkiye A.Ş.", "fatura_turleri": "İthalat"},
]

DEMO_TARAMA = [
    ("Tüm stok kodları listeleniyor",            1842),
    ("Reçeteler kontrol ediliyor",               1842),
    ("Mamül ağaçları kontrol ediliyor",          1842),
    ("Reçete/ağaç dışı stoklar filtreleniyor",    318),
    ("Alış faturaları kontrol ediliyor",          318),
    ("Kesişim kümesi hesaplanıyor",               87),
]

# ══════════════════════════════════════════════════════════════
#  DEMO VERİSİ — Araç 2 (Maliyet Hesaplama)
# ══════════════════════════════════════════════════════════════
DEMO_BOM = {
    "MA-001": {"ad": "Ahşap Masa Ünitesi", "birim": "ADET", "bilesenleri": [
        {"kod": "STK-037", "ad": "MDF Levha 18mm",     "miktar": 2,   "birim": "ADET"},
        {"kod": "STK-022", "ad": "Vida-Somun Seti M8", "miktar": 20,  "birim": "ADET"},
        {"kod": "STK-014", "ad": "Endüstriyel Boya",   "miktar": 0.5, "birim": "LT"},
    ]},
    "MA-002": {"ad": "Metal Raf Sistemi", "birim": "ADET", "bilesenleri": [
        {"kod": "STK-001", "ad": "Çelik Profil 40x40", "miktar": 6,  "birim": "KG"},
        {"kod": "STK-022", "ad": "Vida-Somun Seti M8", "miktar": 12, "birim": "ADET"},
        {"kod": "STK-045", "ad": "Köpük Conta 15mm",   "miktar": 2,  "birim": "ADET"},
    ]},
    "MA-003": {"ad": "Sandalye Grubu", "birim": "ADET", "bilesenleri": [
        {"kod": "STK-037", "ad": "MDF Levha 18mm",     "miktar": 1,   "birim": "ADET"},
        {"kod": "STK-001", "ad": "Çelik Profil 40x40", "miktar": 3,   "birim": "KG"},
        {"kod": "STK-014", "ad": "Endüstriyel Boya",   "miktar": 0.3, "birim": "LT"},
        {"kod": "STK-022", "ad": "Vida-Somun Seti M8", "miktar": 8,   "birim": "ADET"},
    ]},
    "MA-004": {"ad": "Dolap Modülü", "birim": "ADET", "bilesenleri": [
        {"kod": "STK-037", "ad": "MDF Levha 18mm",     "miktar": 4,  "birim": "ADET"},
        {"kod": "STK-022", "ad": "Vida-Somun Seti M8", "miktar": 32, "birim": "ADET"},
        {"kod": "STK-053", "ad": "Rulman 6205-ZZ",     "miktar": 2,  "birim": "ADET"},
        {"kod": "STK-045", "ad": "Köpük Conta 15mm",   "miktar": 4,  "birim": "ADET"},
    ]},
    "MA-005": {"ad": "Montaj Aparatı", "birim": "ADET", "bilesenleri": [
        {"kod": "STK-001", "ad": "Çelik Profil 40x40", "miktar": 2, "birim": "KG"},
        {"kod": "STK-008", "ad": "NYY Kablo 3x2.5mm",  "miktar": 5, "birim": "MT"},
        {"kod": "STK-053", "ad": "Rulman 6205-ZZ",     "miktar": 1, "birim": "ADET"},
    ]},
}

# Fatura geçmişi: tarih (YYYY-MM-DD), birim_fiyat, miktar
DEMO_FIYATLAR = {
    "STK-001": [{"tarih": "2024-01-10", "birim_fiyat": 23.50, "miktar": 500},
                {"tarih": "2024-02-08", "birim_fiyat": 24.80, "miktar": 300},
                {"tarih": "2024-03-15", "birim_fiyat": 26.20, "miktar": 800}],
    "STK-008": [{"tarih": "2024-01-28", "birim_fiyat": 22.50, "miktar": 300}],
    "STK-014": [{"tarih": "2024-02-05", "birim_fiyat": 165.00, "miktar": 60},
                {"tarih": "2024-03-18", "birim_fiyat": 178.00, "miktar": 60}],
    "STK-022": [{"tarih": "2024-01-20", "birim_fiyat": 3.10, "miktar": 1000},
                {"tarih": "2024-03-22", "birim_fiyat": 3.40, "miktar": 1000}],
    "STK-037": [{"tarih": "2024-01-05", "birim_fiyat": 280.00, "miktar": 40},
                {"tarih": "2024-02-20", "birim_fiyat": 295.00, "miktar": 40},
                {"tarih": "2024-03-25", "birim_fiyat": 310.00, "miktar": 80}],
    "STK-045": [{"tarih": "2024-01-15", "birim_fiyat": 2.05, "miktar": 500},
                {"tarih": "2024-03-02", "birim_fiyat": 2.15, "miktar": 500}],
    "STK-053": [{"tarih": "2024-03-01", "birim_fiyat": 92.00, "miktar": 100},
                {"tarih": "2024-04-05", "birim_fiyat": 98.00, "miktar": 100}],
}

# ══════════════════════════════════════════════════════════════
#  MALİYET HESAPLAMA FONKSİYONLARI
# ══════════════════════════════════════════════════════════════
def birim_maliyet(stok_kodu, metod, bas, bit):
    """bas/bit: 'YYYY-MM-DD' string"""
    fiyatlar = DEMO_FIYATLAR.get(stok_kodu, [])
    filtreli = [f for f in fiyatlar if bas <= f["tarih"] <= bit]
    if not filtreli:
        return 0.0
    if metod == "WA":
        top_t = sum(f["birim_fiyat"] * f["miktar"] for f in filtreli)
        top_m = sum(f["miktar"] for f in filtreli)
        return top_t / top_m if top_m else 0.0
    elif metod == "FIFO":
        return min(filtreli, key=lambda x: x["tarih"])["birim_fiyat"]
    elif metod == "LIFO":
        return max(filtreli, key=lambda x: x["tarih"])["birim_fiyat"]
    return 0.0


def mamul_maliyet_hesapla(mamul_kodu, metod, bas, bit):
    mamul = DEMO_BOM.get(mamul_kodu)
    if not mamul:
        return []
    satirlar = []
    toplam = 0.0
    for b in mamul["bilesenleri"]:
        bm = birim_maliyet(b["kod"], metod, bas, bit)
        satir_top = b["miktar"] * bm
        toplam += satir_top
        satirlar.append({
            "tip": "BİLEŞEN",
            "bil_kod": b["kod"], "bil_ad": b["ad"],
            "bom_miktar": b["miktar"], "birim": b["birim"],
            "birim_mal": bm, "satir_top": satir_top,
        })
    return satirlar, toplam


# ══════════════════════════════════════════════════════════════
#  GENEL STİL
# ══════════════════════════════════════════════════════════════
STIL = """
QMainWindow,QWidget{font-family:'Segoe UI';font-size:12px;background:#f0f2f5;}
QGroupBox{border:1.5px solid #c5cae9;border-radius:8px;
    margin-top:8px;padding:10px 8px 8px 8px;background:white;}
QGroupBox::title{subcontrol-origin:margin;left:10px;padding:0 6px;
    color:#3949ab;font-weight:bold;font-size:12px;}
QLineEdit{border:1.5px solid #9fa8da;border-radius:6px;
    padding:6px 10px;background:white;font-size:12px;}
QLineEdit:focus{border-color:#3f51b5;}
QDoubleSpinBox{border:1.5px solid #9fa8da;border-radius:6px;
    padding:4px 8px;background:white;font-size:12px;}
QDoubleSpinBox:focus{border-color:#3f51b5;}
QDateEdit{border:1.5px solid #9fa8da;border-radius:6px;
    padding:5px 8px;background:white;font-size:12px;}
QDateEdit:focus{border-color:#3f51b5;}
QPushButton{border-radius:6px;padding:8px 20px;
    font-size:12px;font-weight:bold;border:none;}
QListWidget{border:1px solid #e8eaf6;border-radius:6px;
    background:white;font-size:12px;}
QListWidget::item{padding:7px 10px;border-bottom:1px solid #f3f3f3;}
QListWidget::item:selected{background:#e8eaf6;color:#1a237e;}
QListWidget::item:hover{background:#f3f4ff;}
QProgressBar{border:none;border-radius:5px;background:#e8eaf6;}
QProgressBar::chunk{border-radius:5px;
    background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #3f51b5,stop:1 #5c6bc0);}
QCheckBox{spacing:8px;}
QCheckBox::indicator{width:16px;height:16px;border-radius:3px;
    border:2px solid #9fa8da;background:white;}
QCheckBox::indicator:checked{background:#3f51b5;border-color:#3f51b5;}
QRadioButton{spacing:8px;}
QRadioButton::indicator{width:15px;height:15px;border-radius:7px;
    border:2px solid #9fa8da;background:white;}
QRadioButton::indicator:checked{background:#3f51b5;border-color:#3f51b5;}
QScrollArea{border:none;background:transparent;}
"""


# ══════════════════════════════════════════════════════════════
#  YARDIMCI WIDGET FONKSİYONLARI
# ══════════════════════════════════════════════════════════════
def etiket(txt, renk="#7986cb", bold=True, size=12):
    w = QLabel(txt)
    s = f"color:{renk};"
    if bold: s += "font-weight:bold;"
    s += f"font-size:{size}px;"
    w.setStyleSheet(s)
    return w

def deger(txt="—", renk="#212121"):
    w = QLabel(txt)
    w.setWordWrap(True)
    w.setStyleSheet(f"color:{renk};")
    return w

def buton(txt, bg, fg="white", min_w=None, h=None):
    b = QPushButton(txt)
    b.setStyleSheet(f"background:{bg};color:{fg};")
    if min_w: b.setMinimumWidth(min_w)
    if h:     b.setFixedHeight(h)
    return b

def ayrac():
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setStyleSheet("color:#e0e0e0;")
    return f


# ══════════════════════════════════════════════════════════════
#  SAYFA 0: ANA MENÜ
# ══════════════════════════════════════════════════════════════
class AnaMenüSayfasi(QWidget):
    arac1_signal = pyqtSignal()
    arac2_signal = pyqtSignal()

    def __init__(self):
        super().__init__()
        ana = QVBoxLayout(self)
        ana.setAlignment(Qt.AlignCenter)
        ana.setSpacing(20)
        ana.setContentsMargins(40, 30, 40, 30)

        baslik = QLabel("CEO ERP Araçları")
        baslik.setFont(QFont("Segoe UI", 20, QFont.Bold))
        baslik.setStyleSheet("color:#1a237e;")
        baslik.setAlignment(Qt.AlignCenter)

        alt = QLabel("Kullanmak istediğiniz aracı seçin")
        alt.setAlignment(Qt.AlignCenter)
        alt.setStyleSheet("color:#7986cb;font-size:13px;")

        kartlar = QHBoxLayout()
        kartlar.setSpacing(24)
        kartlar.addWidget(self._kart(
            "🔗", "Mamül Ağacı\nBağlantı Aracı",
            "Reçete veya mamül ağacında yer almayan\n"
            "stokları tespit et ve doğru mamüle bağla.\n\n"
            "Boşta görünen maliyetleri ortadan kaldır.",
            "#e8eaf6", "#3f51b5", self.arac1_signal,
        ))
        kartlar.addWidget(self._kart(
            "💰", "Maliyet\nHesaplama",
            "Tüm reçete ve mamül ağaçlarını tara.\n"
            "LIFO, FIFO veya Ağırlıklı Ortalama ile\n"
            "ürün bazında maliyet raporu oluştur.",
            "#e8f5e9", "#2e7d32", self.arac2_signal,
        ))

        ana.addStretch()
        ana.addWidget(baslik)
        ana.addWidget(alt)
        ana.addSpacing(10)
        ana.addLayout(kartlar)
        ana.addStretch()

    def _kart(self, ikon, baslik, aciklama, bg, renk, sinyal):
        kart = QFrame()
        kart.setFixedSize(320, 300)
        kart.setStyleSheet(
            f"QFrame{{background:{bg};border-radius:14px;"
            f"border:2px solid {renk}33;}}"
        )
        lay = QVBoxLayout(kart)
        lay.setSpacing(12)
        lay.setContentsMargins(24, 24, 24, 20)

        ikon_lbl = QLabel(ikon)
        ikon_lbl.setStyleSheet(f"font-size:36px;background:transparent;border:none;")
        ikon_lbl.setAlignment(Qt.AlignCenter)

        bas_lbl = QLabel(baslik)
        bas_lbl.setFont(QFont("Segoe UI", 13, QFont.Bold))
        bas_lbl.setStyleSheet(f"color:{renk};background:transparent;border:none;")
        bas_lbl.setAlignment(Qt.AlignCenter)

        ac_lbl = QLabel(aciklama)
        ac_lbl.setStyleSheet("color:#444;font-size:11px;background:transparent;border:none;")
        ac_lbl.setWordWrap(True)
        ac_lbl.setAlignment(Qt.AlignCenter)

        btn_w = QPushButton("Başlat  →")
        btn_w.setStyleSheet(
            f"background:{renk};color:white;border-radius:8px;"
            f"padding:10px;font-weight:bold;font-size:13px;")
        btn_w.clicked.connect(sinyal.emit)

        lay.addWidget(ikon_lbl)
        lay.addWidget(bas_lbl)
        lay.addWidget(ac_lbl, stretch=1)
        lay.addWidget(btn_w)
        return kart


# ══════════════════════════════════════════════════════════════
#  SAYFA 1: BAĞLANTI (Araç 1)
# ══════════════════════════════════════════════════════════════
class BaglantiSayfasi(QWidget):
    devam = pyqtSignal()
    geri  = pyqtSignal()

    def __init__(self):
        super().__init__()
        ana = QVBoxLayout(self)
        ana.setAlignment(Qt.AlignCenter)
        ana.setSpacing(14)

        db_grup = QGroupBox("Veritabanı Bağlantısı")
        db_grup.setMaximumWidth(460)
        grid = QGridLayout(db_grup)
        grid.setSpacing(10)
        self.sunucu    = QLineEdit("localhost\\SQLEXPRESS")
        self.db_adi    = QLineEdit("CEO_ERP")
        self.kullanici = QLineEdit("sa")
        self.sifre     = QLineEdit()
        self.sifre.setEchoMode(QLineEdit.Password)
        self.sifre.setPlaceholderText("••••••••")
        for i, (k, v) in enumerate([("Sunucu:", self.sunucu), ("Veritabanı:", self.db_adi),
                                     ("Kullanıcı:", self.kullanici), ("Şifre:", self.sifre)]):
            grid.addWidget(etiket(k), i, 0)
            grid.addWidget(v, i, 1)

        tur_grup = QGroupBox("Fatura Türleri (alış faturası kontrolü)")
        tur_grup.setMaximumWidth(460)
        tl = QHBoxLayout(tur_grup)
        self.cb_alis    = QCheckBox("Alış");    self.cb_alis.setChecked(True)
        self.cb_masraf  = QCheckBox("Masraf");  self.cb_masraf.setChecked(True)
        self.cb_hizmet  = QCheckBox("Hizmet");  self.cb_hizmet.setChecked(True)
        self.cb_ithalat = QCheckBox("İthalat"); self.cb_ithalat.setChecked(True)
        for cb in [self.cb_alis, self.cb_masraf, self.cb_hizmet, self.cb_ithalat]:
            tl.addWidget(cb)

        baglan_btn = buton("  Bağlan ve Taramayı Başlat  ", "#3f51b5", h=42, min_w=230)
        baglan_btn.clicked.connect(self.devam.emit)
        demo_btn = buton("Demo Modunda Çalıştır", "#eceff1", "#37474f")
        demo_btn.clicked.connect(self.devam.emit)
        geri_btn = buton("← Ana Menü", "#eceff1", "#37474f")
        geri_btn.clicked.connect(self.geri.emit)

        ana.addStretch()
        ana.addWidget(etiket("Mamül Ağacı Bağlantı Aracı", "#1a237e", size=15),
                      alignment=Qt.AlignCenter)
        ana.addSpacing(4)
        ana.addWidget(db_grup,    alignment=Qt.AlignCenter)
        ana.addWidget(tur_grup,   alignment=Qt.AlignCenter)
        ana.addSpacing(6)
        ana.addWidget(baglan_btn, alignment=Qt.AlignCenter)
        ana.addWidget(demo_btn,   alignment=Qt.AlignCenter)
        ana.addSpacing(10)
        ana.addWidget(geri_btn,   alignment=Qt.AlignCenter)
        ana.addStretch()


# ══════════════════════════════════════════════════════════════
#  SAYFA 2: TARAMA (Araç 1)
# ══════════════════════════════════════════════════════════════
class TaramaThread(QThread):
    adim  = pyqtSignal(str, int, int)
    bitti = pyqtSignal()
    def run(self):
        import time
        for mesaj, toplam in DEMO_TARAMA:
            adim = max(1, toplam // 50)
            for i in range(0, toplam + 1, adim):
                self.adim.emit(mesaj, min(i, toplam), toplam)
                time.sleep(0.016)
            self.adim.emit(mesaj, toplam, toplam)
            time.sleep(0.15)
        self.bitti.emit()

class TaramaSayfasi(QWidget):
    bitti = pyqtSignal()
    def __init__(self):
        super().__init__()
        ana = QVBoxLayout(self)
        ana.setAlignment(Qt.AlignCenter)
        ana.setSpacing(14)

        self.adim_lbl   = QLabel("Başlatılıyor...")
        self.adim_lbl.setAlignment(Qt.AlignCenter)
        self.adim_lbl.setStyleSheet("color:#5c6bc0;font-size:13px;")

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setFixedHeight(16)
        self.bar.setMaximumWidth(540)
        self.bar.setTextVisible(False)

        self.sayac = QLabel("")
        self.sayac.setAlignment(Qt.AlignCenter)
        self.sayac.setStyleSheet("color:#9e9e9e;font-size:11px;")

        self.ozet_grup = QGroupBox("Tarama Sonucu")
        self.ozet_grup.setMaximumWidth(540)
        self.ozet_grup.hide()
        ol = QVBoxLayout(self.ozet_grup)
        self.ozet_lbl = QLabel()
        self.ozet_lbl.setStyleSheet("font-size:13px;color:#212121;line-height:1.8;")
        ol.addWidget(self.ozet_lbl)

        self.devam_btn = buton("Eşleştirmeye Geç  →", "#3f51b5", min_w=210, h=42)
        self.devam_btn.hide()
        self.devam_btn.clicked.connect(self.bitti.emit)

        ana.addStretch()
        ana.addWidget(etiket("Stok Analizi Yapılıyor...", "#1a237e", size=15),
                      alignment=Qt.AlignCenter)
        ana.addWidget(self.adim_lbl)
        ana.addWidget(self.bar,        alignment=Qt.AlignCenter)
        ana.addWidget(self.sayac)
        ana.addSpacing(8)
        ana.addWidget(self.ozet_grup,  alignment=Qt.AlignCenter)
        ana.addWidget(self.devam_btn,  alignment=Qt.AlignCenter)
        ana.addStretch()

    def baslat(self):
        self.ozet_grup.hide()
        self.devam_btn.hide()
        self.bar.setValue(0)
        self.thread = TaramaThread()
        self.thread.adim.connect(self._guncelle)
        self.thread.bitti.connect(self._bitti)
        self.thread.start()

    def _guncelle(self, mesaj, mevcut, toplam):
        self.adim_lbl.setText(mesaj)
        self.bar.setValue(int(mevcut / toplam * 100) if toplam else 0)
        self.sayac.setText(f"{mevcut:,} / {toplam:,}")

    def _bitti(self):
        self.bar.setValue(100)
        self.adim_lbl.setText("Analiz tamamlandı.")
        self.sayac.setText("")
        self.ozet_lbl.setText(
            f"<b>Toplam stok kodu:</b>  {DEMO_TARAMA[0][1]:,}<br>"
            f"<b>Reçete/mamül ağacında olmayan:</b>  {DEMO_TARAMA[3][1]:,}<br>"
            f"<b>Bunlardan faturası olan (kesişim):</b>  "
            f"<span style='color:#c62828;font-weight:bold;'>{len(DEMO_STOKLAR)}</span>"
        )
        self.ozet_grup.show()
        self.devam_btn.show()


# ══════════════════════════════════════════════════════════════
#  SAYFA 3: EŞLEŞTİRME (Araç 1)
# ══════════════════════════════════════════════════════════════
class EslestirmeSayfasi(QWidget):
    rapor_signal = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.stoklar = DEMO_STOKLAR
        self.idx = 0
        self.sonuclar = []
        self.secilen_mamul = None
        self._kur()

    def _kur(self):
        ana = QVBoxLayout(self)
        ana.setSpacing(10)
        ana.setContentsMargins(16, 12, 16, 10)
        ana.addWidget(self._ilerleme())

        sp = QSplitter(Qt.Horizontal)
        sp.addWidget(self._stok_paneli())
        sp.addWidget(self._sag_panel())
        sp.setSizes([440, 460])
        sp.setStyleSheet("QSplitter::handle{background:#e0e0e0;width:2px;}")
        ana.addWidget(sp, stretch=1)
        ana.addWidget(self._buton_bar())

    def _ilerleme(self):
        frame = QFrame()
        frame.setStyleSheet("background:white;border-radius:8px;border:1.5px solid #c5cae9;")
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(14, 8, 14, 8)
        self.ilerleme_lbl = QLabel("Stok 1 / ?")
        self.ilerleme_lbl.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.ilerleme_lbl.setStyleSheet("color:#3949ab;border:none;background:transparent;")
        self.prog = QProgressBar()
        self.prog.setRange(0, len(self.stoklar))
        self.prog.setValue(0)
        self.prog.setTextVisible(False)
        self.prog.setFixedHeight(10)
        self.kalan_lbl = QLabel("")
        self.kalan_lbl.setStyleSheet("color:#7986cb;border:none;background:transparent;font-size:11px;")
        self.kalan_lbl.setAlignment(Qt.AlignRight)
        lay.addWidget(self.ilerleme_lbl)
        lay.addWidget(self.prog, stretch=1)
        lay.addWidget(self.kalan_lbl)
        return frame

    def _stok_paneli(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 6, 0)
        lay.setSpacing(10)
        detay = QGroupBox("Stok Detayları")
        grid = QGridLayout(detay)
        grid.setVerticalSpacing(10)
        grid.setHorizontalSpacing(10)
        alanlar = [
            ("Stok Kodu:",     "stok_kodu",     "#c62828"),
            ("Stok Adı:",      "stok_adi",      "#212121"),
            ("Fatura Türleri:","fatura_turleri","#212121"),
            ("Fatura Sayısı:", "fatura_sayisi", "#212121"),
            ("Toplam Tutar:",  "toplam_tutar",  "#1b5e20"),
            ("İlk Fatura:",    "ilk_fatura",    "#212121"),
            ("Son Fatura:",    "son_fatura",    "#212121"),
            ("Tedarikçi:",     "tedarikci",     "#212121"),
        ]
        self.deger_lbls = {}
        for i, (et, key, renk) in enumerate(alanlar):
            e = etiket(et)
            e.setFixedWidth(105)
            v = QLabel("—")
            v.setWordWrap(True)
            v.setStyleSheet(f"color:{renk};")
            if key == "stok_kodu": v.setFont(QFont("Segoe UI", 13, QFont.Bold))
            if key == "toplam_tutar": v.setFont(QFont("Segoe UI", 12, QFont.Bold))
            grid.addWidget(e, i, 0, Qt.AlignTop)
            grid.addWidget(v, i, 1)
            self.deger_lbls[key] = v

        uyari = QLabel("⚠  Bu stok herhangi bir reçete veya mamül ağacında\n"
                       "   yer almıyor. Sağ taraftan atama yapınız.")
        uyari.setStyleSheet("color:#e65100;font-size:11px;background:#fff3e0;"
                            "border-radius:6px;padding:8px;border:1px solid #ffcc80;")
        uyari.setWordWrap(True)
        lay.addWidget(detay)
        lay.addWidget(uyari)
        lay.addStretch()
        return w

    def _sag_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(6, 0, 0, 0)
        lay.setSpacing(10)

        mamul_grup = QGroupBox("Hangi Mamül Ağacına Eklenecek?")
        ml = QVBoxLayout(mamul_grup)
        ml.setSpacing(8)
        self.arama = QLineEdit()
        self.arama.setPlaceholderText("Mamül kodu veya adı ile filtrele...")
        self.arama.textChanged.connect(self._filtrele)
        self.mamul_listesi = QListWidget()
        self.mamul_listesi.setSelectionMode(QAbstractItemView.SingleSelection)
        self.mamul_listesi.itemClicked.connect(self._mamul_secildi)
        self._mamul_doldur()
        self.secilen_lbl = QLabel("Seçim yapılmadı")
        self.secilen_lbl.setStyleSheet("color:#9e9e9e;font-style:italic;padding:7px;"
                                       "background:#f5f5f5;border-radius:5px;font-size:12px;")
        self.secilen_lbl.setWordWrap(True)
        ml.addWidget(etiket("Ara:", bold=False, renk="#555"))
        ml.addWidget(self.arama)
        ml.addWidget(self.mamul_listesi, stretch=1)
        ml.addWidget(self.secilen_lbl)

        gecmis_grup = QGroupBox("İşlem Geçmişi")
        gl = QVBoxLayout(gecmis_grup)
        self.gecmis_liste = QListWidget()
        self.gecmis_liste.setMaximumHeight(150)
        gl.addWidget(self.gecmis_liste)

        lay.addWidget(mamul_grup, stretch=1)
        lay.addWidget(gecmis_grup)
        return w

    def _mamul_doldur(self, filtre=""):
        self.mamul_listesi.clear()
        ft = filtre.lower().strip()
        for kod, ad in DEMO_MAMUL_AGACLARI:
            if ft and ft not in kod.lower() and ft not in ad.lower():
                continue
            item = QListWidgetItem(f"  {kod}   {ad}")
            item.setData(Qt.UserRole, (kod, ad))
            self.mamul_listesi.addItem(item)

    def _filtrele(self, txt):
        self.secilen_mamul = None
        self.secilen_lbl.setText("Seçim yapılmadı")
        self.secilen_lbl.setStyleSheet("color:#9e9e9e;font-style:italic;padding:7px;"
                                       "background:#f5f5f5;border-radius:5px;font-size:12px;")
        self._mamul_doldur(txt)

    def _mamul_secildi(self, item):
        kod, ad = item.data(Qt.UserRole)
        self.secilen_mamul = (kod, ad)
        self.secilen_lbl.setText(f"✓  {kod} — {ad}")
        self.secilen_lbl.setStyleSheet("color:#2e7d32;font-weight:bold;padding:7px;"
                                       "background:#e8f5e9;border-radius:5px;font-size:12px;")

    def _buton_bar(self):
        frame = QFrame()
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(0, 4, 0, 0)
        lay.setSpacing(10)
        self.geri_btn = QPushButton("← Geri")
        self.geri_btn.setStyleSheet("background:#eceff1;color:#37474f;border:1px solid #cfd8dc;"
                                    "border-radius:6px;padding:8px 20px;font-weight:bold;")
        self.geri_btn.clicked.connect(self._geri)
        atla_btn = QPushButton("⊘  Şimdilik Atla")
        atla_btn.setStyleSheet("background:#fff3e0;color:#e65100;border:1px solid #ffcc80;"
                               "border-radius:6px;padding:8px 20px;font-weight:bold;")
        atla_btn.clicked.connect(self._atla)
        kaydet_btn = QPushButton("✓  Mamüle Bağla ve İleri")
        kaydet_btn.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #3f51b5,stop:1 #5c6bc0);"
            "color:white;border-radius:6px;padding:9px 28px;font-weight:bold;font-size:13px;")
        kaydet_btn.clicked.connect(self._kaydet)
        lay.addWidget(self.geri_btn)
        lay.addStretch()
        lay.addWidget(atla_btn)
        lay.addWidget(kaydet_btn)
        return frame

    def baslat(self):
        self.idx = 0
        self.sonuclar = []
        self.secilen_mamul = None
        self.gecmis_liste.clear()
        self._goster()

    def _goster(self):
        n = len(self.stoklar)
        if self.idx >= n:
            self.rapor_signal.emit(self.sonuclar)
            return
        s = self.stoklar[self.idx]
        for key, w in self.deger_lbls.items():
            w.setText(str(s.get(key, "—")))
        self.arama.clear()
        self.secilen_mamul = None
        self.secilen_lbl.setText("Seçim yapılmadı")
        self.secilen_lbl.setStyleSheet("color:#9e9e9e;font-style:italic;padding:7px;"
                                       "background:#f5f5f5;border-radius:5px;font-size:12px;")
        self._mamul_doldur()
        self.ilerleme_lbl.setText(f"Stok {self.idx + 1} / {n}")
        self.prog.setValue(self.idx + 1)
        self.kalan_lbl.setText(f"{n - self.idx} stok bekliyor")
        self.geri_btn.setEnabled(self.idx > 0)

    def _kaydet(self):
        if not self.secilen_mamul:
            QMessageBox.warning(self, "Uyarı", "Lütfen sağ listeden bir mamül ağacı seçin.")
            return
        kod, ad = self.secilen_mamul
        s = self.stoklar[self.idx]
        self.sonuclar.append({**s, "mamul_kodu": kod, "mamul_adi": ad, "islem": "Bağlandı"})
        item = QListWidgetItem(f"✓  {s['stok_kodu']}  →  {kod} ({ad})")
        item.setForeground(QColor("#2e7d32"))
        self.gecmis_liste.insertItem(0, item)
        self.idx += 1
        self._goster()

    def _atla(self):
        s = self.stoklar[self.idx]
        self.sonuclar.append({**s, "mamul_kodu": "—", "mamul_adi": "—", "islem": "Atlandı"})
        item = QListWidgetItem(f"⊘  {s['stok_kodu']}  →  Atlandı")
        item.setForeground(QColor("#e65100"))
        self.gecmis_liste.insertItem(0, item)
        self.idx += 1
        self._goster()

    def _geri(self):
        if self.idx > 0:
            self.idx -= 1
            if self.sonuclar: self.sonuclar.pop()
            if self.gecmis_liste.count() > 0: self.gecmis_liste.takeItem(0)
            self._goster()


# ══════════════════════════════════════════════════════════════
#  SAYFA 4: RAPOR (Araç 1)
# ══════════════════════════════════════════════════════════════
class RaporSayfasi(QWidget):
    def __init__(self):
        super().__init__()
        self.sonuclar = []
        self._kur()

    def _kur(self):
        ana = QVBoxLayout(self)
        ana.setAlignment(Qt.AlignCenter)
        ana.setSpacing(14)

        self.ozet_grup = QGroupBox("Özet")
        self.ozet_grup.setMaximumWidth(540)
        ol = QVBoxLayout(self.ozet_grup)
        self.ozet_lbl = QLabel()
        self.ozet_lbl.setStyleSheet("font-size:13px;color:#212121;line-height:1.8;")
        ol.addWidget(self.ozet_lbl)

        excel_grup = QGroupBox("Excel Raporu (.xlsx)")
        excel_grup.setMaximumWidth(540)
        el = QVBoxLayout(excel_grup)
        bilgi = QLabel("Tek sayfa: stok kodu, fatura türleri, toplam tutar, tedarikçi, atanan mamül.\n"
                       "Bağlananlar yeşil — atlandılar sarı renkte gösterilir.")
        bilgi.setStyleSheet("color:#555;font-size:11px;")
        self.excel_btn = QPushButton("📥  Excel Olarak Kaydet (.xlsx)")
        self.excel_btn.setStyleSheet("background:#2e7d32;color:white;border-radius:6px;"
                                     "padding:10px 30px;font-weight:bold;font-size:13px;")
        self.excel_btn.clicked.connect(self._excel_kaydet)
        el.addWidget(bilgi)
        el.addSpacing(6)
        el.addWidget(self.excel_btn, alignment=Qt.AlignCenter)

        geri_btn = QPushButton("← Ana Menü")
        geri_btn.setStyleSheet("background:#eceff1;color:#37474f;border:1px solid #cfd8dc;"
                               "border-radius:6px;padding:8px 20px;font-weight:bold;")
        geri_btn.clicked.connect(lambda: self.window().sayfa_gec(0))

        ana.addStretch()
        ana.addWidget(etiket("İşlem Tamamlandı ✓", "#1a237e", size=15), alignment=Qt.AlignCenter)
        ana.addWidget(self.ozet_grup,  alignment=Qt.AlignCenter)
        ana.addWidget(excel_grup,      alignment=Qt.AlignCenter)
        ana.addWidget(geri_btn,        alignment=Qt.AlignCenter)
        ana.addStretch()

    def goster(self, sonuclar):
        self.sonuclar = sonuclar
        baglandi = sum(1 for s in sonuclar if s["islem"] == "Bağlandı")
        atlandi  = sum(1 for s in sonuclar if s["islem"] == "Atlandı")
        self.ozet_lbl.setText(
            f"<b>İşlenen stok:</b>  {len(sonuclar)}<br>"
            f"<b>Mamül ağacına bağlanan:</b>  <span style='color:#2e7d32;font-weight:bold;'>{baglandi}</span><br>"
            f"<b>Atlanan / bekleyen:</b>  <span style='color:#e65100;'>{atlandi}</span>"
        )

    def _excel_kaydet(self):
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"ceo_erp_mamul_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel Dosyası (*.xlsx)")
        if not dosya: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mamül Ağacı Raporu"
            hf = PatternFill("solid", fgColor="3F51B5")
            hy = XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11)
            kn = Border(left=Side(style="thin",color="CCCCCC"), right=Side(style="thin",color="CCCCCC"),
                        top=Side(style="thin",color="CCCCCC"),  bottom=Side(style="thin",color="CCCCCC"))
            sutunlar = [("Stok Kodu",14),("Stok Adı",32),("Fatura Türleri",20),
                        ("Fatura Sayısı",14),("Toplam Tutar",18),("İlk Fatura",14),
                        ("Son Fatura",14),("Tedarikçi",30),("Mamül Kodu",14),
                        ("Mamül Adı",30),("İşlem",13)]
            ws.row_dimensions[1].height = 30
            for col,(bas,gen) in enumerate(sutunlar,1):
                h = ws.cell(row=1,column=col,value=bas)
                h.fill=hf; h.font=hy; h.border=kn
                h.alignment=Alignment(horizontal="center",vertical="center")
                ws.column_dimensions[get_column_letter(col)].width=gen
            bf = PatternFill("solid",fgColor="E8F5E9")
            af = PatternFill("solid",fgColor="FFF8E1")
            vf = XFont(name="Segoe UI",size=10)
            alan = ["stok_kodu","stok_adi","fatura_turleri","fatura_sayisi","toplam_tutar",
                    "ilk_fatura","son_fatura","tedarikci","mamul_kodu","mamul_adi","islem"]
            for satir,s in enumerate(self.sonuclar,2):
                df = bf if s["islem"]=="Bağlandı" else af
                ws.row_dimensions[satir].height=18
                for col,key in enumerate(alan,1):
                    h=ws.cell(row=satir,column=col,value=str(s.get(key,"")))
                    h.fill=df; h.font=vf; h.border=kn
                    h.alignment=Alignment(vertical="center")
            ws.auto_filter.ref=f"A1:{get_column_letter(len(sutunlar))}1"
            ws.freeze_panes="A2"
            wb.save(dosya)
            QMessageBox.information(self,"Başarılı",f"Dosya kaydedildi:\n{dosya}")
        except Exception as e:
            QMessageBox.critical(self,"Hata",f"Excel kaydedilemedi:\n{e}")


# ══════════════════════════════════════════════════════════════
#  SAYFA 5: MALİYET HESAPLAMA (Araç 2)
# ══════════════════════════════════════════════════════════════
class MaliyetSayfasi(QWidget):
    def __init__(self):
        super().__init__()
        self._mamul_satirlari = {}   # kod → (checkbox, QDoubleSpinBox)
        self._kur()

    def _kur(self):
        ana = QHBoxLayout(self)
        ana.setContentsMargins(16, 12, 16, 12)
        ana.setSpacing(16)
        ana.addWidget(self._sol_panel(), stretch=4)
        ana.addWidget(self._sag_panel(), stretch=5)

    # ── Sol: Parametreler ────────────────────────────────────────────────────
    def _sol_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        # Başlık
        bas = QLabel("Maliyet Hesaplama")
        bas.setFont(QFont("Segoe UI", 13, QFont.Bold))
        bas.setStyleSheet("color:#1a237e;")
        lay.addWidget(bas)

        # Veritabanı
        db_grup = QGroupBox("Veritabanı Bağlantısı")
        dg = QGridLayout(db_grup)
        dg.setSpacing(8)
        self.m_sunucu    = QLineEdit("localhost\\SQLEXPRESS")
        self.m_db        = QLineEdit("CEO_ERP")
        self.m_kullanici = QLineEdit("sa")
        self.m_sifre     = QLineEdit(); self.m_sifre.setEchoMode(QLineEdit.Password)
        self.m_sifre.setPlaceholderText("••••••••")
        for i,(k,v) in enumerate([("Sunucu:",self.m_sunucu),("Veritabanı:",self.m_db),
                                   ("Kullanıcı:",self.m_kullanici),("Şifre:",self.m_sifre)]):
            dg.addWidget(etiket(k), i, 0)
            dg.addWidget(v, i, 1)
        lay.addWidget(db_grup)

        # Tarih aralığı
        tarih_grup = QGroupBox("Tarih Aralığı")
        tg = QGridLayout(tarih_grup)
        tg.setSpacing(8)
        self.tarih_bas = QDateEdit(QDate(2024, 1, 1))
        self.tarih_bas.setCalendarPopup(True)
        self.tarih_bas.setDisplayFormat("dd.MM.yyyy")
        self.tarih_bit = QDateEdit(QDate(2024, 3, 31))
        self.tarih_bit.setCalendarPopup(True)
        self.tarih_bit.setDisplayFormat("dd.MM.yyyy")
        tg.addWidget(etiket("Başlangıç:"), 0, 0)
        tg.addWidget(self.tarih_bas, 0, 1)
        tg.addWidget(etiket("Bitiş:"),     1, 0)
        tg.addWidget(self.tarih_bit, 1, 1)
        lay.addWidget(tarih_grup)

        # Maliyet yöntemi
        yontem_grup = QGroupBox("Maliyet Yöntemi")
        yg = QVBoxLayout(yontem_grup)
        self.metod_grup = QButtonGroup(self)
        for metin, deger_m in [("Ağırlıklı Ortalama", "WA"),
                                ("FIFO (İlk Giren İlk Çıkar)", "FIFO"),
                                ("LIFO (Son Giren İlk Çıkar)", "LIFO")]:
            rb = QRadioButton(metin)
            rb.setProperty("metod", deger_m)
            if deger_m == "WA": rb.setChecked(True)
            self.metod_grup.addButton(rb)
            yg.addWidget(rb)
        lay.addWidget(yontem_grup)

        lay.addStretch()

        # Alt butonlar
        alt = QHBoxLayout()
        geri_btn = QPushButton("← Ana Menü")
        geri_btn.setStyleSheet("background:#eceff1;color:#37474f;border:1px solid #cfd8dc;"
                               "border-radius:6px;padding:7px 16px;font-weight:bold;")
        geri_btn.clicked.connect(lambda: self.window().sayfa_gec(0))

        self.hesapla_btn = QPushButton("📊  Hesapla ve Excel'e Aktar")
        self.hesapla_btn.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2e7d32,stop:1 #43a047);"
            "color:white;border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;")
        self.hesapla_btn.clicked.connect(self._hesapla)

        alt.addWidget(geri_btn)
        alt.addStretch()
        alt.addWidget(self.hesapla_btn)
        lay.addLayout(alt)
        return w

    # ── Sağ: Mamül seçimi + işçilik ─────────────────────────────────────────
    def _sag_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        mamul_grup = QGroupBox("Mamüller ve İşçilik Tutarları")
        mg = QVBoxLayout(mamul_grup)

        # Tümünü seç / kaldır
        btn_satir = QHBoxLayout()
        tum_sec = QPushButton("Tümünü Seç")
        tum_sec.setStyleSheet("background:#e8eaf6;color:#3949ab;border-radius:5px;padding:5px 12px;font-size:11px;font-weight:bold;")
        tum_kaldir = QPushButton("Tümünü Kaldır")
        tum_kaldir.setStyleSheet("background:#fce4ec;color:#c62828;border-radius:5px;padding:5px 12px;font-size:11px;font-weight:bold;")
        tum_sec.clicked.connect(lambda: self._tum_sec(True))
        tum_kaldir.clicked.connect(lambda: self._tum_sec(False))
        btn_satir.addWidget(tum_sec)
        btn_satir.addWidget(tum_kaldir)
        btn_satir.addStretch()
        mg.addLayout(btn_satir)

        # Sütun başlıkları
        baslik_satir = QHBoxLayout()
        baslik_satir.addWidget(QLabel(""), stretch=1)   # checkbox + ad
        lbl_iscilik = QLabel("İşçilik (₺)")
        lbl_iscilik.setStyleSheet("color:#7986cb;font-weight:bold;font-size:11px;")
        lbl_iscilik.setFixedWidth(110)
        lbl_iscilik.setAlignment(Qt.AlignCenter)
        baslik_satir.addWidget(lbl_iscilik)
        mg.addLayout(baslik_satir)
        mg.addWidget(ayrac())

        # Scroll alanı — mamül listesi
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet("background:white;")
        self.mamul_layout = QVBoxLayout(scroll_widget)
        self.mamul_layout.setSpacing(4)
        self.mamul_layout.setContentsMargins(4, 4, 4, 4)
        self._mamul_satirlari = {}
        for kod, ad in DEMO_BOM.items():
            self._mamul_satiri_ekle(kod, ad["ad"])
        self.mamul_layout.addStretch()
        scroll.setWidget(scroll_widget)
        mg.addWidget(scroll, stretch=1)

        lay.addWidget(mamul_grup, stretch=1)

        # Durum etiketi
        self.durum_lbl = QLabel("")
        self.durum_lbl.setStyleSheet("color:#555;font-size:11px;padding:4px;")
        self.durum_lbl.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.durum_lbl)
        return w

    def _mamul_satiri_ekle(self, kod, ad):
        satir = QFrame()
        satir.setStyleSheet("QFrame{background:white;border-radius:5px;}"
                            "QFrame:hover{background:#f5f5f5;}")
        satir_lay = QHBoxLayout(satir)
        satir_lay.setContentsMargins(4, 4, 4, 4)
        satir_lay.setSpacing(8)

        cb = QCheckBox(f"{kod}  —  {ad}")
        cb.setChecked(True)
        cb.setStyleSheet("font-size:12px;color:#212121;")
        cb.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        spin = QDoubleSpinBox()
        spin.setRange(0, 9_999_999)
        spin.setDecimals(2)
        spin.setSuffix(" ₺")
        spin.setValue(0.00)
        spin.setFixedWidth(110)
        spin.setAlignment(Qt.AlignRight)

        satir_lay.addWidget(cb, stretch=1)
        satir_lay.addWidget(spin)
        self.mamul_layout.addWidget(satir)
        self._mamul_satirlari[kod] = (cb, spin)

    def _tum_sec(self, durum):
        for cb, _ in self._mamul_satirlari.values():
            cb.setChecked(durum)

    def _secili_metod(self):
        for btn_w in self.metod_grup.buttons():
            if btn_w.isChecked():
                return btn_w.property("metod")
        return "WA"

    # ── Hesaplama ve Excel ───────────────────────────────────────────────────
    def _hesapla(self):
        secili = [(kod, cb, spin) for kod, (cb, spin) in self._mamul_satirlari.items()
                  if cb.isChecked()]
        if not secili:
            QMessageBox.warning(self, "Uyarı", "Lütfen en az bir mamül seçin.")
            return

        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"ceo_erp_maliyet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel Dosyası (*.xlsx)")
        if not dosya: return

        metod = self._secili_metod()
        bas   = self.tarih_bas.date().toString("yyyy-MM-dd")
        bit   = self.tarih_bit.date().toString("yyyy-MM-dd")
        bas_g = self.tarih_bas.date().toString("dd.MM.yyyy")
        bit_g = self.tarih_bit.date().toString("dd.MM.yyyy")

        try:
            self._excel_olustur(dosya, secili, metod, bas, bit, bas_g, bit_g)
            self.durum_lbl.setText(f"✓  Rapor oluşturuldu: {dosya}")
            self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;padding:4px;")
            QMessageBox.information(self, "Başarılı", f"Maliyet raporu kaydedildi:\n{dosya}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel oluşturulamadı:\n{e}")

    def _excel_olustur(self, dosya, secili, metod, bas, bit, bas_g, bit_g):
        metod_ad = {"WA": "Ağırlıklı Ortalama", "FIFO": "FIFO", "LIFO": "LIFO"}[metod]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maliyet Raporu"

        # ── Stiller ──
        kn = Border(left=Side(style="thin",color="DDDDDD"),right=Side(style="thin",color="DDDDDD"),
                    top=Side(style="thin",color="DDDDDD"), bottom=Side(style="thin",color="DDDDDD"))
        orta = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sol  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        sag  = Alignment(horizontal="right",  vertical="center")

        fill = {
            "baslik":   PatternFill("solid", fgColor="1A237E"),
            "mamul":    PatternFill("solid", fgColor="3F51B5"),
            "birlesen": PatternFill("solid", fgColor="F5F5F5"),
            "iscilik":  PatternFill("solid", fgColor="FFF3E0"),
            "toplam":   PatternFill("solid", fgColor="E8F5E9"),
            "bilgi":    PatternFill("solid", fgColor="E8EAF6"),
        }
        font = {
            "baslik":  XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11),
            "mamul":   XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11),
            "birlesen":XFont(name="Segoe UI", size=10),
            "iscilik": XFont(italic=True, name="Segoe UI", size=10, color="E65100"),
            "toplam":  XFont(bold=True, name="Segoe UI", size=11, color="1B5E20"),
            "bilgi":   XFont(name="Segoe UI", size=10, color="555555"),
        }

        # ── Sütunlar ──
        # Tip | Mamül Kodu | Mamül Adı | Bileşen Kodu | Bileşen Adı | BOM Miktarı | Birim | Birim Maliyet | Satır Maliyeti | Mamül Toplam | +İşçilik | Genel Toplam | Metod
        sutunlar = [
            ("Tip",            10), ("Mamül Kodu",  14), ("Mamül Adı",       28),
            ("Bileşen Kodu",   14), ("Bileşen Adı", 30), ("BOM Miktarı",     13),
            ("Birim",          10), ("Birim Maliyet ₺", 18), ("Satır Maliyeti ₺", 18),
            ("Hammadde Toplamı ₺", 20), ("İşçilik ₺", 16), ("Genel Toplam ₺", 20),
        ]

        # ── Bilgi satırı (üst kısım) ──
        ws.merge_cells(f"A1:{get_column_letter(len(sutunlar))}1")
        h = ws.cell(row=1, column=1,
                    value=f"CEO ERP — Maliyet Raporu  |  Yöntem: {metod_ad}  |  "
                          f"Dönem: {bas_g} – {bit_g}  |  "
                          f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        h.fill = fill["bilgi"]
        h.font = font["bilgi"]
        h.alignment = sol
        h.border = kn
        ws.row_dimensions[1].height = 24

        # ── Başlık satırı ──
        ws.row_dimensions[2].height = 28
        for col, (bas_txt, gen) in enumerate(sutunlar, 1):
            h = ws.cell(row=2, column=col, value=bas_txt)
            h.fill = fill["baslik"]
            h.font = font["baslik"]
            h.alignment = orta
            h.border = kn
            ws.column_dimensions[get_column_letter(col)].width = gen

        satir = 3

        for mamul_kodu, cb, spin in secili:
            mamul = DEMO_BOM.get(mamul_kodu)
            if not mamul: continue

            iscilik_tutari = spin.value()
            bilesenleri, hammadde_top = mamul_maliyet_hesapla(mamul_kodu, metod, bas, bit)
            genel_top = hammadde_top + iscilik_tutari

            # ── Mamül başlık satırı ──
            ws.row_dimensions[satir].height = 22
            degerler = ["MAMÜL", mamul_kodu, mamul["ad"], "", "", "", "", "", "",
                        hammadde_top, iscilik_tutari, genel_top]
            for col, val in enumerate(degerler, 1):
                h = ws.cell(row=satir, column=col)
                if isinstance(val, float):
                    h.value = round(val, 2)
                    h.number_format = '#,##0.00 ₺'
                    h.alignment = sag
                else:
                    h.value = val
                    h.alignment = orta if col in (1, 2) else sol
                h.fill = fill["mamul"]
                h.font = font["mamul"]
                h.border = kn
            satir += 1

            # ── Bileşen satırları ──
            for b in bilesenleri:
                ws.row_dimensions[satir].height = 18
                degerler = [
                    "BİLEŞEN", mamul_kodu, mamul["ad"],
                    b["bil_kod"], b["bil_ad"],
                    b["bom_miktar"], b["birim"],
                    b["birim_mal"], b["satir_top"],
                    "", "", ""
                ]
                for col, val in enumerate(degerler, 1):
                    h = ws.cell(row=satir, column=col)
                    if isinstance(val, float):
                        h.value = round(val, 4)
                        h.number_format = '#,##0.00'
                        h.alignment = sag
                    else:
                        h.value = val
                        h.alignment = orta if col in (1,2,6,7) else sol
                    h.fill = fill["birlesen"]
                    h.font = font["birlesen"]
                    h.border = kn
                satir += 1

            # ── İşçilik satırı ──
            if iscilik_tutari > 0:
                ws.row_dimensions[satir].height = 18
                degerler = ["İŞÇİLİK", mamul_kodu, mamul["ad"],
                            "—", "Manuel İşçilik Tutarı", "", "", "", iscilik_tutari,
                            "", "", ""]
                for col, val in enumerate(degerler, 1):
                    h = ws.cell(row=satir, column=col)
                    if isinstance(val, float):
                        h.value = round(val, 2)
                        h.number_format = '#,##0.00 ₺'
                        h.alignment = sag
                    else:
                        h.value = val
                        h.alignment = orta if col in (1,2) else sol
                    h.fill = fill["iscilik"]
                    h.font = font["iscilik"]
                    h.border = kn
                satir += 1

            # ── Toplam satırı ──
            ws.row_dimensions[satir].height = 20
            degerler = ["TOPLAM", mamul_kodu, mamul["ad"], "", "", "", "", "", "",
                        hammadde_top, iscilik_tutari, genel_top]
            for col, val in enumerate(degerler, 1):
                h = ws.cell(row=satir, column=col)
                if isinstance(val, float):
                    h.value = round(val, 2)
                    h.number_format = '#,##0.00 ₺'
                    h.alignment = sag
                else:
                    h.value = val
                    h.alignment = orta if col in (1,2) else sol
                h.fill = fill["toplam"]
                h.font = font["toplam"]
                h.border = kn
            satir += 1

            # Boş ayraç satırı
            ws.row_dimensions[satir].height = 6
            satir += 1

        ws.auto_filter.ref = f"A2:{get_column_letter(len(sutunlar))}2"
        ws.freeze_panes = "A3"
        wb.save(dosya)


# ══════════════════════════════════════════════════════════════
#  ANA PENCERE
# ══════════════════════════════════════════════════════════════
ADIMLAR_ARAC1 = ["① Bağlantı", "② Tarama", "③ Eşleştirme", "④ Rapor"]

class AnaPencere(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CEO ERP — Araçlar")
        self.setMinimumSize(920, 680)
        self.resize(1020, 750)
        self.setStyleSheet(STIL)

        self.adim_bar = self._adim_bar_olustur()
        self.stack = QStackedWidget()

        self.s0_menu       = AnaMenüSayfasi()    # 0
        self.s1_baglanti   = BaglantiSayfasi()   # 1
        self.s2_tarama     = TaramaSayfasi()      # 2
        self.s3_eslestirme = EslestirmeSayfasi()  # 3
        self.s4_rapor      = RaporSayfasi()       # 4
        self.s5_maliyet    = MaliyetSayfasi()     # 5

        for s in [self.s0_menu, self.s1_baglanti, self.s2_tarama,
                  self.s3_eslestirme, self.s4_rapor, self.s5_maliyet]:
            self.stack.addWidget(s)

        # Sinyal bağlantıları
        self.s0_menu.arac1_signal.connect(lambda: self.sayfa_gec(1))
        self.s0_menu.arac2_signal.connect(lambda: self.sayfa_gec(5))
        self.s1_baglanti.devam.connect(lambda: self.sayfa_gec(2))
        self.s1_baglanti.geri.connect(lambda: self.sayfa_gec(0))
        self.s2_tarama.bitti.connect(lambda: self.sayfa_gec(3))
        self.s3_eslestirme.rapor_signal.connect(self._rapor_goster)

        merkez = QWidget()
        lay = QVBoxLayout(merkez)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        lay.addWidget(self.adim_bar)
        lay.addWidget(self.stack, stretch=1)
        self.setCentralWidget(merkez)

        sb = QStatusBar()
        sb.setStyleSheet("background:#f5f5f5;color:#555;font-size:11px;")
        self.setStatusBar(sb)
        sb.showMessage("Demo modu — gerçek bağlantı için sunucu bilgilerini girin.")

    def _adim_bar_olustur(self):
        frame = QFrame()
        frame.setFixedHeight(44)
        frame.setStyleSheet("background:#3f51b5;")
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(20, 0, 20, 0)
        self._adim_lbller = []
        for i, ad in enumerate(ADIMLAR_ARAC1):
            lbl_w = QLabel(ad)
            lbl_w.setAlignment(Qt.AlignCenter)
            lbl_w.setFont(QFont("Segoe UI", 10, QFont.Bold))
            lbl_w.setStyleSheet("color:rgba(255,255,255,0.4);border:none;")
            lay.addWidget(lbl_w)
            self._adim_lbller.append(lbl_w)
            if i < len(ADIMLAR_ARAC1) - 1:
                ay = QLabel("›")
                ay.setStyleSheet("color:rgba(255,255,255,0.3);border:none;font-size:16px;")
                ay.setAlignment(Qt.AlignCenter)
                lay.addWidget(ay)
        return frame

    def _adim_guncelle(self, sayfa_idx):
        # Sayfa 0 veya 5: özel başlık
        if sayfa_idx == 0:
            for w in self._adim_lbller:
                w.setStyleSheet("color:rgba(255,255,255,0.5);border:none;font-size:10px;")
            return
        if sayfa_idx == 5:
            for i, w in enumerate(self._adim_lbller):
                w.setText(["💰 Maliyet","Hesaplama","",""][i] if i < 2 else "")
                w.setStyleSheet("color:white;border:none;font-size:11px;" if i==0
                                else "color:rgba(255,255,255,0.5);border:none;")
            return
        # Araç 1 adımları (sayfa 1–4 → adım 0–3)
        for i, w in enumerate(self._adim_lbller):
            w.setText(ADIMLAR_ARAC1[i])
            aktif = sayfa_idx - 1
            if i == aktif:
                w.setStyleSheet("color:white;border:none;font-size:11px;")
            elif i < aktif:
                w.setStyleSheet("color:rgba(255,255,255,0.6);border:none;font-size:10px;")
            else:
                w.setStyleSheet("color:rgba(255,255,255,0.35);border:none;font-size:10px;")

    def sayfa_gec(self, idx):
        self.stack.setCurrentIndex(idx)
        self._adim_guncelle(idx)
        if idx == 2: self.s2_tarama.baslat()
        elif idx == 3: self.s3_eslestirme.baslat()

    def _rapor_goster(self, sonuclar):
        self.s4_rapor.goster(sonuclar)
        self.sayfa_gec(4)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    AnaPencere().show()
    sys.exit(app.exec_())
