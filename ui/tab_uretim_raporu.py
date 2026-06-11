"""tab_uretim_raporu.py — Üretim Eksik Stok Raporu sekmesi (BOM patlatmalı)."""
import logging
from datetime import datetime

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QListWidgetItem, QTreeWidget, QTreeWidgetItem,
    QHeaderView, QMessageBox, QFileDialog, QProgressBar, QGroupBox,
    QSplitter,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QColor, QBrush

from config import DB_DEFAULTS
from db.baglanti import get_connection
from db.sorgular import (
    uretim_emirleri_listesi, uretim_emir_bom_patlat,
    tum_emirler_eksik_stok, muhasebe_eksik_raporu_olustur,
)

_BTN_MAVI = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #1976d2);"
    "color:white;border-radius:6px;padding:8px 18px;font-weight:bold;font-size:12px;border:none;"
)
_BTN_YESIL = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2e7d32,stop:1 #43a047);"
    "color:white;border-radius:6px;padding:8px 18px;font-weight:bold;font-size:12px;border:none;"
)
_BTN_PASIF = (
    "background:#bdbdbd;color:#757575;"
    "border-radius:6px;padding:8px 18px;font-weight:bold;font-size:12px;border:none;"
)
_BTN_TURUNCU = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #e65100,stop:1 #f57c00);"
    "color:white;border-radius:6px;padding:8px 18px;font-weight:bold;font-size:12px;border:none;"
)
_BTN_MOR = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #4a148c,stop:1 #7b1fa2);"
    "color:white;border-radius:6px;padding:8px 18px;font-weight:bold;font-size:12px;border:none;"
)
_AGAC_KOLONLAR = [
    "Malzeme Kodu",
    "Malzeme Adı",
    "İhtiyaç Miktarı",
    "Emri Tarihindeki Bakiye",
    "Eksik Miktar",
]

_RENK_EKSIK      = QColor("#c62828")
_RENK_ALT_MONTAJ = QColor("#e8eaf6")   # alt montaj kök satırları — açık indigo


# ── Thread'ler ────────────────────────────────────────────────────────────────

class EmirlerThread(QThread):
    bitti = pyqtSignal(list)
    hata  = pyqtSignal(str)

    def __init__(self, conn):
        super().__init__()
        self.conn = conn

    def run(self):
        try:
            self.bitti.emit(uretim_emirleri_listesi(self.conn))
        except Exception as e:
            logging.error("EmirlerThread: %s", e)
            self.hata.emit(str(e))


class AnalizThread(QThread):
    bitti = pyqtSignal(list)
    hata  = pyqtSignal(str)

    def __init__(self, conn, emir_id: int):
        super().__init__()
        self.conn = conn
        self.emir_id = emir_id

    def run(self):
        try:
            self.bitti.emit(uretim_emir_bom_patlat(self.conn, self.emir_id))
        except Exception as e:
            logging.error("AnalizThread emir_id=%s: %s", self.emir_id, e)
            self.hata.emit(str(e))


class MuhasebeRaporuThread(QThread):
    bitti = pyqtSignal(list)
    hata  = pyqtSignal(str)

    def __init__(self, conn):
        super().__init__()
        self.conn = conn

    def run(self):
        try:
            self.bitti.emit(muhasebe_eksik_raporu_olustur(self.conn))
        except Exception as e:
            logging.error("MuhasebeRaporuThread: %s", e)
            self.hata.emit(str(e))


class TumEmirlerThread(QThread):
    ilerleme = pyqtSignal(int, int, str)   # (tamamlanan, toplam, emir_kodu)
    bitti    = pyqtSignal(list)
    hata     = pyqtSignal(str)

    def __init__(self, conn, emirler: list):
        super().__init__()
        self.conn    = conn
        self.emirler = emirler

    def run(self):
        try:
            sonuc = []
            toplam = len(self.emirler)
            for i, e in enumerate(self.emirler):
                self.ilerleme.emit(i + 1, toplam, e['kodu'])
                eksikler = uretim_emir_bom_patlat(self.conn, e['id'])
                if eksikler:
                    sonuc.append({**e, 'eksikler': eksikler})
            self.bitti.emit(sonuc)
        except Exception as ex:
            logging.error("TumEmirlerThread: %s", ex)
            self.hata.emit(str(ex))


# ── Ana Sekme ─────────────────────────────────────────────────────────────────

class UretimRaporuTab(QWidget):
    durum_guncelle = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.conn = None
        self._emirler: list[dict] = []
        self._secili_emir: dict | None = None
        self._analiz_veri: list[dict] = []
        self._emirler_thread     = None
        self._analiz_thread      = None
        self._tum_emirler_thread = None
        self._muhasebe_thread    = None
        self._kur()
        self._baglan_otomatik()

    def _kur(self):
        ana = QVBoxLayout(self)
        ana.setContentsMargins(12, 10, 12, 10)
        ana.setSpacing(8)

        # ── Başlık satırı ──────────────────────────────────────────────────────
        ust = QHBoxLayout()
        baslik = QLabel("Üretim Eksik Stok Raporu")
        baslik.setFont(QFont("Segoe UI", 14, QFont.Bold))
        baslik.setStyleSheet("color:#1a237e;")
        ust.addWidget(baslik)
        ust.addStretch()
        self.durum_lbl = QLabel("Bağlanılıyor…")
        self.durum_lbl.setStyleSheet("color:#555;font-size:11px;")
        ust.addWidget(self.durum_lbl)
        self.yenile_btn = QPushButton("🔄  Emirleri Yenile")
        self.yenile_btn.setStyleSheet(_BTN_MAVI)
        self.yenile_btn.setEnabled(False)
        self.yenile_btn.clicked.connect(self._emirleri_yukle)
        ust.addWidget(self.yenile_btn)

        self.tum_excel_btn = QPushButton("📊  Tüm Emirleri Excel'e Aktar")
        self.tum_excel_btn.setStyleSheet(_BTN_PASIF)
        self.tum_excel_btn.setEnabled(False)
        self.tum_excel_btn.clicked.connect(self._tum_excel_baslat)
        ust.addWidget(self.tum_excel_btn)

        self.muhasebe_btn = QPushButton("🧾  Muhasebe Eksik Raporu (Excel)")
        self.muhasebe_btn.setStyleSheet(_BTN_PASIF)
        self.muhasebe_btn.setEnabled(False)
        self.muhasebe_btn.clicked.connect(self._muhasebe_raporu_baslat)
        ust.addWidget(self.muhasebe_btn)
        ana.addLayout(ust)

        # İlerleme çubuğu
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setFixedHeight(4)
        self.progress.setStyleSheet(
            "QProgressBar{border:none;background:#e8eaf6;border-radius:2px;}"
            "QProgressBar::chunk{background:#3f51b5;border-radius:2px;}"
        )
        self.progress.setVisible(False)
        ana.addWidget(self.progress)

        # ── İki sütun: sol liste | sağ panel ──────────────────────────────────
        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(6)
        splitter.setStyleSheet("QSplitter::handle{background:#e0e0e0;}")

        # SOL — İş emri listesi
        sol = QWidget()
        sol_lay = QVBoxLayout(sol)
        sol_lay.setContentsMargins(0, 0, 4, 0)
        sol_lay.setSpacing(6)

        sol_baslik = QLabel("Devam Eden İş Emirleri")
        sol_baslik.setFont(QFont("Segoe UI", 11, QFont.Bold))
        sol_baslik.setStyleSheet("color:#3949ab;")
        sol_lay.addWidget(sol_baslik)

        self.emir_listesi = QListWidget()
        self.emir_listesi.setStyleSheet("""
            QListWidget {
                border: 1px solid #c5cae9;
                border-radius: 6px;
                background: white;
                font-size: 12px;
            }
            QListWidget::item {
                padding: 8px 10px;
                border-bottom: 1px solid #f0f2f5;
            }
            QListWidget::item:selected {
                background: #3f51b5;
                color: white;
            }
            QListWidget::item:hover:!selected {
                background: #e8eaf6;
            }
        """)
        self.emir_listesi.currentItemChanged.connect(self._emir_secildi)
        sol_lay.addWidget(self.emir_listesi)

        self.liste_ozet = QLabel("")
        self.liste_ozet.setStyleSheet("color:#757575;font-size:10px;padding:2px;")
        self.liste_ozet.setAlignment(Qt.AlignCenter)
        sol_lay.addWidget(self.liste_ozet)
        splitter.addWidget(sol)

        # SAĞ — Analiz paneli
        sag = QWidget()
        sag_lay = QVBoxLayout(sag)
        sag_lay.setContentsMargins(4, 0, 0, 0)
        sag_lay.setSpacing(6)

        # Seçili emir bilgi kartı
        self.emir_kart = QGroupBox("Seçili İş Emri")
        self.emir_kart.setStyleSheet(
            "QGroupBox{border:1.5px solid #c5cae9;border-radius:8px;"
            "margin-top:8px;padding:10px 8px 8px 8px;background:white;}"
            "QGroupBox::title{subcontrol-origin:margin;left:10px;padding:0 6px;"
            "color:#3949ab;font-weight:bold;font-size:12px;}"
        )
        kart_lay = QHBoxLayout(self.emir_kart)
        kart_lay.setSpacing(24)
        self.kart_kodu     = self._bilgi_widget("Emir Kodu", "—")
        self.kart_tarih    = self._bilgi_widget("Emri Tarihi", "—")
        self.kart_aciklama = self._bilgi_widget("Açıklama", "—")
        kart_lay.addWidget(self.kart_kodu)
        kart_lay.addWidget(self.kart_tarih)
        kart_lay.addWidget(self.kart_aciklama, stretch=1)
        kart_lay.addStretch()

        self.analiz_btn = QPushButton("🔍  Analiz Et")
        self.analiz_btn.setStyleSheet(_BTN_MAVI)
        self.analiz_btn.setEnabled(False)
        self.analiz_btn.setFixedHeight(36)
        self.analiz_btn.clicked.connect(self._analiz_et)
        kart_lay.addWidget(self.analiz_btn)

        self.excel_btn = QPushButton("📥  Excel'e Aktar")
        self.excel_btn.setStyleSheet(_BTN_PASIF)
        self.excel_btn.setEnabled(False)
        self.excel_btn.setFixedHeight(36)
        self.excel_btn.clicked.connect(self._excel_aktar)
        kart_lay.addWidget(self.excel_btn)
        sag_lay.addWidget(self.emir_kart)

        # Eksik malzeme ağacı (BOM patlatmalı)
        self.agac = QTreeWidget()
        self.agac.setColumnCount(len(_AGAC_KOLONLAR))
        self.agac.setHeaderLabels(_AGAC_KOLONLAR)
        self.agac.setEditTriggers(QTreeWidget.NoEditTriggers)
        self.agac.setSelectionBehavior(QTreeWidget.SelectRows)
        self.agac.setAlternatingRowColors(False)
        self.agac.setRootIsDecorated(True)
        self.agac.setStyleSheet("""
            QTreeWidget {
                border: 1px solid #c5cae9;
                border-radius: 6px;
                background: white;
                font-size: 12px;
            }
            QTreeWidget::item { padding: 4px 8px; }
            QTreeWidget::item:selected { background: #e8eaf6; color: #1a237e; }
            QTreeWidget::item:hover:!selected { background: #f5f5ff; }
            QHeaderView::section {
                background: #3f51b5; color: white;
                padding: 7px 8px; font-weight: bold; font-size: 12px;
                border: none; border-right: 1px solid #5c6bc0;
            }
        """)
        h = self.agac.header()
        h.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.Stretch)
        h.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        sag_lay.addWidget(self.agac)

        # Alt özet
        self.tablo_ozet = QLabel("Sol listeden bir iş emri seçip 'Analiz Et' düğmesine basın.")
        self.tablo_ozet.setStyleSheet(
            "color:#3949ab;font-size:12px;font-weight:bold;padding:3px;"
        )
        sag_lay.addWidget(self.tablo_ozet)
        splitter.addWidget(sag)

        splitter.setSizes([280, 800])
        ana.addWidget(splitter)

    def _bilgi_widget(self, etiket: str, deger: str) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(2)
        lbl_e = QLabel(etiket)
        lbl_e.setStyleSheet("color:#9e9e9e;font-size:10px;")
        lbl_d = QLabel(deger)
        lbl_d.setFont(QFont("Segoe UI", 11, QFont.Bold))
        lbl_d.setStyleSheet("color:#1a237e;")
        lbl_d.setObjectName("deger")
        lay.addWidget(lbl_e)
        lay.addWidget(lbl_d)
        return w

    def _bilgi_guncelle(self, widget: QWidget, deger: str):
        lbl = widget.findChild(QLabel, "deger")
        if lbl:
            lbl.setText(deger)

    # ── Bağlantı & Emirler ────────────────────────────────────────────────────

    def _baglan_otomatik(self):
        try:
            self.conn = get_connection(
                DB_DEFAULTS["sunucu"], DB_DEFAULTS["veritabani"],
                DB_DEFAULTS["kullanici"], DB_DEFAULTS["sifre"],
            )
            self.yenile_btn.setEnabled(True)
            self._emirleri_yukle()
        except Exception as e:
            logging.error("UretimRaporuTab baglanti: %s", e)
            self.durum_lbl.setText(f"Bağlantı hatası: {e}")
            self.durum_lbl.setStyleSheet("color:#c62828;font-size:11px;")

    def _emirleri_yukle(self):
        if not self.conn:
            return
        self.yenile_btn.setEnabled(False)
        self.emir_listesi.clear()
        self._emirler = []
        self.progress.setVisible(True)
        self.durum_lbl.setText("İş emirleri yükleniyor…")
        self.durum_lbl.setStyleSheet("color:#1565c0;font-size:11px;")

        self._emirler_thread = EmirlerThread(self.conn)
        self._emirler_thread.bitti.connect(self._emirler_geldi)
        self._emirler_thread.hata.connect(self._hata)
        self._emirler_thread.start()

    def _emirler_geldi(self, emirler: list[dict]):
        self.progress.setVisible(False)
        self.yenile_btn.setEnabled(True)
        self._emirler = emirler
        self.emir_listesi.clear()

        for e in emirler:
            item = QListWidgetItem()
            aciklama = e['aciklama'][:55] + "…" if len(e['aciklama']) > 55 else e['aciklama']
            item.setText(f"{e['kodu']}\n{e['tarih']}  •  {aciklama}")
            item.setData(Qt.UserRole, e)
            item.setSizeHint(QSize(0, 52))
            self.emir_listesi.addItem(item)

        self.liste_ozet.setText(f"{len(emirler)} iş emri")
        self.durum_lbl.setText(f"Yüklendi: {datetime.now().strftime('%H:%M:%S')}")
        self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;")
        self.durum_guncelle.emit(f"{len(emirler)} devam eden iş emri yüklendi.")
        if emirler:
            self.tum_excel_btn.setEnabled(True)
            self.tum_excel_btn.setStyleSheet(_BTN_TURUNCU)
            self.muhasebe_btn.setEnabled(True)
            self.muhasebe_btn.setStyleSheet(_BTN_MOR)

    # ── Emir Seçimi ───────────────────────────────────────────────────────────

    def _emir_secildi(self, current: QListWidgetItem, _prev):
        if current is None:
            return
        e = current.data(Qt.UserRole)
        self._secili_emir = e
        aciklama = e['aciklama'][:70] + "…" if len(e['aciklama']) > 70 else e['aciklama']
        self._bilgi_guncelle(self.kart_kodu,     e['kodu'])
        self._bilgi_guncelle(self.kart_tarih,    e['tarih'])
        self._bilgi_guncelle(self.kart_aciklama, aciklama)
        self.analiz_btn.setEnabled(True)
        self.excel_btn.setEnabled(False)
        self.excel_btn.setStyleSheet(_BTN_PASIF)
        self.agac.clear()
        self.tablo_ozet.setText("'Analiz Et' düğmesine basarak eksik malzemeleri görün.")
        self._analiz_veri = []

    # ── Analiz ────────────────────────────────────────────────────────────────

    def _analiz_et(self):
        if not self._secili_emir or not self.conn:
            return
        self.analiz_btn.setEnabled(False)
        self.excel_btn.setEnabled(False)
        self.excel_btn.setStyleSheet(_BTN_PASIF)
        self.agac.clear()
        self._analiz_veri = []
        self.progress.setVisible(True)
        self.tablo_ozet.setText("Analiz ediliyor…")
        self.durum_guncelle.emit(f"Analiz ediliyor: {self._secili_emir['kodu']}")

        self._analiz_thread = AnalizThread(self.conn, self._secili_emir['id'])
        self._analiz_thread.bitti.connect(self._analiz_geldi)
        self._analiz_thread.hata.connect(self._hata)
        self._analiz_thread.start()

    def _dugum_olustur(self, m: dict, is_child: bool = False) -> QTreeWidgetItem:
        """Tek bir malzeme kaydı için QTreeWidgetItem oluşturur."""
        item = QTreeWidgetItem([
            m['kodu'],
            m['adi'],
            f"{m['ihtiyac']:.2f}",
            f"{m['bakiye']:.2f}",
            f"{m['eksik']:.2f}",
        ])
        for col in range(5):
            item.setTextAlignment(
                col,
                Qt.AlignVCenter | (Qt.AlignRight if col >= 2 else Qt.AlignLeft)
            )
        # Alt montaj kök satırı: açık indigo arka plan
        if not is_child and m.get('is_alt_montaj'):
            for col in range(5):
                item.setBackground(col, QBrush(_RENK_ALT_MONTAJ))
        # Bakiye sıfır veya negatif: kırmızı
        if m['bakiye'] <= 0:
            item.setForeground(3, QBrush(_RENK_EKSIK))
        # Eksik negatif: kırmızı ve kalın
        if m['eksik'] < 0:
            item.setForeground(4, QBrush(_RENK_EKSIK))
            f = QFont()
            f.setBold(True)
            item.setFont(4, f)
        return item

    def _ekle_dugum(self, parent, m: dict, is_child: bool = False):
        """Düğümü ve alt bileşenlerini recursive olarak ekler."""
        item = self._dugum_olustur(m, is_child=is_child)
        if parent is None:
            self.agac.addTopLevelItem(item)
        else:
            parent.addChild(item)
        for bil in m.get('bilesenler', []):
            self._ekle_dugum(item, bil, is_child=True)
        return item

    def _analiz_geldi(self, veri: list[dict]):
        self.progress.setVisible(False)
        self.analiz_btn.setEnabled(True)
        self._analiz_veri = veri

        self.agac.clear()
        if not veri:
            self.tablo_ozet.setText(
                f"✅  '{self._secili_emir['kodu']}' için tüm malzemeler stokta mevcut — eksik yok."
            )
            self.tablo_ozet.setStyleSheet("color:#2e7d32;font-size:12px;font-weight:bold;padding:3px;")
            self.durum_guncelle.emit(f"{self._secili_emir['kodu']}: eksik malzeme yok.")
            return

        for m in veri:
            self._ekle_dugum(None, m, is_child=False)

        self.agac.expandAll()

        n_kok = len(veri)
        n_alt = sum(1 for m in veri if m.get('bilesenler'))
        ozet  = f"{n_kok} eksik kalem"
        if n_alt:
            ozet += f" ({n_alt} alt montaj açıldı)"
        self.tablo_ozet.setText(
            f"{self._secili_emir['kodu']} için {ozet} bulundu."
        )
        self.tablo_ozet.setStyleSheet("color:#c62828;font-size:12px;font-weight:bold;padding:3px;")
        self.durum_guncelle.emit(f"{self._secili_emir['kodu']}: {ozet}.")
        self.excel_btn.setEnabled(True)
        self.excel_btn.setStyleSheet(_BTN_YESIL)

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _excel_aktar(self):
        if not self._analiz_veri or not self._secili_emir:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            emir = self._secili_emir
            oneri = f"eksik_{emir['kodu'].replace('.', '_').replace('/', '_')}.xlsx"
            dosya, _ = QFileDialog.getSaveFileName(
                self, "Excel Olarak Kaydet", oneri,
                "Excel Dosyaları (*.xlsx)"
            )
            if not dosya:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Eksik Stok"

            ws.append(["Emir Kodu",  emir['kodu']])
            ws.append(["Emir Tarihi", emir['tarih']])
            ws.append(["Açıklama",   emir['aciklama']])
            ws.append([])
            ws.append(_AGAC_KOLONLAR)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.fill      = PatternFill("solid", fgColor="3F51B5")
                cell.font      = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            eksik_font  = Font(name="Segoe UI", size=11, color="C62828", bold=True)
            normal_font = Font(name="Segoe UI", size=11)
            alt_fill    = PatternFill("solid", fgColor="E8EAF6")

            def _yaz(m: dict, derinlik: int = 0):
                girinti = "  " * derinlik + ("└─ " if derinlik > 0 else "")
                ws.append([
                    girinti + m['kodu'],
                    m['adi'],
                    round(m['ihtiyac'], 2),
                    round(m['bakiye'],  2),
                    round(m['eksik'],   2),
                ])
                r = ws.max_row
                for c in range(3, 6):
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
                ws.cell(row=r, column=4).font = eksik_font if m['bakiye'] <= 0 else normal_font
                ws.cell(row=r, column=5).font = eksik_font if m['eksik'] < 0 else normal_font
                if derinlik == 0 and m.get('is_alt_montaj'):
                    for c in range(1, 6):
                        ws.cell(row=r, column=c).fill = alt_fill
                for bil in m.get('bilesenler', []):
                    _yaz(bil, derinlik + 1)

            for m in self._analiz_veri:
                _yaz(m)

            for col, w in zip("ABCDE", [28, 50, 16, 16, 14]):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = f"A{header_row + 1}"

            wb.save(dosya)
            QMessageBox.information(self, "Başarılı", f"Excel kaydedildi:\n{dosya}")
            self.durum_guncelle.emit(f"Excel kaydedildi: {dosya}")
        except Exception as e:
            logging.error("UretimRaporuTab excel: %s", e)
            QMessageBox.critical(self, "Excel Hatası", str(e))

    # ── Tüm Emirler Excel ─────────────────────────────────────────────────────

    def _tum_excel_baslat(self):
        if not self._emirler or not self.conn:
            return
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "tum_emirler_eksik_stok.xlsx",
            "Excel Dosyaları (*.xlsx)"
        )
        if not dosya:
            return

        self.tum_excel_btn.setEnabled(False)
        self.tum_excel_btn.setStyleSheet(_BTN_PASIF)
        self.muhasebe_btn.setEnabled(False)
        self.muhasebe_btn.setStyleSheet(_BTN_PASIF)
        self.yenile_btn.setEnabled(False)
        self.analiz_btn.setEnabled(False)
        self.progress.setRange(0, len(self._emirler))
        self.progress.setValue(0)
        self.progress.setVisible(True)
        self.durum_lbl.setText("Tüm emirler analiz ediliyor…")
        self.durum_lbl.setStyleSheet("color:#e65100;font-size:11px;")

        self._tum_excel_dosya = dosya
        self._tum_emirler_thread = TumEmirlerThread(self.conn, self._emirler)
        self._tum_emirler_thread.ilerleme.connect(self._tum_ilerleme)
        self._tum_emirler_thread.bitti.connect(self._tum_excel_kaydet)
        self._tum_emirler_thread.hata.connect(self._hata)
        self._tum_emirler_thread.start()

    def _tum_ilerleme(self, tamamlanan: int, toplam: int, emir_kodu: str):
        self.progress.setValue(tamamlanan)
        self.durum_lbl.setText(f"İşleniyor {tamamlanan}/{toplam}: {emir_kodu}")

    def _tum_excel_kaydet(self, veri: list[dict]):
        self.progress.setVisible(False)
        self.progress.setRange(0, 0)
        self.yenile_btn.setEnabled(True)
        self.tum_excel_btn.setEnabled(True)
        self.tum_excel_btn.setStyleSheet(_BTN_TURUNCU)
        self.muhasebe_btn.setEnabled(True)
        self.muhasebe_btn.setStyleSheet(_BTN_MOR)
        if self._secili_emir:
            self.analiz_btn.setEnabled(True)

        if not veri:
            QMessageBox.information(self, "Sonuç", "Tüm açık emirlerde eksik malzeme bulunamadı.")
            self.durum_lbl.setText("Tüm emirler: eksik yok.")
            self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Eksik Stok"

            # Başlık satırı
            basliklar = [
                "Emir Kodu", "Emir Tarihi", "Açıklama",
                "Malzeme Kodu", "Malzeme Adı",
                "İhtiyaç Miktarı", "Emri Tarihindeki Bakiye", "Eksik Miktar",
            ]
            ws.append(basliklar)
            hrow = ws.max_row
            for cell in ws[hrow]:
                cell.fill      = PatternFill("solid", fgColor="1A237E")
                cell.font      = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[hrow].height = 22

            # Stiller
            emir_fill   = PatternFill("solid", fgColor="E8F5E9")   # yeşilimsi — emir header
            alt_fill    = PatternFill("solid", fgColor="E8EAF6")   # indigo — alt montaj
            eksik_font  = Font(name="Segoe UI", size=11, color="C62828", bold=True)
            normal_font = Font(name="Segoe UI", size=11)
            emir_font   = Font(name="Segoe UI", size=11, bold=True, color="1B5E20")
            ince_kenarl = Border(
                bottom=Side(style="thin", color="E0E0E0")
            )

            def _yaz_malzeme(m: dict, emir_kodu: str, emir_tarih: str,
                             emir_aciklama: str, derinlik: int):
                girinti = "    " * derinlik + ("└─ " if derinlik > 0 else "")
                ws.append([
                    "" if derinlik > 0 else emir_kodu,
                    "" if derinlik > 0 else emir_tarih,
                    "" if derinlik > 0 else emir_aciklama,
                    girinti + m['kodu'],
                    m['adi'],
                    round(m['ihtiyac'], 2),
                    round(m['bakiye'],  2),
                    round(m['eksik'],   2),
                ])
                r = ws.max_row
                for c in range(6, 9):
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
                ws.cell(row=r, column=7).font = eksik_font if m['bakiye'] <= 0 else normal_font
                ws.cell(row=r, column=8).font = eksik_font if m['eksik']  < 0 else normal_font
                if derinlik == 0 and m.get('is_alt_montaj'):
                    for c in range(1, 9):
                        ws.cell(row=r, column=c).fill = alt_fill
                for bil in m.get('bilesenler', []):
                    _yaz_malzeme(bil, emir_kodu, emir_tarih, emir_aciklama, derinlik + 1)

            for e in veri:
                # Emir header boş ayırıcı satır
                ws.append([
                    e['kodu'], e['tarih'],
                    e['aciklama'][:80] if e['aciklama'] else "",
                    "", "", "", "", "",
                ])
                r = ws.max_row
                for c in range(1, 9):
                    ws.cell(row=r, column=c).fill = emir_fill
                    ws.cell(row=r, column=c).font = emir_font
                ws.row_dimensions[r].height = 18

                for m in e['eksikler']:
                    _yaz_malzeme(m, e['kodu'], e['tarih'], e['aciklama'], 0)

            # Sütun genişlikleri
            for col, w in zip("ABCDEFGH", [22, 12, 45, 30, 50, 16, 20, 14]):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = f"A{hrow + 1}"

            wb.save(self._tum_excel_dosya)

            toplam_eksik = sum(len(e['eksikler']) for e in veri)
            QMessageBox.information(
                self, "Başarılı",
                f"{len(veri)} emirde toplam {toplam_eksik} eksik kalem.\n"
                f"Excel kaydedildi:\n{self._tum_excel_dosya}"
            )
            self.durum_lbl.setText(
                f"Tüm emirler: {len(veri)} emirde {toplam_eksik} eksik kalem — Excel kaydedildi."
            )
            self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;")
            self.durum_guncelle.emit(
                f"Tüm emirler Excel'e aktarıldı: {len(veri)} emir, {toplam_eksik} eksik kalem."
            )

        except Exception as ex:
            logging.error("TumEmirler excel: %s", ex)
            QMessageBox.critical(self, "Excel Hatası", str(ex))

    # ── Muhasebe Raporu ───────────────────────────────────────────────────────

    def _muhasebe_raporu_baslat(self):
        if not self._emirler or not self.conn:
            return
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Muhasebe Eksik Raporu — Kaydet",
            "muhasebe_eksik_raporu.xlsx",
            "Excel Dosyaları (*.xlsx)",
        )
        if not dosya:
            return

        self.muhasebe_btn.setEnabled(False)
        self.muhasebe_btn.setStyleSheet(_BTN_PASIF)
        self.tum_excel_btn.setEnabled(False)
        self.tum_excel_btn.setStyleSheet(_BTN_PASIF)
        self.yenile_btn.setEnabled(False)
        self.analiz_btn.setEnabled(False)
        self.progress.setRange(0, 0)
        self.progress.setVisible(True)
        self.durum_lbl.setText("Muhasebe raporu hesaplanıyor…")
        self.durum_lbl.setStyleSheet("color:#4a148c;font-size:11px;")

        self._muhasebe_dosya = dosya
        self._muhasebe_thread = MuhasebeRaporuThread(self.conn)
        self._muhasebe_thread.bitti.connect(self._muhasebe_excel_kaydet)
        self._muhasebe_thread.hata.connect(self._hata)
        self._muhasebe_thread.start()

    def _muhasebe_excel_kaydet(self, satirlar: list):
        self.progress.setVisible(False)
        self.progress.setRange(0, 0)
        self.yenile_btn.setEnabled(True)
        self.muhasebe_btn.setEnabled(True)
        self.muhasebe_btn.setStyleSheet(_BTN_MOR)
        self.tum_excel_btn.setEnabled(True)
        self.tum_excel_btn.setStyleSheet(_BTN_TURUNCU)
        if self._secili_emir:
            self.analiz_btn.setEnabled(True)

        if not satirlar:
            QMessageBox.information(
                self, "Muhasebe Raporu",
                "Tüm açık emirlerde (ATP dahil) eksik malzeme bulunamadı.",
            )
            self.durum_lbl.setText("Muhasebe raporu: eksik yok.")
            self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Muhasebe Eksik Raporu"

            # Başlık satırı
            BASLIKLAR = [
                "İş Emri Tarihi", "İş Emri No", "Açıklama",
                "Stok Kodu", "Stok Adı",
                "İhtiyaç Miktarı",
                "O Tarihteki ERP Bakiyesi",
                "Önceki Emirlerin Rezervasyonu",
                "Net Eksik Miktar",
            ]
            ws.append(BASLIKLAR)
            hrow = ws.max_row
            for cell in ws[hrow]:
                cell.fill      = PatternFill("solid", fgColor="4A148C")
                cell.font      = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[hrow].height = 28

            # Stiller
            eksik_font  = Font(name="Segoe UI", size=11, color="C62828", bold=True)
            normal_font = Font(name="Segoe UI", size=11)
            emir_fill   = PatternFill("solid", fgColor="EDE7F6")
            ince_alt    = Border(bottom=Side(style="thin", color="D1C4E9"))

            secili_emir = None
            for s in satirlar:
                # Emir değiştiğinde ayırıcı satır
                if s['emir_kodu'] != secili_emir:
                    secili_emir = s['emir_kodu']
                    ws.append([
                        s['emir_tarihi'],
                        s['emir_kodu'],
                        s['emir_aciklama'][:80] if s['emir_aciklama'] else "",
                        "", "", "", "", "", "",
                    ])
                    r = ws.max_row
                    for c in range(1, 10):
                        ws.cell(row=r, column=c).fill = emir_fill
                        ws.cell(row=r, column=c).font = Font(
                            name="Segoe UI", bold=True, color="4A148C", size=11
                        )
                    ws.row_dimensions[r].height = 18

                ws.append([
                    s['emir_tarihi'],
                    s['emir_kodu'],
                    "",
                    s['stok_kodu'],
                    s['stok_adi'],
                    s['ihtiyac'],
                    s['erp_bakiye'],
                    s['onceki_rezervasyon'],
                    s['net_eksik'],
                ])
                r = ws.max_row
                for c in range(6, 10):
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
                    ws.cell(row=r, column=c).font = normal_font
                # Net Eksik: kırmızı kalın
                ws.cell(row=r, column=9).font = eksik_font
                # ERP Bakiye negatifse: kırmızı
                if s['erp_bakiye'] < 0:
                    ws.cell(row=r, column=7).font = eksik_font
                for c in range(1, 10):
                    ws.cell(row=r, column=c).border = ince_alt

            for col, w in zip("ABCDEFGHI", [14, 22, 40, 22, 45, 16, 22, 24, 18]):
                ws.column_dimensions[col].width = w
            ws.freeze_panes = f"A{hrow + 1}"
            ws.auto_filter.ref = f"A{hrow}:I{ws.max_row}"

            wb.save(self._muhasebe_dosya)

            toplam_emir = len({s['emir_kodu'] for s in satirlar})
            QMessageBox.information(
                self, "Muhasebe Raporu Oluşturuldu",
                f"{toplam_emir} emirde toplam {len(satirlar)} eksik kalem.\n"
                f"Excel kaydedildi:\n{self._muhasebe_dosya}",
            )
            self.durum_lbl.setText(
                f"Muhasebe raporu: {toplam_emir} emir, {len(satirlar)} satır — kaydedildi."
            )
            self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;")
            self.durum_guncelle.emit(
                f"Muhasebe eksik raporu oluşturuldu: {toplam_emir} emir, {len(satirlar)} satır."
            )

        except Exception as ex:
            logging.error("MuhasebeRaporu excel: %s", ex)
            QMessageBox.critical(self, "Excel Hatası", str(ex))

    # ── Hata ─────────────────────────────────────────────────────────────────

    def _hata(self, mesaj: str):
        self.progress.setVisible(False)
        self.progress.setRange(0, 0)
        self.yenile_btn.setEnabled(True)
        self.analiz_btn.setEnabled(bool(self._secili_emir))
        if self._emirler:
            self.tum_excel_btn.setEnabled(True)
            self.tum_excel_btn.setStyleSheet(_BTN_TURUNCU)
            self.muhasebe_btn.setEnabled(True)
            self.muhasebe_btn.setStyleSheet(_BTN_MOR)
        self.durum_lbl.setText(f"Hata: {mesaj}")
        self.durum_lbl.setStyleSheet("color:#c62828;font-size:11px;")
        QMessageBox.critical(self, "Hata", mesaj)
