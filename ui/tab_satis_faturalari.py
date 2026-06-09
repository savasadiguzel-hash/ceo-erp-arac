"""tab_satis_faturalari.py — Satış Faturaları sekmesi."""
import logging
from datetime import datetime, date as date_type

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox, QGridLayout,
    QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QFileDialog,
)
from PyQt5.QtCore import Qt, QEvent, QThread, pyqtSignal
from PyQt5.QtGui import QFont

from config import DB_DEFAULTS
from db.baglanti import get_connection
from db.sorgular import satis_faturalari
from ui.stil import etiket

_BTN_MAVI = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #1976d2);"
    "color:white;border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)
_BTN_YESIL = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2e7d32,stop:1 #43a047);"
    "color:white;border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)
_BTN_PASIF = (
    "background:#bdbdbd;color:#757575;"
    "border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)

_KOLONLAR = [
    "Tarih", "Belge Türü", "Belge No", "Müşteri",
    "Stok Kodu", "Stok Adı", "Miktar", "Birim Fiyat (₺)", "Tutar (₺)",
]


# ── Thread'ler ────────────────────────────────────────────────────────────────

class BaglantiThread(QThread):
    bitti = pyqtSignal(object)
    hata  = pyqtSignal(str)

    def __init__(self, sunucu, veritabani, kullanici, sifre):
        super().__init__()
        self.sunucu = sunucu
        self.veritabani = veritabani
        self.kullanici = kullanici
        self.sifre = sifre

    def run(self):
        try:
            conn = get_connection(self.sunucu, self.veritabani, self.kullanici, self.sifre)
            self.bitti.emit(conn)
        except Exception as e:
            logging.error("SatisFaturalari BaglantiThread: %s", e)
            self.hata.emit(str(e))


class VeriCekmeThread(QThread):
    bitti = pyqtSignal(list)
    hata  = pyqtSignal(str)

    def __init__(self, conn, bas_tarih, bit_tarih):
        super().__init__()
        self.conn = conn
        self.bas_tarih = bas_tarih
        self.bit_tarih = bit_tarih

    def run(self):
        try:
            veri = satis_faturalari(self.conn, self.bas_tarih, self.bit_tarih)
            self.bitti.emit(veri)
        except Exception as e:
            logging.error("SatisFaturalari VeriCekmeThread: %s", e)
            self.hata.emit(str(e))


# ── Ana Sekme ─────────────────────────────────────────────────────────────────

class SatisFaturalariTab(QWidget):
    durum_guncelle = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.conn = None
        self._veri: list[dict] = []
        self._baglanti_thread = None
        self._veri_thread = None
        self._kur()

    def _kur(self):
        ana = QHBoxLayout(self)
        ana.setContentsMargins(16, 12, 16, 12)
        ana.setSpacing(16)
        ana.addWidget(self._sol_panel(), stretch=1)
        ana.addWidget(self._sag_panel(), stretch=3)

    # ── Sol panel (ayarlar) ───────────────────────────────────────────────────
    def _sol_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        bas = QLabel("Satış Faturaları")
        bas.setFont(QFont("Segoe UI", 13, QFont.Bold))
        bas.setStyleSheet("color:#1a237e;")
        lay.addWidget(bas)

        # DB bağlantı
        db_grup = QGroupBox("Veritabanı Bağlantısı")
        dg = QGridLayout(db_grup)
        dg.setSpacing(8)
        self.sf_sunucu    = QLineEdit(DB_DEFAULTS["sunucu"])
        self.sf_db        = QLineEdit(DB_DEFAULTS["veritabani"])
        self.sf_kullanici = QLineEdit(DB_DEFAULTS["kullanici"])
        self.sf_sifre     = QLineEdit(DB_DEFAULTS.get("sifre", ""))
        self.sf_sifre.setEchoMode(QLineEdit.Password)
        self.sf_sifre.setPlaceholderText("••••••••")
        for i, (k, v) in enumerate([
            ("Sunucu:",    self.sf_sunucu),
            ("Veritabanı:", self.sf_db),
            ("Kullanıcı:", self.sf_kullanici),
            ("Şifre:",     self.sf_sifre),
        ]):
            dg.addWidget(etiket(k), i, 0)
            dg.addWidget(v, i, 1)
        lay.addWidget(db_grup)

        self.baglan_btn = QPushButton("🔌  Bağlan")
        self.baglan_btn.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #1976d2);"
            "color:white;border-radius:6px;padding:8px 16px;font-weight:bold;font-size:12px;"
        )
        self.baglan_btn.clicked.connect(self._baglan)
        lay.addWidget(self.baglan_btn)

        # Tarih aralığı
        tarih_grup = QGroupBox("Tarih Aralığı")
        tg = QGridLayout(tarih_grup)
        tg.setSpacing(8)
        self.tarih_bas = QLineEdit()
        self.tarih_bas.setPlaceholderText("GG.AA.YYYY")
        self.tarih_bit = QLineEdit()
        self.tarih_bit.setPlaceholderText("GG.AA.YYYY")
        hint = QLabel("Bitiş kutusunda Ctrl+N → bugün.")
        hint.setStyleSheet("color:#9e9e9e;font-size:10px;")
        tg.addWidget(etiket("Başlangıç:"), 0, 0)
        tg.addWidget(self.tarih_bas,       0, 1)
        tg.addWidget(etiket("Bitiş:"),     1, 0)
        tg.addWidget(self.tarih_bit,       1, 1)
        tg.addWidget(hint,                 2, 0, 1, 2)
        self.tarih_bas.editingFinished.connect(self._bas_validate)
        self.tarih_bit.editingFinished.connect(self._bit_validate)
        self.tarih_bit.installEventFilter(self)
        lay.addWidget(tarih_grup)

        lay.addStretch()

        self.getir_btn = QPushButton("🔍  Faturaları Getir")
        self.getir_btn.setStyleSheet(_BTN_MAVI)
        self.getir_btn.clicked.connect(self._getir)
        lay.addWidget(self.getir_btn)

        self.excel_btn = QPushButton("📥  Excel'e Aktar")
        self.excel_btn.setStyleSheet(_BTN_PASIF)
        self.excel_btn.setEnabled(False)
        self.excel_btn.clicked.connect(self._excel_aktar)
        lay.addWidget(self.excel_btn)

        self.durum_lbl = QLabel("")
        self.durum_lbl.setStyleSheet("color:#555;font-size:11px;padding:4px;")
        self.durum_lbl.setAlignment(Qt.AlignCenter)
        self.durum_lbl.setWordWrap(True)
        lay.addWidget(self.durum_lbl)
        return w

    # ── Sağ panel (tablo) ─────────────────────────────────────────────────────
    def _sag_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        self.arama = QLineEdit()
        self.arama.setPlaceholderText("🔍  Müşteri, stok kodu veya belge no ile filtrele…")
        self.arama.setStyleSheet(
            "border:1px solid #c5cae9;border-radius:6px;"
            "padding:6px 10px;font-size:12px;background:white;"
        )
        self.arama.textChanged.connect(self._filtrele)
        lay.addWidget(self.arama)

        self.ozet_lbl = QLabel(
            "Bağlanın ve tarih seçip 'Faturaları Getir' düğmesine basın."
        )
        self.ozet_lbl.setStyleSheet(
            "color:#3949ab;font-size:12px;font-weight:bold;padding:4px;"
        )
        lay.addWidget(self.ozet_lbl)

        self.tablo = QTableWidget()
        self.tablo.setColumnCount(len(_KOLONLAR))
        self.tablo.setHorizontalHeaderLabels(_KOLONLAR)
        self.tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo.setSelectionBehavior(QTableWidget.SelectRows)
        self.tablo.setAlternatingRowColors(True)
        self.tablo.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tablo.horizontalHeader().setStretchLastSection(True)
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setStyleSheet(
            "QTableWidget{background:white;gridline-color:#e0e0e0;font-size:12px;}"
            "QTableWidget::item{padding:4px 8px;}"
            "QHeaderView::section{background:#e8eaf6;color:#1a237e;font-weight:bold;"
            "font-size:12px;padding:6px;border:none;border-right:1px solid #c5cae9;}"
            "QTableWidget::item:selected{background:#e8eaf6;color:#1a237e;}"
        )
        lay.addWidget(self.tablo)
        return w

    # ── Tarih doğrulama ───────────────────────────────────────────────────────
    def eventFilter(self, obj, event):
        if (obj is self.tarih_bit
                and event.type() == QEvent.KeyPress
                and event.key() == Qt.Key_N
                and event.modifiers() == Qt.ControlModifier):
            self.tarih_bit.setText(date_type.today().strftime("%d.%m.%Y"))
            return True
        return super().eventFilter(obj, event)

    @staticmethod
    def _parse(metin: str):
        metin = metin.strip()
        if len(metin) == 8 and metin.isdigit():
            metin = f"{metin[:2]}.{metin[2:4]}.{metin[4:]}"
        try:
            dt = datetime.strptime(metin, "%d.%m.%Y").date()
            return None if dt > date_type.today() else dt
        except ValueError:
            return None

    def _bas_validate(self):
        dt = self._parse(self.tarih_bas.text())
        if dt is None:
            self.tarih_bas.clear()
            return
        self.tarih_bas.setText(dt.strftime("%d.%m.%Y"))
        bit_dt = self._parse(self.tarih_bit.text())
        if bit_dt is not None and dt > bit_dt:
            self.tarih_bit.setText(dt.strftime("%d.%m.%Y"))

    def _bit_validate(self):
        dt = self._parse(self.tarih_bit.text())
        if dt is None:
            self.tarih_bit.clear()
            return
        self.tarih_bit.setText(dt.strftime("%d.%m.%Y"))
        bas_dt = self._parse(self.tarih_bas.text())
        if bas_dt is not None and dt < bas_dt:
            self.tarih_bit.setText(bas_dt.strftime("%d.%m.%Y"))

    # ── Bağlantı ──────────────────────────────────────────────────────────────
    def _baglan(self):
        self.baglan_btn.setEnabled(False)
        self.baglan_btn.setText("⏳  Bağlanıyor…")
        self._baglanti_thread = BaglantiThread(
            self.sf_sunucu.text(), self.sf_db.text(),
            self.sf_kullanici.text(), self.sf_sifre.text(),
        )
        self._baglanti_thread.bitti.connect(self._baglanti_tamam)
        self._baglanti_thread.hata.connect(self._baglanti_hatasi)
        self._baglanti_thread.start()

    def _baglanti_tamam(self, conn):
        self.conn = conn
        self.baglan_btn.setText("✅  Bağlı")
        self.baglan_btn.setStyleSheet(
            "background:#2e7d32;color:white;border-radius:6px;"
            "padding:8px 16px;font-weight:bold;font-size:12px;"
        )
        self.durum_lbl.setText("Bağlantı başarılı.")
        self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;padding:4px;")

    def _baglanti_hatasi(self, mesaj: str):
        self.baglan_btn.setEnabled(True)
        self.baglan_btn.setText("🔌  Bağlan")
        QMessageBox.critical(self, "Bağlantı Hatası",
                             f"Veritabanına bağlanılamadı:\n\n{mesaj}")

    # ── Veri çekme ────────────────────────────────────────────────────────────
    def _getir(self):
        if self.conn is None:
            QMessageBox.warning(self, "Bağlantı Yok", "Önce 'Bağlan' butonuna basın.")
            return
        bas_dt = self._parse(self.tarih_bas.text())
        bit_dt = self._parse(self.tarih_bit.text())
        if bas_dt is None or bit_dt is None:
            QMessageBox.warning(
                self, "Eksik Bilgi",
                "Lütfen geçerli başlangıç ve bitiş tarihi girin.\n\nFormat: GG.AA.YYYY",
            )
            return

        self.getir_btn.setEnabled(False)
        self.getir_btn.setText("⏳  Yükleniyor…")
        self.excel_btn.setEnabled(False)
        self.excel_btn.setStyleSheet(_BTN_PASIF)
        self.durum_lbl.setText("Veriler çekiliyor…")
        self.durum_lbl.setStyleSheet("color:#1565c0;font-size:11px;padding:4px;")

        self._veri_thread = VeriCekmeThread(
            self.conn,
            bas_dt.strftime("%d.%m.%Y"),
            bit_dt.strftime("%d.%m.%Y"),
        )
        self._veri_thread.bitti.connect(self._veri_yuklendi)
        self._veri_thread.hata.connect(self._veri_hatasi)
        self._veri_thread.start()

    def _veri_yuklendi(self, veri: list):
        self._veri = veri
        self.getir_btn.setEnabled(True)
        self.getir_btn.setText("🔍  Faturaları Getir")
        self._tabloyu_doldur(veri)

        toplam = sum(r['tutar'] for r in veri)
        self.ozet_lbl.setText(
            f"{len(veri):,} kayıt  |  Toplam tutar: {toplam:,.2f} ₺"
        )
        self.durum_lbl.setText(f"{len(veri)} satır yüklendi.")
        self.durum_lbl.setStyleSheet("color:#2e7d32;font-size:11px;padding:4px;")
        self.durum_guncelle.emit(f"Satış Faturaları: {len(veri)} satır")

        if veri:
            self.excel_btn.setEnabled(True)
            self.excel_btn.setStyleSheet(_BTN_YESIL)

    def _veri_hatasi(self, mesaj: str):
        self.getir_btn.setEnabled(True)
        self.getir_btn.setText("🔍  Faturaları Getir")
        self.durum_lbl.setText("Hata oluştu.")
        self.durum_lbl.setStyleSheet("color:#c62828;font-size:11px;padding:4px;")
        QMessageBox.critical(self, "Sorgu Hatası", f"Veriler alınamadı:\n\n{mesaj}")

    def _tabloyu_doldur(self, veri: list):
        self.tablo.setRowCount(0)
        self.tablo.setRowCount(len(veri))
        for row_idx, r in enumerate(veri):
            def _it(val, align=Qt.AlignLeft | Qt.AlignVCenter):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(align)
                return item

            sag = Qt.AlignRight | Qt.AlignVCenter
            self.tablo.setItem(row_idx, 0, _it(r['tarih']))
            self.tablo.setItem(row_idx, 1, _it(r['islem_turu']))
            self.tablo.setItem(row_idx, 2, _it(r['belge_no']))
            self.tablo.setItem(row_idx, 3, _it(r['musteri']))
            self.tablo.setItem(row_idx, 4, _it(r['stok_kodu']))
            self.tablo.setItem(row_idx, 5, _it(r['stok_adi']))
            self.tablo.setItem(row_idx, 6, _it(f"{r['miktar']:,.2f}", sag))
            self.tablo.setItem(row_idx, 7, _it(f"{r['birim_fiyat']:,.4f}", sag))
            self.tablo.setItem(row_idx, 8, _it(f"{r['tutar']:,.2f}", sag))

    # ── Arama filtresi ────────────────────────────────────────────────────────
    def _filtrele(self, metin: str):
        metin = metin.strip().lower()
        for row in range(self.tablo.rowCount()):
            if not metin:
                self.tablo.setRowHidden(row, False)
                continue
            gorunsun = any(
                (self.tablo.item(row, col) is not None
                 and metin in self.tablo.item(row, col).text().lower())
                for col in (0, 1, 2, 3, 4, 5)
            )
            self.tablo.setRowHidden(row, not gorunsun)

    # ── Excel'e aktar ─────────────────────────────────────────────────────────
    def _excel_aktar(self):
        if not self._veri:
            return
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"satis_faturalari_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel Dosyası (*.xlsx)",
        )
        if not dosya:
            return
        try:
            self._excel_olustur(dosya)
            QMessageBox.information(self, "Başarılı",
                                    f"Excel dosyası kaydedildi:\n{dosya}")
        except Exception as e:
            logging.error("Excel olusturulamadi: %s", e)
            QMessageBox.critical(self, "Hata", f"Excel oluşturulamadı:\n{e}")

    def _excel_olustur(self, dosya: str):
        import openpyxl
        from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Satış Faturaları"

        fill_baslik = PatternFill("solid", fgColor="1A237E")
        fill_cift   = PatternFill("solid", fgColor="F5F5F5")
        fill_toplam = PatternFill("solid", fgColor="E8F5E9")
        font_baslik = XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11)
        font_veri   = XFont(name="Segoe UI", size=10)
        font_toplam = XFont(bold=True, name="Segoe UI", size=11, color="1B5E20")
        kenar = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )
        hiza_sol = Alignment(horizontal="left",   vertical="center")
        hiza_sag = Alignment(horizontal="right",  vertical="center")
        hiza_ort = Alignment(horizontal="center", vertical="center")

        for col_idx, baslik in enumerate(_KOLONLAR, 1):
            h = ws.cell(row=1, column=col_idx, value=baslik)
            h.fill = fill_baslik
            h.font = font_baslik
            h.border = kenar
            h.alignment = hiza_ort

        for row_idx, r in enumerate(self._veri, 2):
            fill = fill_cift if row_idx % 2 == 0 else None
            satir = [
                r['tarih'], r['islem_turu'], r['belge_no'], r['musteri'],
                r['stok_kodu'], r['stok_adi'],
                r['miktar'], r['birim_fiyat'], r['tutar'],
            ]
            for col_idx, deger in enumerate(satir, 1):
                hucre = ws.cell(row=row_idx, column=col_idx, value=deger)
                hucre.font = font_veri
                hucre.border = kenar
                if fill:
                    hucre.fill = fill
                if col_idx == 7:
                    hucre.number_format = '#,##0.##'
                    hucre.alignment = hiza_sag
                elif col_idx == 8:
                    hucre.number_format = '#,##0.0000 "₺"'
                    hucre.alignment = hiza_sag
                elif col_idx == 9:
                    hucre.number_format = '#,##0.00 "₺"'
                    hucre.alignment = hiza_sag
                else:
                    hucre.alignment = hiza_sol

        toplam_satir = len(self._veri) + 2
        toplam_tutar = sum(r['tutar'] for r in self._veri)
        for col_idx in range(1, len(_KOLONLAR) + 1):
            hucre = ws.cell(row=toplam_satir, column=col_idx)
            hucre.fill = fill_toplam
            hucre.font = font_toplam
            hucre.border = kenar
            if col_idx == 1:
                hucre.value = f"TOPLAM  ({len(self._veri):,} kayıt)"
                hucre.alignment = hiza_sol
            elif col_idx == 9:
                hucre.value = toplam_tutar
                hucre.number_format = '#,##0.00 "₺"'
                hucre.alignment = hiza_sag

        genislikler = [12, 16, 14, 32, 14, 32, 10, 16, 16]
        for col_idx, gen in enumerate(genislikler, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = gen

        ws.freeze_panes = "A2"
        wb.save(dosya)
