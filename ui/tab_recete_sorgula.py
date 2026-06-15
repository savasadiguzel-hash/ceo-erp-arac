"""tab_recete_sorgula.py — Bileşen → Mamül Ağacı sorgulama sekmesi (8. sekme)."""
from pathlib import Path
from datetime import datetime

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QGroupBox, QGridLayout,
    QLineEdit, QPushButton, QTextEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QFileDialog,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor

from config import DB_DEFAULTS, config_kaydet
from db.baglanti import get_connection
from db.sorgular import bilesen_mamul_bul
from ui.stil import etiket

BASE = Path(__file__).resolve().parent.parent

_BTN_MAVI = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1565c0,stop:1 #1976d2);"
    "color:white;border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)
_BTN_YESIL = (
    "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2e7d32,stop:1 #43a047);"
    "color:white;border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)
_BTN_PASIF = (
    "background:#bdbdbd;color:#757575;"
    "border-radius:6px;padding:9px 20px;font-weight:bold;font-size:13px;"
)

_RENK_BULUNAMADI = QColor("#ffcdd2")  # kırmızı — geçersiz veya hiç mamülde yok
_RENK_DOGRUDAN   = QColor("#c8e6c9")  # yeşil — seviye 1 (direkt)
_RENK_DOLAYLI    = QColor("#e8eaf6")  # mavi — seviye 2+ (alt montaj üzerinden)

_TABLO_STIL = (
    "QTableWidget{border:1px solid #e8eaf6;border-radius:6px;background:white;}"
    "QHeaderView::section{background:#e8eaf6;color:#3949ab;font-weight:bold;"
    "padding:6px;border:none;border-bottom:1px solid #c5cae9;}"
)


# ── Thread'ler ─────────────────────────────────────────────────────────────────

class _BaglantiThread(QThread):
    bitti = pyqtSignal(object)
    hata  = pyqtSignal(str)

    def __init__(self, sunucu, veritabani, kullanici, sifre):
        super().__init__()
        self._s = sunucu; self._v = veritabani
        self._k = kullanici; self._p = sifre

    def run(self):
        try:
            self.bitti.emit(get_connection(self._s, self._v, self._k, self._p))
        except Exception as e:
            self.hata.emit(str(e))


class _SorguThread(QThread):
    bitti = pyqtSignal(dict)
    hata  = pyqtSignal(str)

    def __init__(self, conn, kodlar: list):
        super().__init__()
        self._conn = conn
        self._kodlar = kodlar

    def run(self):
        try:
            self.bitti.emit(bilesen_mamul_bul(self._conn, self._kodlar))
        except Exception as e:
            import traceback
            self.hata.emit(f"{type(e).__name__}: {e}\n\n{traceback.format_exc()}")


# ── Ana Sekme ─────────────────────────────────────────────────────────────────

class ReceteSorgulaTab(QWidget):
    durum_guncelle = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._conn = None
        self._son_sonuc: dict = {}
        self._t_bag = None
        self._t_sorgu = None

        ana = QVBoxLayout(self)
        ana.setContentsMargins(12, 12, 12, 12)
        ana.setSpacing(10)

        ana.addWidget(self._baglanti_grubu())
        ana.addWidget(self._giris_grubu())
        ana.addWidget(self._sonuc_grubu(), stretch=1)

    # ── UI inşa ───────────────────────────────────────────────────────────────

    def _baglanti_grubu(self) -> QGroupBox:
        grp = QGroupBox("Veritabanı Bağlantısı")
        g = QGridLayout(grp)
        g.setSpacing(8)

        g.addWidget(etiket("Sunucu:"), 0, 0)
        self._sunucu = QLineEdit(DB_DEFAULTS["sunucu"])
        g.addWidget(self._sunucu, 0, 1)

        g.addWidget(etiket("Veritabanı:"), 0, 2)
        self._veritabani = QLineEdit(DB_DEFAULTS["veritabani"])
        g.addWidget(self._veritabani, 0, 3)

        g.addWidget(etiket("Kullanıcı:"), 1, 0)
        self._kullanici = QLineEdit(DB_DEFAULTS["kullanici"])
        g.addWidget(self._kullanici, 1, 1)

        g.addWidget(etiket("Şifre:"), 1, 2)
        self._sifre = QLineEdit(DB_DEFAULTS["sifre"])
        self._sifre.setEchoMode(QLineEdit.Password)
        g.addWidget(self._sifre, 1, 3)

        self._bag_btn = QPushButton("🔌  Bağlan")
        self._bag_btn.setStyleSheet(_BTN_MAVI)
        self._bag_btn.setFixedHeight(36)
        self._bag_btn.clicked.connect(self._baglan)
        g.addWidget(self._bag_btn, 0, 4, 2, 1)

        self._bag_durum = QLabel("")
        self._bag_durum.setStyleSheet("color:#757575;font-size:11px;")
        g.addWidget(self._bag_durum, 2, 0, 1, 5)

        return grp

    def _giris_grubu(self) -> QGroupBox:
        self._giris_grp = QGroupBox("Bileşen Kodları")
        self._giris_grp.setEnabled(False)
        gl = QHBoxLayout(self._giris_grp)
        gl.setSpacing(14)

        # Sol — metin alanı
        sol = QVBoxLayout()
        sol.addWidget(QLabel("Her satıra bir stok kodu girin (veya Excel'den yükleyin):"))
        self._kod_metin = QTextEdit()
        self._kod_metin.setPlaceholderText("KOD001\nKOD002\nKOD003")
        self._kod_metin.setMaximumHeight(140)
        self._kod_metin.setStyleSheet(
            "border:1.5px solid #9fa8da;border-radius:6px;background:white;"
            "font-family:'Consolas','Courier New';font-size:12px;padding:6px;"
        )
        sol.addWidget(self._kod_metin)
        gl.addLayout(sol, stretch=1)

        # Sağ — butonlar
        sag = QVBoxLayout()
        sag.setAlignment(Qt.AlignTop)
        sag.setSpacing(8)

        self._excel_yukle_btn = QPushButton("📥  Excel'den Yükle")
        self._excel_yukle_btn.setStyleSheet(_BTN_MAVI)
        self._excel_yukle_btn.setFixedHeight(36)
        self._excel_yukle_btn.setMinimumWidth(175)
        self._excel_yukle_btn.clicked.connect(self._excel_yukle)
        sag.addWidget(self._excel_yukle_btn)

        self._temizle_btn = QPushButton("🗑  Temizle")
        self._temizle_btn.setStyleSheet(
            "background:#ef9a9a;color:white;border-radius:6px;"
            "padding:9px 20px;font-weight:bold;font-size:13px;"
        )
        self._temizle_btn.setFixedHeight(36)
        self._temizle_btn.setMinimumWidth(175)
        self._temizle_btn.clicked.connect(lambda: self._kod_metin.clear())
        sag.addWidget(self._temizle_btn)

        sag.addStretch()

        self._sorgula_btn = QPushButton("🔍  Sorgula")
        self._sorgula_btn.setStyleSheet(_BTN_YESIL)
        self._sorgula_btn.setFixedHeight(36)
        self._sorgula_btn.setMinimumWidth(175)
        self._sorgula_btn.clicked.connect(self._sorgula)
        sag.addWidget(self._sorgula_btn)

        gl.addLayout(sag)
        return self._giris_grp

    def _sonuc_grubu(self) -> QGroupBox:
        grp = QGroupBox("Sonuçlar")
        sl = QVBoxLayout(grp)
        sl.setSpacing(8)

        # Renk açıklaması
        acik = QHBoxLayout()
        for hex_renk, metin in [
            ("#c8e6c9", "Direkt bileşen (Seviye 1)"),
            ("#e8eaf6", "Dolaylı bileşen (Alt montaj üzerinden)"),
            ("#ffcdd2", "Hiçbir mamülde yok / Kod geçersiz"),
        ]:
            kare = QLabel("■")
            kare.setStyleSheet(f"color:{hex_renk};font-size:20px;")
            lbl = QLabel(metin)
            lbl.setStyleSheet("color:#555;font-size:11px;")
            acik.addWidget(kare)
            acik.addWidget(lbl)
            acik.addSpacing(16)
        acik.addStretch()
        sl.addLayout(acik)

        self._tablo = QTableWidget(0, 5)
        self._tablo.setHorizontalHeaderLabels([
            "Aranan Kod", "Mamül Kodu", "Mamül Adı", "Seviye", "Bağlantı Türü",
        ])
        hdr = self._tablo.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self._tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tablo.setSelectionBehavior(QTableWidget.SelectRows)
        self._tablo.verticalHeader().setVisible(False)
        self._tablo.setStyleSheet(_TABLO_STIL)
        sl.addWidget(self._tablo)

        btn_satir = QHBoxLayout()
        btn_satir.addStretch()
        self._excel_kaydet_btn = QPushButton("📤  Excel'e Aktar")
        self._excel_kaydet_btn.setStyleSheet(_BTN_YESIL)
        self._excel_kaydet_btn.setFixedHeight(34)
        self._excel_kaydet_btn.setEnabled(False)
        self._excel_kaydet_btn.clicked.connect(self._excel_kaydet)
        btn_satir.addWidget(self._excel_kaydet_btn)
        sl.addLayout(btn_satir)

        return grp

    # ── Bağlantı ──────────────────────────────────────────────────────────────

    def _baglan(self):
        self._bag_btn.setEnabled(False)
        self._bag_btn.setStyleSheet(_BTN_PASIF)
        self._bag_durum.setText("Bağlanılıyor…")
        self._t_bag = _BaglantiThread(
            self._sunucu.text(), self._veritabani.text(),
            self._kullanici.text(), self._sifre.text(),
        )
        self._t_bag.bitti.connect(self._baglanti_ok)
        self._t_bag.hata.connect(self._baglanti_hata)
        self._t_bag.start()

    def _baglanti_ok(self, conn):
        self._conn = conn
        config_kaydet(
            self._sunucu.text(), self._veritabani.text(),
            self._kullanici.text(), self._sifre.text(),
        )
        self._bag_durum.setText("✅ Bağlantı kuruldu.")
        self._bag_btn.setText("🔄  Yeniden Bağlan")
        self._bag_btn.setStyleSheet(_BTN_MAVI)
        self._bag_btn.setEnabled(True)
        self._giris_grp.setEnabled(True)
        self.durum_guncelle.emit("Reçete Sorgula: Bağlantı kuruldu.")

    def _baglanti_hata(self, msg: str):
        self._bag_btn.setEnabled(True)
        self._bag_btn.setStyleSheet(_BTN_MAVI)
        self._bag_durum.setText("❌ Bağlantı hatası.")
        QMessageBox.critical(self, "Bağlantı Hatası", f"Veritabanına bağlanılamadı:\n\n{msg}")
        self.durum_guncelle.emit("Reçete Sorgula: Bağlantı başarısız.")

    # ── Excel İçe Aktar ───────────────────────────────────────────────────────

    def _excel_yukle(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", str(BASE),
            "Excel Dosyaları (*.xlsx *.xls)",
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            kodlar = []
            for row in ws.iter_rows(values_only=True):
                if row and row[0] is not None:
                    val = str(row[0]).strip()
                    # Başlık satırını atla (sadece "Kod", "KOD" gibi metin ise)
                    if val and val.upper() not in ("KOD", "STOK KODU", "CODE"):
                        kodlar.append(val)
            wb.close()
            if not kodlar:
                QMessageBox.information(self, "Boş Dosya", "Excel'de okunabilecek kod bulunamadı.")
                return
            mevcut = self._kod_metin.toPlainText().strip()
            ek = "\n".join(kodlar)
            self._kod_metin.setPlainText((mevcut + "\n" + ek).strip())
            self.durum_guncelle.emit(f"{len(kodlar)} kod Excel'den yüklendi.")
        except Exception as e:
            QMessageBox.critical(self, "Excel Hatası", f"Dosya okunamadı:\n{e}")

    # ── Sorgula ───────────────────────────────────────────────────────────────

    def _sorgula(self):
        if self._conn is None:
            return
        metin = self._kod_metin.toPlainText().strip()
        if not metin:
            QMessageBox.warning(self, "Boş Liste", "Lütfen en az bir stok kodu girin.")
            return

        # Tekrar eden kodları koru ama sıralamayla teke indir
        tum = [k.strip() for k in metin.splitlines() if k.strip()]
        kodlar = list(dict.fromkeys(tum))  # sıra korunarak tekilleştir

        self._sorgula_btn.setEnabled(False)
        self._sorgula_btn.setStyleSheet(_BTN_PASIF)
        self._sorgula_btn.setText("🔍  Sorgulanıyor…")
        self.durum_guncelle.emit(f"{len(kodlar)} kod sorgulanıyor…")

        self._t_sorgu = _SorguThread(self._conn, kodlar)
        self._t_sorgu.bitti.connect(self._sonuclari_goster)
        self._t_sorgu.hata.connect(self._sorgu_hata)
        self._t_sorgu.start()

    def _sorgu_hata(self, msg: str):
        self._sorgula_btn.setEnabled(True)
        self._sorgula_btn.setStyleSheet(_BTN_YESIL)
        self._sorgula_btn.setText("🔍  Sorgula")
        QMessageBox.critical(self, "Sorgu Hatası", msg)
        self.durum_guncelle.emit("Reçete Sorgula: Sorgu hatası.")

    def _sonuclari_goster(self, sonuc: dict):
        self._son_sonuc = sonuc
        self._sorgula_btn.setEnabled(True)
        self._sorgula_btn.setStyleSheet(_BTN_YESIL)
        self._sorgula_btn.setText("🔍  Sorgula")
        self._tablo.setRowCount(0)

        for giris_kodu, bilgi in sonuc.items():
            if not bilgi["kod_gecerli"]:
                self._satir_ekle(
                    giris_kodu, "—", "Kod veritabanında bulunamadı", "—",
                    "GEÇERSİZ KOD", _RENK_BULUNAMADI,
                )
            elif not bilgi["bulundu"]:
                self._satir_ekle(
                    giris_kodu, "—", "Hiçbir mamül ağacında kullanılmıyor", "—",
                    "BAĞLANTISIZ", _RENK_BULUNAMADI,
                )
            else:
                for m in bilgi["mamullar"]:
                    tur = "Direkt" if m["seviye"] == 1 else f"Dolaylı (Sev. {m['seviye']})"
                    renk = _RENK_DOGRUDAN if m["seviye"] == 1 else _RENK_DOLAYLI
                    self._satir_ekle(
                        giris_kodu, m["kodu"], m["adi"], str(m["seviye"]), tur, renk,
                    )

        bulunan  = sum(1 for b in sonuc.values() if b.get("bulundu"))
        gecersiz = sum(1 for b in sonuc.values() if not b.get("kod_gecerli"))
        toplam   = len(sonuc)
        self._excel_kaydet_btn.setEnabled(bool(sonuc))
        self.durum_guncelle.emit(
            f"Tamamlandı — {toplam} kod: {bulunan} mamülde bulundu, "
            f"{toplam - bulunan - gecersiz} bağlantısız, {gecersiz} geçersiz."
        )

    def _satir_ekle(self, kod, mamul_kodu, mamul_adi, seviye, tur, renk: QColor):
        r = self._tablo.rowCount()
        self._tablo.insertRow(r)
        for c, deger in enumerate([kod, mamul_kodu, mamul_adi, seviye, tur]):
            item = QTableWidgetItem(str(deger))
            item.setBackground(renk)
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            self._tablo.setItem(r, c, item)

    # ── Excel Dışa Aktar ──────────────────────────────────────────────────────

    def _excel_kaydet(self):
        if not self._son_sonuc:
            return
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            str(BASE / f"recete_sorgu_{datetime.now():%Y%m%d_%H%M%S}.xlsx"),
            "Excel Dosyaları (*.xlsx)",
        )
        if not dosya:
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font as XFont, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reçete Sorgu"

            baslik_fill = PatternFill("solid", fgColor="1A237E")
            baslik_font = XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11)
            orta = Alignment(horizontal="center", vertical="center")

            for c, baslik in enumerate(
                ["Aranan Kod", "Mamül Kodu", "Mamül Adı", "Seviye", "Bağlantı Türü"], 1
            ):
                cell = ws.cell(1, c, baslik)
                cell.fill = baslik_fill
                cell.font = baslik_font
                cell.alignment = orta

            fill_yesil   = PatternFill("solid", fgColor="C8E6C9")
            fill_mavi    = PatternFill("solid", fgColor="E8EAF6")
            fill_kirmizi = PatternFill("solid", fgColor="FFCDD2")

            for giris_kodu, bilgi in self._son_sonuc.items():
                if not bilgi["kod_gecerli"]:
                    ws.append([giris_kodu, "—", "Kod veritabanında bulunamadı", "—", "GEÇERSİZ KOD"])
                    fill = fill_kirmizi
                    for c in range(1, 6):
                        ws.cell(ws.max_row, c).fill = fill
                elif not bilgi["bulundu"]:
                    ws.append([giris_kodu, "—", "Hiçbir mamül ağacında kullanılmıyor", "—", "BAĞLANTISIZ"])
                    for c in range(1, 6):
                        ws.cell(ws.max_row, c).fill = fill_kirmizi
                else:
                    for m in bilgi["mamullar"]:
                        tur = "Direkt" if m["seviye"] == 1 else f"Dolaylı (Sev. {m['seviye']})"
                        ws.append([giris_kodu, m["kodu"], m["adi"], m["seviye"], tur])
                        fill = fill_yesil if m["seviye"] == 1 else fill_mavi
                        for c in range(1, 6):
                            ws.cell(ws.max_row, c).fill = fill

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 52
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 28

            wb.save(dosya)
            self.durum_guncelle.emit(f"Excel kaydedildi: {Path(dosya).name}")
        except Exception as e:
            QMessageBox.critical(self, "Kayıt Hatası", f"Excel kaydedilemedi:\n{e}")
