"""tab_stok_hazirla.py — Stok Hazırlık sekmesi (9. sekme).

Hammadde / Yarımamül / Mamül / Reçete kodlarını QTreeWidget ağacında
organize eder; CEO ERP'de kontrol eder, eksik kartları açar, BOM bağlantılarını yazar.
"""
from __future__ import annotations

import copy
from pathlib import Path
from typing import Optional

import pyodbc

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QPlainTextEdit, QTreeWidget, QTreeWidgetItem, QHeaderView,
    QMessageBox, QFileDialog, QComboBox, QStyledItemDelegate,
    QAbstractItemView, QInputDialog,
    QDialog, QLineEdit, QListWidget, QListWidgetItem, QDialogButtonBox, QMenu,
    QGraphicsView, QGraphicsScene, QGraphicsPathItem,
    QGraphicsTextItem,
)
from PyQt5.QtCore import Qt, QThread, QTimer, pyqtSignal, QRectF
from PyQt5.QtGui import QColor, QTextCursor, QPen, QBrush, QPainterPath, QPainter

from config import DB_DEFAULTS

try:
    from sw import erp_handler as _erp
    ERP_OK = True
except Exception as _e:
    _erp = None          # type: ignore[assignment]
    ERP_OK = False
    _ERP_HATA = str(_e)


def _baglan() -> pyodbc.Connection:
    """config.json kimlik bilgileriyle CEO ERP'ye bağlan (CEO_SQL_CONN gerekmez)."""
    cfg = DB_DEFAULTS
    dsn = (
        "DRIVER={SQL Server};"
        "SERVER=%(sunucu)s;DATABASE=%(veritabani)s;"
        "UID=%(kullanici)s;PWD=%(sifre)s;"
        "TrustServerCertificate=yes;"
    ) % cfg
    return pyodbc.connect(dsn, autocommit=False, timeout=10)


BASE = Path(__file__).resolve().parent.parent

_BTN_MAVI    = ("background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                "stop:0 #1565c0,stop:1 #1976d2);"
                "color:white;border-radius:6px;padding:7px 16px;"
                "font-weight:bold;font-size:12px;")
_BTN_YESIL   = ("background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                "stop:0 #2e7d32,stop:1 #43a047);"
                "color:white;border-radius:6px;padding:7px 16px;"
                "font-weight:bold;font-size:12px;")
_BTN_TURUNCU = ("background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                "stop:0 #e65100,stop:1 #f57c00);"
                "color:white;border-radius:6px;padding:7px 16px;"
                "font-weight:bold;font-size:12px;")
_BTN_GERI    = ("background:#eceff1;color:#37474f;border-radius:6px;"
                "padding:7px 16px;font-size:12px;")
_BTN_PASIF   = ("background:#bdbdbd;color:#757575;border-radius:6px;"
                "padding:7px 16px;font-size:12px;")
_LOG_STIL    = ("background:#1e1e1e;color:#d4d4d4;"
                "font-family:Consolas,monospace;font-size:11px;")
_AGAC_STIL   = (
    "QTreeWidget{border:1px solid #e8eaf6;border-radius:6px;background:white;"
    "outline:0;}"
    "QHeaderView::section{background:#e8eaf6;color:#3949ab;"
    "font-weight:bold;padding:6px;border:none;border-bottom:1px solid #c5cae9;}"
    "QTreeWidget::item{padding:5px 2px;min-height:22px;}"
    "QTreeWidget::item:selected{background:#bbdefb;color:#0d47a1;}"
    "QTreeWidget::item:selected:active{background:#bbdefb;color:#0d47a1;}"
    "QTreeWidget::item:selected:!active{background:#e3f2fd;color:#1a237e;}"
)

_TIP_SECENEKLER = ["Hammadde", "Yarımamül", "Mamül", "Reçete", "Masraf"]
_ADET_GUID      = "4acc21e3-3140-4863-922d-a0b3d35de8c1"

_ROLE_GIRDI_ID   = Qt.UserRole + 1   # int: UretimReceteHatPlaniGirdi.Id veya None
_ROLE_ORG_MIKTAR = Qt.UserRole + 2   # float: yüklendiğindeki orijinal Miktar
_BASLIK_ATLA    = {"tip", "kod", "stok kodu", "stok adi", "stok adı",
                   "ad", "adi", "adı", "type", "code", "name", "seviye"}

_C_MEVCUT = QColor("#c8e6c9")
_C_YENI   = QColor("#ffe0b2")
_C_TAMAM  = QColor("#a5d6a7")
_C_HATA   = QColor("#ffcdd2")
_C_KURU   = QColor("#fff9c4")
_C_BOSTA  = QColor("#f5f5f5")

# Sütun sabitleri
_COL_TIP    = 0
_COL_KOD    = 1
_COL_ADI    = 2
_COL_ADI2   = 3
_COL_MIKTAR = 4
_COL_BIRIM  = 5
_COL_DURUM  = 6


# ── Delegate'ler ──────────────────────────────────────────────────────────────

class _TipDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        cb = QComboBox(parent)
        cb.addItems(_TIP_SECENEKLER)
        return cb

    def setEditorData(self, editor, index):
        val = index.data() or _TIP_SECENEKLER[0]
        editor.setCurrentText(val if val in _TIP_SECENEKLER else _TIP_SECENEKLER[0])

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)


class _MetinDelegate(QStyledItemDelegate):
    """Metin sütunları için delegate — global QLineEdit stilini ezip okunabilir editör açar."""
    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if editor is not None:
            editor.setStyleSheet(
                "QLineEdit{background:white;color:#212121;"
                "border:1px solid #3f51b5;border-radius:0px;"
                "padding:0px 3px;font-size:12px;"
                "selection-background-color:#bbdefb;selection-color:#0d47a1;}"
            )
        return editor

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class _BirimDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._secenekler: list[str] = ["ADET"]

    def set_birimler(self, birimler: list[str]):
        self._secenekler = birimler if birimler else ["ADET"]

    def createEditor(self, parent, option, index):
        cb = QComboBox(parent)
        cb.addItems(self._secenekler)
        return cb

    def setEditorData(self, editor, index):
        val = index.data() or "ADET"
        idx = editor.findText(val)
        editor.setCurrentIndex(idx if idx >= 0 else 0)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)


class _NoEditDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return None


# ── Thread yardımcısı: ağaç verisini pre-order DFS ile düzleştir ─────────────

def _dfs(nodes: list) -> list:
    """nested [{...children:[...]}, ...] → [(node, depth), ...] pre-order."""
    result = []
    def _visit(lst, depth):
        for node in lst:
            result.append((node, depth))
            _visit(node.get("children", []), depth + 1)
    _visit(nodes, 0)
    return result


# ── Thread: Birimleri Yükle ───────────────────────────────────────────────────

class _BirimleriYukleThread(QThread):
    bitti = pyqtSignal(dict)   # {adi: guid}

    def run(self):
        try:
            conn = _baglan()
            cur  = conn.cursor()
            cur.execute("SELECT Id, Adi FROM StokBirim WHERE Aktif = 1 ORDER BY Adi")
            birimler = {str(r[1]): str(r[0]) for r in cur.fetchall()}
            conn.close()
            self.bitti.emit(birimler or {"ADET": _ADET_GUID})
        except Exception:
            self.bitti.emit({"ADET": _ADET_GUID})


# ── Thread: Tüm reçete listesini çek ─────────────────────────────────────────

class _TumRecetelerThread(QThread):
    bitti = pyqtSignal(list)   # [(kod, tanim), ...]
    hata  = pyqtSignal(str)

    def run(self):
        try:
            conn = _baglan()
            cur = conn.cursor()
            cur.execute(
                "SELECT Kodu, ISNULL(Tanim,'') FROM UretimRecete "
                "WHERE KullanimDisi IS NULL OR KullanimDisi = 0 "
                "ORDER BY Kodu"
            )
            rows = [(str(r[0]), str(r[1])) for r in cur.fetchall()]
            cur.close()
            conn.close()
            self.bitti.emit(rows)
        except Exception as e:
            self.hata.emit(str(e))


# ── Dialog: Reçete seçim penceresi ───────────────────────────────────────────

class _ReceteSecDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Reçete Seç")
        self.resize(560, 480)
        self.setModal(True)
        self._secilen_kod: str = ""

        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(12, 12, 12, 12)

        # Arama kutusu
        self._ara = QLineEdit()
        self._ara.setPlaceholderText("Kod veya ada göre filtrele…")
        self._ara.setStyleSheet(
            "QLineEdit{border:1.5px solid #9fa8da;border-radius:6px;"
            "padding:6px 10px;background:white;font-size:12px;}"
            "QLineEdit:focus{border-color:#3f51b5;}"
        )
        self._ara.textChanged.connect(self._filtrele)
        lay.addWidget(self._ara)

        # Yükleniyor etiketi
        self._yukleniyor = QLabel("Reçeteler yükleniyor…")
        self._yukleniyor.setStyleSheet("color:#757575;font-size:12px;padding:4px 0;")
        lay.addWidget(self._yukleniyor)

        # Liste
        self._liste = QListWidget()
        self._liste.setStyleSheet(
            "QListWidget{border:1px solid #e8eaf6;border-radius:6px;background:white;}"
            "QListWidget::item{padding:5px 8px;font-size:12px;}"
            "QListWidget::item:selected{background:#bbdefb;color:#0d47a1;}"
            "QListWidget::item:hover{background:#e3f2fd;}"
        )
        self._liste.setFont(self._liste.font())
        self._liste.itemDoubleClicked.connect(self._kabul_et)
        lay.addWidget(self._liste, stretch=1)

        # Sayaç etiketi
        self._sayac = QLabel("")
        self._sayac.setStyleSheet("color:#9e9e9e;font-size:11px;")
        lay.addWidget(self._sayac)

        # Butonlar
        self._butonlar = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._butonlar.button(QDialogButtonBox.Ok).setText("Seç")
        self._butonlar.button(QDialogButtonBox.Cancel).setText("İptal")
        self._butonlar.button(QDialogButtonBox.Ok).setEnabled(False)
        self._butonlar.button(QDialogButtonBox.Ok).setStyleSheet(_BTN_MAVI)
        self._butonlar.accepted.connect(self._kabul_et)
        self._butonlar.rejected.connect(self.reject)
        lay.addWidget(self._butonlar)

        self._liste.currentItemChanged.connect(self._secim_degisti)

        self._tum_receteler: list[tuple[str, str]] = []

    def doldur(self, receteler: list):
        """Thread sonucu gelince çağrılır."""
        self._tum_receteler = receteler
        self._yukleniyor.hide()
        self._filtrele(self._ara.text())
        self._ara.setFocus()

    def _filtrele(self, metin: str):
        filtre = metin.strip().lower()
        self._liste.clear()
        for kod, tanim in self._tum_receteler:
            if not filtre or filtre in kod.lower() or filtre in tanim.lower():
                item = QListWidgetItem("%-30s  %s" % (kod, tanim))
                item.setData(Qt.UserRole, kod)
                self._liste.addItem(item)
        n = self._liste.count()
        toplam = len(self._tum_receteler)
        self._sayac.setText("%d / %d reçete" % (n, toplam))
        if n > 0:
            self._liste.setCurrentRow(0)

    def _secim_degisti(self, current, _prev):
        self._butonlar.button(QDialogButtonBox.Ok).setEnabled(current is not None)

    def _kabul_et(self):
        item = self._liste.currentItem()
        if item:
            self._secilen_kod = item.data(Qt.UserRole)
            self.accept()

    def secilen_kod(self) -> str:
        return self._secilen_kod


class _EbeveynSecDialog(QDialog):
    """BOM diyagramında düğüm taşıma için yeni ebeveyn seçim dialog'u."""

    def __init__(self, hareketli_node: dict, hedefler: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Yeni Ebeveyn Seç — %s" % hareketli_node.get("kod", ""))
        self.setMinimumSize(480, 400)
        self._secilen: Optional[str] = None

        # İlk seçenek: kök'e taşı
        self._hedefler = [{"kod": None, "adi": "[ Kök seviyesine taşı ]", "tip": ""}] + hedefler

        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(12, 12, 12, 12)

        self._ara = QLineEdit()
        self._ara.setPlaceholderText("Kod veya ada göre filtrele…")
        self._ara.setStyleSheet(
            "QLineEdit{border:1.5px solid #9fa8da;border-radius:6px;"
            "padding:6px 10px;background:white;font-size:12px;}"
            "QLineEdit:focus{border-color:#3f51b5;}"
        )
        self._ara.textChanged.connect(self._filtrele)
        lay.addWidget(self._ara)

        self._liste = QListWidget()
        self._liste.setStyleSheet(
            "QListWidget{border:1px solid #e8eaf6;border-radius:6px;background:white;}"
            "QListWidget::item{padding:5px 8px;font-size:12px;}"
            "QListWidget::item:selected{background:#bbdefb;color:#0d47a1;}"
            "QListWidget::item:hover{background:#e3f2fd;}"
        )
        self._liste.itemDoubleClicked.connect(self.accept)
        lay.addWidget(self._liste, stretch=1)

        self._sayac = QLabel("")
        self._sayac.setStyleSheet("color:#9e9e9e;font-size:11px;")
        lay.addWidget(self._sayac)

        btn = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn.button(QDialogButtonBox.Ok).setText("Seç")
        btn.button(QDialogButtonBox.Cancel).setText("İptal")
        btn.button(QDialogButtonBox.Ok).setStyleSheet(_BTN_MAVI)
        btn.accepted.connect(self.accept)
        btn.rejected.connect(self.reject)
        lay.addWidget(btn)

        self._filtrele("")
        self._ara.setFocus()

    def _filtrele(self, metin: str):
        self._liste.clear()
        metin = metin.strip().lower()
        for h in self._hedefler:
            if h["kod"] is None:
                etiket = "[ Kök seviyesine taşı ]"
            else:
                etiket = "%s — %s (%s)" % (h["kod"], h["adi"], h["tip"])
            if not metin or metin in etiket.lower():
                it = QListWidgetItem(etiket)
                it.setData(Qt.UserRole, h["kod"])
                self._liste.addItem(it)
        n = self._liste.count()
        self._sayac.setText("%d hedef" % n)
        if n:
            self._liste.setCurrentRow(0)

    def accept(self):
        cur = self._liste.currentItem()
        if cur:
            self._secilen = cur.data(Qt.UserRole)
        super().accept()

    def secilen_kod(self) -> Optional[str]:
        return self._secilen


# ── Thread: Reçete Yükle ─────────────────────────────────────────────────────

class _ReceteYukleThread(QThread):
    bitti = pyqtSignal(object)   # nested dict veya None
    hata  = pyqtSignal(str)

    def __init__(self, mamul_kod: str, birimler: dict):
        super().__init__()
        self._kod = mamul_kod.strip()
        # {guid_normalize: adi} — ters harita
        self._guid_adi = {
            v.lower().strip("{}"): k
            for k, v in birimler.items()
        }

    def _birim_coz(self, node: dict):
        """birim_guid → birim_adi dönüşümü (recursive)."""
        guid = str(node.get("birim_guid", "")).lower().strip("{}")
        node["birim_adi"] = self._guid_adi.get(guid, "ADET")
        for child in node.get("children", []):
            self._birim_coz(child)

    def run(self):
        try:
            from db.sorgular import recete_yukle
            conn = _baglan()
            sonuc = recete_yukle(conn, self._kod)
            conn.close()
            if sonuc:
                self._birim_coz(sonuc)
            self.bitti.emit(sonuc)
        except Exception as e:
            self.hata.emit(str(e))


# ── Thread: Kontrol (kuru) ────────────────────────────────────────────────────

class _KontrolThread(QThread):
    satir_guncelle = pyqtSignal(int, str, object)   # uid, metin, QColor
    log_mesaj      = pyqtSignal(str)
    bitti          = pyqtSignal(dict)
    hata           = pyqtSignal(str)

    def __init__(self, nodes: list):
        super().__init__()
        self._nodes = nodes

    def run(self):
        try:
            conn = _baglan()
            cur  = conn.cursor()
            flat = _dfs(self._nodes)
            ozet = {"mevcut": 0, "yeni": 0, "bos": 0}
            self.log_mesaj.emit("[Kontrol] Başladı — %d öğe" % len(flat))

            for node, depth in flat:
                uid = node["uid"]
                kod = (node.get("kod") or "").strip()
                if not kod:
                    self.satir_guncelle.emit(uid, "—", _C_BOSTA)
                    ozet["bos"] += 1
                    continue

                tip = node.get("tip", "")
                if tip == "Masraf":
                    kart_id = _erp.masraf_karti_var_mi(cur, kod)
                else:
                    kart_id = _erp.kod_var_mi(cur, kod)

                if kart_id is not None:
                    self.satir_guncelle.emit(uid, "Mevcut", _C_MEVCUT)
                    self.log_mesaj.emit("  [MEVCUT] %s (Id=%d)" % (kod, kart_id))
                    ozet["mevcut"] += 1
                else:
                    self.satir_guncelle.emit(uid, "Yeni açılacak", _C_YENI)
                    self.log_mesaj.emit("  [YENİ]   %s" % kod)
                    ozet["yeni"] += 1

                # BOM kontrolü: child öğeler için parent reçetesi var mı?
                if depth > 0 and node.get("parent_kod"):
                    p_kod = node["parent_kod"]
                    cur.execute(
                        "SELECT TOP 1 Id FROM UretimRecete WHERE Kodu = ? "
                        "AND (KullanimDisi IS NULL OR KullanimDisi = 0)", p_kod)
                    r = cur.fetchone()
                    if r:
                        self.log_mesaj.emit(
                            "  [BOM-VAR]  %s → %s reçetesi mevcut" % (kod, p_kod))
                    else:
                        self.log_mesaj.emit(
                            "  [BOM-YENİ] %s → %s reçetesi yok, oluşturulacak" % (kod, p_kod))

            conn.close()
            self.log_mesaj.emit(
                "[Kontrol] Tamamlandı — mevcut: %d · yeni: %d" % (
                    ozet["mevcut"], ozet["yeni"]))
            self.bitti.emit(ozet)
        except Exception as e:
            self.hata.emit(str(e))


# ── Thread: CANLI Aktar ───────────────────────────────────────────────────────

class _AktarThread(QThread):
    satir_guncelle = pyqtSignal(int, str, object)
    log_mesaj      = pyqtSignal(str)
    bitti          = pyqtSignal(dict)
    hata           = pyqtSignal(str)

    def __init__(self, nodes: list, live: bool, birimler: dict,
                 deleted_girdi_ids: set | None = None):
        super().__init__()
        self._nodes             = nodes
        self._live              = live
        self._birimler          = birimler
        self._deleted_girdi_ids = deleted_girdi_ids or set()

    def run(self):
        try:
            conn = _baglan()
            mod  = "CANLI YAZMA" if self._live else "KURU ÇALIŞMA"
            flat = _dfs(self._nodes)
            self.log_mesaj.emit("[Aktar] %s — %d öğe" % (mod, len(flat)))
            ozet = {"olusturuldu": 0, "atlandi": 0, "olusturulacak": 0, "hata": 0, "bom": 0}

            # ── Faz 1: Stok / Masraf Kartları ───────────────────────────────
            kod_id:       dict[str, int] = {}   # stok kod → StokKarti.Id
            masraf_kod_id: dict[str, int] = {}  # masraf kod → StokMasrafKarti.Id

            for node, _depth in flat:
                uid        = node["uid"]
                kod        = (node.get("kod") or "").strip()
                adi        = (node.get("adi") or "").strip()
                adi2       = (node.get("adi2") or "").strip()
                tip        = node.get("tip", "")
                birim_adi  = node.get("birim") or "ADET"
                birim_guid = self._birimler.get(birim_adi, _ADET_GUID)

                if not kod:
                    self.satir_guncelle.emit(uid, "—", _C_BOSTA)
                    continue

                if tip == "Masraf":
                    r = _erp.masraf_karti_ac(
                        kod, adi, live=self._live,
                        birim_guid=birim_guid, conn=conn)
                    if r.get("id"):
                        masraf_kod_id[kod] = r["id"]
                else:
                    r = _erp.stok_karti_ac(
                        kod, adi, live=self._live,
                        adi2=adi2, birim_guid=birim_guid, conn=conn)
                    if r.get("id"):
                        kod_id[kod] = r["id"]

                d = r["durum"]
                ozet[d] = ozet.get(d, 0) + 1

                if d == "atlandi":
                    self.satir_guncelle.emit(uid, "Mevcut", _C_MEVCUT)
                    self.log_mesaj.emit("  [MEVCUT]      %s" % kod)
                elif d == "olusturuldu":
                    self.satir_guncelle.emit(uid, "Oluşturuldu ✓", _C_TAMAM)
                    self.log_mesaj.emit("  [OLUŞTURULDU] %s (Id=%s)" % (kod, r.get("id")))
                elif d == "olusturulacak":
                    self.satir_guncelle.emit(uid, "Açılacak (Kuru)", _C_KURU)
                    self.log_mesaj.emit("  [KURU]        %s" % kod)
                else:
                    self.satir_guncelle.emit(uid, "Hata", _C_HATA)
                    self.log_mesaj.emit("  [HATA]        %s — %s" % (kod, r.get("mesaj", "")))

            # ── Faz 2a: Yüklenen satirlarda Miktar değişikliği → UPDATE ─────
            for node, _depth in flat:
                girdi_id   = node.get("girdi_id")
                org_miktar = node.get("org_miktar")
                if girdi_id is None or org_miktar is None:
                    continue
                try:
                    cur_mik = float(node.get("miktar") or "1")
                    org_mik = float(org_miktar)
                except ValueError:
                    continue
                if abs(cur_mik - org_mik) > 0.0001:
                    r = _erp.recete_satiri_guncelle(
                        girdi_id, cur_mik, live=self._live, conn=conn)
                    tag = "MIKTAR-OK" if r["durum"] == "guncellendi" else (
                          "MIKTAR-KURU" if r["durum"] == "olusturulacak" else "MIKTAR-HATA")
                    if r["durum"] in ("guncellendi", "olusturulacak"):
                        ozet["guncellendi"] = ozet.get("guncellendi", 0) + 1
                    self.log_mesaj.emit(
                        "  [%s] %s: %g → %g" % (tag, node.get("kod", "?"), org_mik, cur_mik))

            # ── Faz 2b: YENİ öğeler için BOM bağlantısı ─────────────────────
            bom_ciftler = [
                (n, d) for n, d in flat
                if d > 0 and n.get("parent_kod") and n.get("girdi_id") is None
            ]
            if bom_ciftler:
                self.log_mesaj.emit("[BOM] %d yeni bağlantı işlenecek" % len(bom_ciftler))
                for node, _depth in bom_ciftler:
                    child_kod  = (node.get("kod") or "").strip()
                    parent_kod = node.get("parent_kod", "")
                    parent_adi = node.get("parent_adi", "")
                    parent_tip = node.get("parent_tip", "")
                    child_tip  = node.get("tip", "")
                    hat_tipi   = 1 if parent_tip == "Mamül" else 2
                    try:
                        miktar = float(node.get("miktar") or "1")
                    except ValueError:
                        miktar = 1.0

                    if not child_kod or not parent_kod:
                        continue

                    parent_id = kod_id.get(parent_kod)

                    if child_tip == "Masraf":
                        child_id = masraf_kod_id.get(child_kod)
                        if child_id is None or parent_id is None:
                            self.log_mesaj.emit(
                                "  [BOM-ATLA] %s → %s (Id bilinmiyor)" % (child_kod, parent_kod))
                            continue
                        r = _erp.recete_masraf_bagla(
                            parent_kod, parent_adi, parent_id, child_id, miktar,
                            hat_tipi=hat_tipi, live=self._live, conn=conn)
                    else:
                        child_id = kod_id.get(child_kod)
                        if child_id is None or parent_id is None:
                            self.log_mesaj.emit(
                                "  [BOM-ATLA] %s → %s (Id bilinmiyor)" % (child_kod, parent_kod))
                            continue
                        r = _erp.recete_bagla(
                            parent_kod, parent_adi, parent_id, child_id, miktar,
                            hat_tipi=hat_tipi, live=self._live, conn=conn)

                    if r["durum"] == "olusturuldu":
                        ozet["bom"] += 1
                        self.log_mesaj.emit(
                            "  [BOM-OK]   %s ← %s (x%g)" % (parent_kod, child_kod, miktar))
                    elif r["durum"] == "atlandi":
                        self.log_mesaj.emit(
                            "  [BOM-VAR]  %s ← %s zaten bağlı" % (parent_kod, child_kod))
                    elif r["durum"] == "olusturulacak":
                        self.log_mesaj.emit(
                            "  [BOM-KURU] %s ← %s" % (parent_kod, child_kod))
                    else:
                        self.log_mesaj.emit(
                            "  [BOM-HATA] %s ← %s: %s" % (
                                parent_kod, child_kod, r.get("mesaj", "")))

            # ── Faz 3: Silinen satırlar → DELETE ─────────────────────────────
            if self._deleted_girdi_ids:
                self.log_mesaj.emit(
                    "[Sil] %d bileşen CEO ERP'den siliniyor" % len(self._deleted_girdi_ids))
                for gid in self._deleted_girdi_ids:
                    r = _erp.recete_satiri_sil(gid, live=self._live, conn=conn)
                    if r["durum"] == "silindi":
                        ozet["silindi"] = ozet.get("silindi", 0) + 1
                        self.log_mesaj.emit("  [SİLİNDİ]  GirdiId=%d" % gid)
                    elif r["durum"] == "olusturulacak":
                        ozet["silindi"] = ozet.get("silindi", 0) + 1
                        self.log_mesaj.emit("  [SİL-KURU] GirdiId=%d" % gid)
                    else:
                        self.log_mesaj.emit(
                            "  [SİL-HATA] GirdiId=%d: %s" % (gid, r.get("mesaj", "")))

            conn.close()
            self.log_mesaj.emit(
                "[Aktar] Bitti — oluşturuldu=%d atlandı=%d hata=%d  BOM=%d  "
                "güncellendi=%d  silindi=%d" % (
                    ozet.get("olusturuldu", 0), ozet.get("atlandi", 0),
                    ozet.get("hata", 0), ozet.get("bom", 0),
                    ozet.get("guncellendi", 0), ozet.get("silindi", 0)))
            self.bitti.emit(ozet)
        except Exception as e:
            self.hata.emit(str(e))


# ── BOM harita sabitleri ─────────────────────────────────────────────────────

_NODE_W, _NODE_H, _GAP_X, _GAP_Y = 180, 72, 24, 56
_TIP_RENK = {
    "Mamül":     ("#e8eaf6", "#3949ab"),
    "Yarımamül": ("#f3e5f5", "#8e24aa"),
    "Hammadde":  ("#e8f5e9", "#43a047"),
    "Masraf":    ("#fff3e0", "#fb8c00"),
    "Reçete":    ("#e1f5fe", "#039be5"),
}
_TIP_RENK_DEFAULT = ("#f5f5f5", "#757575")


class _DugumItem(QGraphicsPathItem):
    """BOM düğüm kutusu — context menu VIEW seviyesinde yakalanır."""

    def __init__(self, path, node_dict: dict, harita_ref):
        super().__init__(path)
        self._node = node_dict
        self._harita = harita_ref
        self.setFlag(QGraphicsPathItem.ItemIsSelectable, True)

    def paint(self, painter, option, widget=None):
        from PyQt5.QtWidgets import QStyleOptionGraphicsItem, QStyle
        opt = QStyleOptionGraphicsItem(option)
        opt.state &= ~QStyle.State_Selected   # Qt'nin varsayılan kesik çizgisini kapat
        super().paint(painter, opt, widget)
        if self.isSelected():
            painter.setPen(QPen(QColor("#f57c00"), 3.0))
            painter.setBrush(Qt.NoBrush)
            painter.drawPath(self.path())


class _ReceteHaritaWidget(QGraphicsView):
    """Salt-okunur BOM ağaç diyagramı — kutucuklar ve L-bağlantı çizgileri."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._sahne = QGraphicsScene(self)
        self.setScene(self._sahne)
        self.setFrameShape(QGraphicsView.NoFrame)
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.TextAntialiasing)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setBackgroundBrush(QBrush(QColor("#fafafa")))
        self.setMinimumWidth(280)
        self.setFocusPolicy(Qt.StrongFocus)
        self.setToolTip(
            "BOM Diyagramı — tekerlek: zoom · sol tık: pan\n"
            "CTRL+sol tık: çoklu seçim · sağ tık: menü"
        )
        self._nodes: list = []

    def guncelle(self, nodes_list: list):
        self._nodes = copy.deepcopy(nodes_list)
        self._sahne.clear()
        if not self._nodes:
            return
        x = 0.0
        for root in self._nodes:
            self._yerles(root, x, 0.0)
            x += self._genislik(root) + _GAP_X
        rect = self._sahne.itemsBoundingRect()
        self.fitInView(rect, Qt.KeepAspectRatio)
        # Büyük ağaçlarda minimum okunaklı zoom: kök üstte, scroll ile gez
        if self.transform().m11() < 0.30:
            self.resetTransform()
            self.scale(0.30, 0.30)
            vp_half = self.viewport().height() / (2 * 0.30)
            self.centerOn(rect.center().x(), rect.top() + vp_half)

    def temizle(self):
        self._sahne.clear()
        self._nodes = []

    def degisiklikleri_al(self) -> list:
        return self._nodes

    # ── ağaç manipülasyonu ────────────────────────────────────────────────────

    def _valid_targets(self, nodes: list, haric_kod: str, sonuc=None) -> list:
        if sonuc is None:
            sonuc = []
        for n in nodes:
            if n["kod"] == haric_kod:
                continue  # bu düğümün tüm alt ağacını atla
            sonuc.append({"kod": n["kod"], "adi": n["adi"], "tip": n["tip"]})
            self._valid_targets(n.get("children", []), haric_kod, sonuc)
        return sonuc

    def _dugumu_kaldir(self, nodes: list, kod: str) -> Optional[dict]:
        for i, n in enumerate(nodes):
            if n["kod"] == kod:
                return nodes.pop(i)
            buldu = self._dugumu_kaldir(n.get("children", []), kod)
            if buldu is not None:
                return buldu
        return None

    def _dugumu_ekle(self, nodes: list, parent_kod: Optional[str], dugum: dict) -> bool:
        if parent_kod is None:
            nodes.append(dugum)
            return True
        for n in nodes:
            if n["kod"] == parent_kod:
                n.setdefault("children", []).append(dugum)
                return True
            if self._dugumu_ekle(n.get("children", []), parent_kod, dugum):
                return True
        return False

    def _ebeveyn_degistir(self, node_dict: dict):
        hedefler = self._valid_targets(self._nodes, node_dict["kod"])
        dlg = _EbeveynSecDialog(node_dict, hedefler, parent=self)
        if dlg.exec_() != QDialog.Accepted:
            return
        yeni_parent_kod = dlg.secilen_kod()
        tasınan = self._dugumu_kaldir(self._nodes, node_dict["kod"])
        if tasınan is None:
            return
        self._dugumu_ekle(self._nodes, yeni_parent_kod, tasınan)
        # contextMenuEvent hâlâ bu item içinde yürütülüyor;
        # sahneyi anında temizlemek (clear) use-after-free çökmesine yol açar.
        # singleShot(0) event handler dönene kadar erteleyerek bunu önler.
        nodes_snap = copy.deepcopy(self._nodes)
        QTimer.singleShot(0, lambda: self.guncelle(nodes_snap))

    def _dugum_atasi_mi(self, node: dict, hedef_kod: str) -> bool:
        """node'un alt ağacında hedef_kod var mı? (döngü koruması)"""
        for child in node.get("children", []):
            if child["kod"] == hedef_kod:
                return True
            if self._dugum_atasi_mi(child, hedef_kod):
                return True
        return False

    def _secilileri_bagla(self, hedef_node: dict, secili_dugumler: list):
        """Seçili düğümleri (alt ağaçlarıyla) hedef_node'un çocuğu yap."""
        hedef_kod = hedef_node["kod"]

        # Hedef düğümü seçimden çıkar; döngü yaratacakları filtrele
        tasınacaklar = [
            it for it in secili_dugumler
            if it._node["kod"] != hedef_kod
            and not self._dugum_atasi_mi(it._node, hedef_kod)
        ]
        if not tasınacaklar:
            QMessageBox.information(
                self, "Bağlama",
                "Taşınabilecek düğüm bulunamadı.\n"
                "(Hedef, seçili düğümlerin alt ağacında olabilir.)")
            return

        # Yarımamül → Reçete dönüşümü onayı
        if hedef_node.get("tip") == "Yarımamül":
            cevap = QMessageBox.question(
                self, "Tip Dönüşümü",
                "Hedef düğüm <b>%s</b> Yarımamül tipinde.\n\n"
                "Reçete'ye dönüştürülsün mü?" % hedef_kod,
                QMessageBox.Yes | QMessageBox.No,
            )
            if cevap == QMessageBox.Yes:
                hedef_node["tip"] = "Reçete"

        for item in tasınacaklar:
            tasınan = self._dugumu_kaldir(self._nodes, item._node["kod"])
            if tasınan is not None:
                hedef_node.setdefault("children", []).append(tasınan)

        nodes_snap = copy.deepcopy(self._nodes)
        QTimer.singleShot(0, lambda: self.guncelle(nodes_snap))

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Control:
            self.setDragMode(QGraphicsView.NoDrag)
            self.viewport().setCursor(Qt.ArrowCursor)
        super().keyPressEvent(event)

    def keyReleaseEvent(self, event):
        if event.key() == Qt.Key_Control:
            self.setDragMode(QGraphicsView.ScrollHandDrag)
        super().keyReleaseEvent(event)

    def wheelEvent(self, event):
        factor = 1.25 if event.angleDelta().y() > 0 else 0.8
        self.scale(factor, factor)

    def contextMenuEvent(self, event):
        # QGraphicsTextItem üstte olabileceği için itemAt tek item döndürmez;
        # tüm item'ları tara ve ilk _DugumItem'ı bul.
        scene_pos = self.mapToScene(event.pos())
        dugum = next(
            (it for it in self._sahne.items(scene_pos) if isinstance(it, _DugumItem)),
            None
        )
        if dugum is None:
            super().contextMenuEvent(event)
            return

        secili = [it for it in self._sahne.selectedItems() if isinstance(it, _DugumItem)]
        menu = QMenu(self)

        if secili:
            act = menu.addAction(
                "🔗  Seçilileri Buraya Bağla  (%d düğüm)" % len(secili))
            chosen = menu.exec_(event.globalPos())
            if chosen == act:
                self._secilileri_bagla(dugum._node, secili)
        else:
            act = menu.addAction("🔁  Ebeveynini Değiştir")
            chosen = menu.exec_(event.globalPos())
            if chosen == act:
                self._ebeveyn_degistir(dugum._node)

    def _genislik(self, node: dict) -> float:
        children = node.get("children", [])
        if not children:
            return float(_NODE_W)
        total = sum(self._genislik(c) for c in children) + _GAP_X * (len(children) - 1)
        return max(float(_NODE_W), total)

    def _yerles(self, node: dict, x: float, y: float):
        self._dugum_ciz(node, x, y)
        children = node.get("children", [])
        if not children:
            return
        toplam = sum(self._genislik(c) for c in children) + _GAP_X * (len(children) - 1)
        cx = x + _NODE_W / 2 - toplam / 2
        cy = y + _NODE_H + _GAP_Y
        for child in children:
            cw = self._genislik(child)
            child_x = cx + cw / 2 - _NODE_W / 2
            self._baglanti_ciz(x + _NODE_W / 2, y + _NODE_H, child_x + _NODE_W / 2, cy)
            self._yerles(child, child_x, cy)
            cx += cw + _GAP_X

    def _dugum_ciz(self, node: dict, x: float, y: float):
        tip = node.get("tip", "")
        fill, border = _TIP_RENK.get(tip, _TIP_RENK_DEFAULT)
        path = QPainterPath()
        path.addRoundedRect(QRectF(x, y, _NODE_W, _NODE_H), 8, 8)
        box = _DugumItem(path, node, self)
        box.setBrush(QBrush(QColor(fill)))
        box.setPen(QPen(QColor(border), 1.8))
        self._sahne.addItem(box)
        kod = node.get("kod", "—")
        adi = node.get("adi", "")
        miktar = node.get("miktar", "")
        birim = node.get("birim", "")
        adi_k = adi[:22] + ("…" if len(adi) > 22 else "")
        html = (
            f'<div style="font-size:9px;color:{border};">{tip}</div>'
            f'<div style="font-size:11px;font-weight:bold;color:#212121;">{kod}</div>'
            f'<div style="font-size:9px;color:#555;">{adi_k}</div>'
            f'<div style="font-size:9px;color:#757575;">{miktar} {birim}</div>'
        )
        t = QGraphicsTextItem()
        t.setHtml(html)
        t.setTextWidth(_NODE_W - 8)
        t.setPos(x + 4, y + 3)
        self._sahne.addItem(t)

    def _baglanti_ciz(self, px: float, py: float, cx: float, cy: float):
        mid_y = (py + cy) / 2
        path = QPainterPath()
        path.moveTo(px, py)
        path.lineTo(px, mid_y)
        path.lineTo(cx, mid_y)
        path.lineTo(cx, cy)
        item = QGraphicsPathItem(path)
        item.setPen(QPen(QColor("#9e9e9e"), 1.5, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
        self._sahne.addItem(item)


# ── Ana sekme widget'ı ────────────────────────────────────────────────────────

class StokHazirlaTab(QWidget):
    durum_guncelle = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._thread               : Optional[QThread]           = None
        self._birim_thread         : Optional[QThread]           = None
        self._birimler             : dict[str, str]              = {"ADET": _ADET_GUID}
        self._uid_say              : int                         = 0
        self._uid_harita           : dict[int, QTreeWidgetItem]  = {}
        self._birim_yuklendi       : bool                        = False
        self._yuklenen_girdi_ids   : set[int]                    = set()
        self._bom_nodes            : list                          = []

        ana = QVBoxLayout(self)
        ana.setContentsMargins(12, 12, 12, 10)
        ana.setSpacing(10)

        baslik = QLabel("<b>CEO ERP — Stok Hazırlık</b>")
        baslik.setStyleSheet("font-size:15px;color:#1a237e;")
        ana.addWidget(baslik)

        if not ERP_OK:
            uyari = QLabel("⚠  ERP modülü yüklenemedi:\n%s" % _ERP_HATA)
            uyari.setWordWrap(True)
            uyari.setStyleSheet("color:#b00000;padding:20px;font-size:12px;")
            ana.addWidget(uyari)
            return

        self._birim_delegate = _BirimDelegate(self)

        ana.addWidget(self._arac_cubugu())
        ana.addWidget(self._agac_olustur(), stretch=1)

        self._ozet_lbl = QLabel("Ağaca öğe ekleyin veya Excel'den yükleyin.")
        self._ozet_lbl.setStyleSheet("color:#555;font-size:12px;padding:2px 0;")
        ana.addWidget(self._ozet_lbl)

        self._log = QPlainTextEdit()
        self._log.setReadOnly(True)
        self._log.setStyleSheet(_LOG_STIL)
        self._log.setFixedHeight(110)
        self._log.setPlaceholderText("İşlem çıktıları burada görünür…")
        ana.addWidget(self._log)

    def showEvent(self, event):
        super().showEvent(event)
        if ERP_OK and not self._birim_yuklendi:
            self._birim_yuklendi = True
            self._birim_thread = _BirimleriYukleThread()
            self._birim_thread.bitti.connect(self._birimler_yuklendi)
            self._birim_thread.start()

    def _birimler_yuklendi(self, birimler: dict):
        self._birimler = birimler
        self._birim_delegate.set_birimler(sorted(birimler.keys()))
        self._log_yaz("[Birim] %d birim yüklendi (CEO ERP)" % len(birimler))
        self.durum_guncelle.emit("Stok Hazırlık: %d birim yüklendi" % len(birimler))

    # ── araç çubuğu ──────────────────────────────────────────────────────────

    def _arac_cubugu(self) -> QWidget:
        w = QWidget()
        vlay = QVBoxLayout(w)
        vlay.setContentsMargins(0, 0, 0, 0)
        vlay.setSpacing(4)

        # Satır 1: Reçete yükleme
        row1 = QHBoxLayout()
        row1.setSpacing(6)

        self._yukle_btn = QPushButton("🔄  Reçete Yükle")
        self._yukle_btn.setStyleSheet(_BTN_MAVI)
        self._yukle_btn.setToolTip(
            "CEO ERP'deki mevcut reçeteyi ağaca yükle (stok kodu sorar)")
        self._yukle_btn.clicked.connect(self._recete_yukle)
        row1.addWidget(self._yukle_btn)

        self._bom_btn = QPushButton("🗺  BOM Diyagramı")
        self._bom_btn.setStyleSheet(
            "QPushButton{background:#4a148c;color:white;border:none;"
            "border-radius:4px;padding:5px 14px;font-size:13px;}"
            "QPushButton:hover{background:#6a1b9a;}"
            "QPushButton:disabled{background:#bdbdbd;color:#757575;}"
        )
        self._bom_btn.setToolTip("BOM ağacını görsel diyagram penceresinde aç")
        self._bom_btn.setEnabled(False)
        self._bom_btn.clicked.connect(self._bom_goster)
        row1.addWidget(self._bom_btn)

        row1.addStretch()
        vlay.addLayout(row1)

        # Satır 2: Düzenleme + aktarım
        row2 = QHBoxLayout()
        row2.setSpacing(6)

        self._excel_btn = QPushButton("📥  Excel'den Yükle")
        self._excel_btn.setStyleSheet(_BTN_GERI)
        self._excel_btn.clicked.connect(self._excel_yukle)
        row2.addWidget(self._excel_btn)

        self._excel_aktar_btn = QPushButton("📤  Excel'e Aktar")
        self._excel_aktar_btn.setStyleSheet(_BTN_GERI)
        self._excel_aktar_btn.clicked.connect(self._excel_aktar)
        row2.addWidget(self._excel_aktar_btn)

        ust_btn = QPushButton("➕  Üst Satır")
        ust_btn.setStyleSheet(_BTN_GERI)
        ust_btn.setToolTip("Ağaca kök öğe ekle")
        ust_btn.clicked.connect(self._ust_satir_ekle)
        row2.addWidget(ust_btn)

        alt_btn = QPushButton("⬇  Alt Satır")
        alt_btn.setStyleSheet(_BTN_GERI)
        alt_btn.setToolTip("Seçili öğenin altına bileşen ekle")
        alt_btn.clicked.connect(self._alt_satir_ekle)
        row2.addWidget(alt_btn)

        sil_btn = QPushButton("🗑  Sil")
        sil_btn.setStyleSheet(_BTN_GERI)
        sil_btn.clicked.connect(self._satir_sil)
        row2.addWidget(sil_btn)

        temizle_btn = QPushButton("Temizle")
        temizle_btn.setStyleSheet(_BTN_GERI)
        temizle_btn.clicked.connect(self._temizle)
        row2.addWidget(temizle_btn)

        row2.addStretch()

        self._kontrol_btn = QPushButton("🔍  Kontrol Et (Kuru)")
        self._kontrol_btn.setStyleSheet(_BTN_TURUNCU)
        self._kontrol_btn.clicked.connect(self._kontrol_et)
        row2.addWidget(self._kontrol_btn)

        self._canli_btn = QPushButton("✅  CANLI Aktar")
        self._canli_btn.setStyleSheet(_BTN_YESIL)
        self._canli_btn.clicked.connect(self._canli_aktar)
        row2.addWidget(self._canli_btn)

        vlay.addLayout(row2)
        return w

    def _bom_goster(self):
        self._bom_nodes = self._agac_verisi_al()
        if not self._bom_nodes:
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("BOM Diyagramı")
        dlg.setWindowFlags(
            dlg.windowFlags()
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowMinimizeButtonHint
        )
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(0, 4, 0, 8)
        lay.setSpacing(6)

        harita = _ReceteHaritaWidget()
        lay.addWidget(harita, stretch=1)

        bar = QHBoxLayout()
        bar.setContentsMargins(12, 0, 12, 0)
        bar.addStretch()

        sifirla_btn = QPushButton("↩  Sıfırla")
        sifirla_btn.setStyleSheet(_BTN_GERI)
        sifirla_btn.setToolTip("Tüm değişiklikleri geri al ve orijinal ağacı yükle")
        bar.addWidget(sifirla_btn)

        kaydet_btn = QPushButton("💾  Kaydet")
        kaydet_btn.setStyleSheet(_BTN_YESIL)
        kaydet_btn.setToolTip("Diyagram değişikliklerini Stok Hazırlık ağacına yansıt")
        bar.addWidget(kaydet_btn)
        lay.addLayout(bar)

        nodes_orijinal = copy.deepcopy(self._bom_nodes)
        QTimer.singleShot(0, lambda: harita.guncelle(nodes_orijinal))

        def _kaydet():
            degisen = harita.degisiklikleri_al()
            self._bom_nodes = copy.deepcopy(degisen)
            self._agac_nodes_yukle(degisen)
            self._log_yaz("[BOM] Diyagram değişiklikleri ağaca yansıtıldı.")
            dlg.accept()

        def _sifirla():
            harita.guncelle(nodes_orijinal)

        kaydet_btn.clicked.connect(_kaydet)
        sifirla_btn.clicked.connect(_sifirla)

        dlg.showMaximized()
        dlg.exec_()

    def _agac_nodes_yukle(self, nodes_list: list):
        """BOM node listesini QTreeWidget'a yansıt (diyagram kaydet sonrası)."""
        self._agac.clear()
        self._uid_harita.clear()
        self._uid_say = 0
        self._yuklenen_girdi_ids.clear()

        def _ekle(node: dict, parent):
            item = self._item_olustur(
                tip=node.get("tip", "Hammadde"),
                kod=node.get("kod", ""),
                adi=node.get("adi", ""),
                adi2=node.get("adi2", ""),
                miktar=str(node.get("miktar", "1")),
                birim=node.get("birim", "ADET"),
            )
            girdi_id   = node.get("girdi_id")
            org_miktar = node.get("org_miktar")
            item.setData(_COL_TIP, _ROLE_GIRDI_ID,   girdi_id)
            item.setData(_COL_TIP, _ROLE_ORG_MIKTAR, org_miktar)
            if girdi_id is not None:
                self._yuklenen_girdi_ids.add(int(girdi_id))
            if parent is None:
                self._agac.addTopLevelItem(item)
            else:
                parent.addChild(item)
            for child in node.get("children", []):
                _ekle(child, item)

        for node in nodes_list:
            _ekle(node, None)
        self._agac.expandAll()

    # ── ağaç ─────────────────────────────────────────────────────────────────

    def _agac_olustur(self) -> QTreeWidget:
        self._agac = QTreeWidget()
        self._agac.setColumnCount(7)
        self._agac.setHeaderLabels(
            ["Tip", "Stok Kodu", "Stok Adı", "Stok Adı-2", "Miktar", "Birim", "Durum"])
        hdr = self._agac.header()
        hdr.setSectionResizeMode(_COL_TIP,    QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(_COL_KOD,    QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(_COL_ADI,    QHeaderView.Stretch)
        hdr.setSectionResizeMode(_COL_ADI2,   QHeaderView.Interactive)
        hdr.setSectionResizeMode(_COL_MIKTAR, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(_COL_BIRIM,  QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(_COL_DURUM,  QHeaderView.ResizeToContents)
        self._agac.setColumnWidth(_COL_ADI2, 150)
        self._agac.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._agac.setStyleSheet(_AGAC_STIL)
        self._agac.setAllColumnsShowFocus(True)
        self._agac.setExpandsOnDoubleClick(False)
        _metin = _MetinDelegate(self._agac)
        self._agac.setItemDelegateForColumn(_COL_TIP,   _TipDelegate(self._agac))
        self._agac.setItemDelegateForColumn(_COL_KOD,   _metin)
        self._agac.setItemDelegateForColumn(_COL_ADI,   _metin)
        self._agac.setItemDelegateForColumn(_COL_ADI2,  _metin)
        self._agac.setItemDelegateForColumn(_COL_MIKTAR,_metin)
        self._agac.setItemDelegateForColumn(_COL_BIRIM, self._birim_delegate)
        self._agac.setItemDelegateForColumn(_COL_DURUM, _NoEditDelegate(self._agac))
        return self._agac

    # ── öğe oluşturma ─────────────────────────────────────────────────────────

    def _yeni_uid(self) -> int:
        self._uid_say += 1
        return self._uid_say

    def _item_olustur(self, tip="Hammadde", kod="", adi="", adi2="",
                      miktar="1", birim="ADET") -> QTreeWidgetItem:
        uid  = self._yeni_uid()
        item = QTreeWidgetItem()
        item.setData(_COL_TIP, Qt.UserRole,    uid)
        item.setData(_COL_TIP, _ROLE_GIRDI_ID,   None)   # yeni öğe; yükleme sırasında dolar
        item.setData(_COL_TIP, _ROLE_ORG_MIKTAR, None)
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        item.setText(_COL_TIP,    tip)
        item.setText(_COL_KOD,    kod)
        item.setText(_COL_ADI,    adi)
        item.setText(_COL_ADI2,   adi2)
        item.setText(_COL_MIKTAR, miktar)
        item.setText(_COL_BIRIM,  birim)
        item.setText(_COL_DURUM,  "—")
        item.setBackground(_COL_DURUM, _C_BOSTA)
        self._uid_harita[uid] = item
        return item

    def _ust_satir_ekle(self):
        item = self._item_olustur()
        self._agac.addTopLevelItem(item)
        self._agac.setCurrentItem(item)
        self._agac.scrollToItem(item)
        self._agac.editItem(item, _COL_KOD)

    def _alt_satir_ekle(self):
        parent = self._agac.currentItem()
        if parent is None:
            self._ust_satir_ekle()
            return
        item = self._item_olustur()
        parent.addChild(item)
        parent.setExpanded(True)
        self._agac.setCurrentItem(item)
        self._agac.scrollToItem(item)
        self._agac.editItem(item, _COL_KOD)

    def _satir_sil(self):
        secili = self._agac.currentItem()
        if secili is None:
            return
        cevap = QMessageBox.question(
            self, "Silme Onayı",
            "Bu öğeyi ve tüm alt öğelerini silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No)
        if cevap != QMessageBox.Yes:
            return
        self._uid_haritadan_cikar(secili)
        p = secili.parent()
        if p:
            p.removeChild(secili)
        else:
            self._agac.takeTopLevelItem(
                self._agac.indexOfTopLevelItem(secili))

    def _uid_haritadan_cikar(self, item: QTreeWidgetItem):
        uid = item.data(_COL_TIP, Qt.UserRole)
        if uid is not None:
            self._uid_harita.pop(uid, None)
        for i in range(item.childCount()):
            self._uid_haritadan_cikar(item.child(i))

    def _temizle(self):
        self._agac.clear()
        self._uid_harita.clear()
        self._uid_say = 0
        self._yuklenen_girdi_ids.clear()
        self._log.clear()
        self._ozet_lbl.setText("Ağaca öğe ekleyin veya Excel'den yükleyin.")
        self._bom_nodes = []
        self._bom_btn.setEnabled(False)

    # ── Excel yükleme ─────────────────────────────────────────────────────────

    def _excel_yukle(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", str(BASE),
            "Excel Dosyaları (*.xlsx *.xls)",
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            # ── ADLASMKE format tespiti (Stok Kodu başlıklı flat liste) ────
            _ilk = next((r for r in all_rows if any(v for v in r)), None)
            if _ilk and str(_ilk[0] or "").strip().lower() == "stok kodu":
                eklendi = self._excel_yukle_adlasmke(all_rows)
                self._log_yaz("Excel: %d satır eklendi (%s)" % (eklendi, Path(path).name))
                self.durum_guncelle.emit("Stok Hazırlık: %d satır yüklendi." % eklendi)
                self._bom_nodes = self._agac_verisi_al()
                self._bom_btn.setEnabled(bool(self._bom_nodes))
                return

            eklendi = 0
            parent_stack: list = []   # [(depth, QTreeWidgetItem)]

            for row in all_rows:
                if not row:
                    continue
                raw = [(str(v) if v is not None else "") for v in row[:7]]
                while len(raw) < 7:
                    raw.append("")

                col_a = raw[0].strip()
                if col_a.lower() in _BASLIK_ATLA:
                    continue

                # ── Format tespiti ───────────────────────────────────────────
                # Yeni format (Excel Aktar çıktısı):
                #   Col A = Seviye (int), B = Tip, C = Stok Kodu, D = Adı,
                #           E = Adı-2, F = Miktar, G = Birim
                # Eski 6-kolon format:
                #   Col A = Tip, B = Stok Kodu (girintili olabilir), C-F = diğer
                # Çok eski format: A = Kod, B = Ad
                try:
                    depth = int(float(col_a))   # "0", "1", "2.0" hepsini yakalar
                    # Yeni format
                    tip_val    = raw[1].strip()
                    kod_val    = raw[2].strip()
                    adi_val    = raw[3].strip()
                    adi2_val   = raw[4].strip()
                    miktar_val = raw[5].strip() or "1"
                    birim_val  = raw[6].strip() or "ADET"
                    if tip_val not in _TIP_SECENEKLER:
                        tip_val = "Hammadde"
                except (ValueError, TypeError):
                    # Eski formatlar
                    if col_a in _TIP_SECENEKLER:
                        # 6-kolon (girintili Stok Kodu)
                        raw_b      = raw[1]
                        stripped_b = raw_b.lstrip(" ")
                        depth      = (len(raw_b) - len(stripped_b)) // 2
                        tip_val    = col_a
                        kod_val    = stripped_b.strip()
                        adi_val    = raw[2].strip()
                        adi2_val   = raw[3].strip()
                        miktar_val = raw[4].strip() or "1"
                        birim_val  = raw[5].strip() or "ADET"
                    else:
                        # Çok eski: A=Kod, B=Ad
                        tip_val    = "Hammadde"
                        kod_val    = col_a
                        adi_val    = raw[1].strip()
                        adi2_val   = ""
                        miktar_val = "1"
                        birim_val  = "ADET"
                        depth      = 0

                if not kod_val:
                    continue

                item = self._item_olustur(
                    tip=tip_val, kod=kod_val, adi=adi_val,
                    adi2=adi2_val, miktar=miktar_val, birim=birim_val)

                while parent_stack and parent_stack[-1][0] >= depth:
                    parent_stack.pop()

                if parent_stack:
                    parent_item = parent_stack[-1][1]
                    parent_item.addChild(item)
                    parent_item.setExpanded(True)
                else:
                    self._agac.addTopLevelItem(item)

                parent_stack.append((depth, item))
                eklendi += 1

            self._log_yaz("Excel: %d satır eklendi (%s)" % (eklendi, Path(path).name))
            self.durum_guncelle.emit("Stok Hazırlık: %d satır yüklendi." % eklendi)
            self._bom_nodes = self._agac_verisi_al()
            self._bom_btn.setEnabled(bool(self._bom_nodes))
        except Exception as e:
            QMessageBox.critical(self, "Excel Hatası", "Dosya okunamadı:\n%s" % e)

    def _excel_yukle_adlasmke(self, all_rows: list) -> int:
        """ADLASMKE / CEO ERP stok listesi: flat liste, ':XX' sonek ile hiyerarşi çıkarımı.

        Baz kodu listede olan ':XX' kodlar o baz kodun çocuğu olarak eklenir;
        baz kodu listede olmayan ':XX' kodlar üst-düzey Hammadde olarak kalır.
        Baz kodu olan kodlar 'Reçete' tipiyle işaretlenir.
        Sütun: A=Stok Kodu, B=Stok Adı, C=Stok Adı-2, D=Gerçek Bakiye (görmezden gel), E=Birim.
        """
        rows: list = []
        all_codes: set = set()

        for row in all_rows:
            if not row:
                continue
            col_a = str(row[0] or "").strip()
            if not col_a or col_a.lower() in _BASLIK_ATLA:
                continue
            adi   = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
            adi2  = str(row[2] if len(row) > 2 and row[2] is not None else "").strip()
            birim = str(row[4] if len(row) > 4 and row[4] is not None else "").strip() or "ADET"
            rows.append((col_a, adi, adi2, birim))
            all_codes.add(col_a)

        # ':XX' sonek taşıyan HER kod → baz kodunun çocuğu (baz listede olsa da olmasa da)
        parent_of: dict = {}
        for kod, *_ in rows:
            if ":" in kod:
                parent_of[kod] = kod.rsplit(":", 1)[0]

        has_children: set = set(parent_of.values())
        item_map: dict = {}
        eklendi = 0

        # Geçiş 1 — üst-düzey öğeler (çocuk OLMAYAN kodlar, dosya sırasıyla)
        for kod, adi, adi2, birim in rows:
            if kod in parent_of:
                continue
            tip = "Reçete" if kod in has_children else "Hammadde"
            item = self._item_olustur(
                tip=tip, kod=kod, adi=adi, adi2=adi2, miktar="1", birim=birim,
            )
            self._agac.addTopLevelItem(item)
            if kod in has_children:
                item.setExpanded(True)
            item_map[kod] = item
            eklendi += 1

        # Geçiş 2 — çocuk öğeler; baz kodu listede yoksa sanal Reçete ebeveyni oluştur
        for kod, adi, adi2, birim in rows:
            if kod not in parent_of:
                continue
            base = parent_of[kod]
            if base not in item_map:
                vitem = self._item_olustur(
                    tip="Reçete", kod=base, adi="", adi2="", miktar="1", birim=birim,
                )
                self._agac.addTopLevelItem(vitem)
                vitem.setExpanded(True)
                item_map[base] = vitem
                eklendi += 1
            child = self._item_olustur(
                tip="Hammadde", kod=kod, adi=adi, adi2=adi2, miktar="1", birim=birim,
            )
            item_map[base].addChild(child)
            item_map[kod] = child
            eklendi += 1

        return eklendi

    # ── Excel'e Aktar ────────────────────────────────────────────────────────

    # Derinlik bazlı satır renkleri (maks 6 seviye; daha derin → son renk)
    _DERINLIK_RENKLERI = [
        "E8EAF6",   # 0: indigo 50  (Mamül/kök)
        "FFFFFF",   # 1: beyaz
        "F3E5F5",   # 2: mor 50
        "E1F5FE",   # 3: açık mavi 50
        "E8F5E9",   # 4: yeşil 50
        "FFF3E0",   # 5+: turuncu 50
    ]

    def _excel_aktar(self):
        if self._agac.topLevelItemCount() == 0:
            QMessageBox.information(self, "Boş Ağaç", "Aktarılacak öğe yok.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Aktar", str(BASE), "Excel Dosyaları (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stok Hazırlık"

            # Excel native outline: özet (parent) satır üstte olsun
            ws.sheet_properties.outlinePr.summaryBelow = False
            ws.sheet_properties.outlinePr.summaryRight = False

            # Başlık satırı
            basliklar = ["Seviye", "Tip", "Stok Kodu", "Stok Adı",
                         "Stok Adı-2", "Miktar", "Birim"]
            ws.append(basliklar)
            hdr_fill = PatternFill("solid", fgColor="3949AB")
            hdr_font = Font(bold=True, color="FFFFFF")
            for ci in range(1, len(basliklar) + 1):
                c = ws.cell(row=1, column=ci)
                c.font      = hdr_font
                c.fill      = hdr_fill
                c.alignment = Alignment(horizontal="center")

            # Veri satırları — pre-order DFS
            flat = _dfs(self._agac_verisi_al())
            for ri, (node, depth) in enumerate(flat, start=2):
                ws.append([
                    depth,
                    node.get("tip",    ""),
                    node.get("kod",    ""),
                    node.get("adi",    ""),
                    node.get("adi2",   ""),
                    node.get("miktar", "1"),
                    node.get("birim",  "ADET"),
                ])

                # Excel outline seviyesi (native +/- gruplama)
                if depth > 0:
                    ws.row_dimensions[ri].outline_level = depth

                # Satır arka plan rengi
                renk = self._DERINLIK_RENKLERI[
                    min(depth, len(self._DERINLIK_RENKLERI) - 1)]
                row_fill = PatternFill("solid", fgColor=renk)
                row_font = Font(bold=(depth == 0))
                for ci in range(1, len(basliklar) + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.fill = row_fill
                    if depth == 0:
                        cell.font = row_font

            # Sütun genişlikleri
            for col, w in zip("ABCDEFG", [8, 12, 26, 34, 22, 10, 12]):
                ws.column_dimensions[col].width = w

            # Başlık satırını sabitle
            ws.freeze_panes = "A2"

            wb.save(path)
            n = len(flat)
            self._log_yaz("Excel Aktar: %d satır → %s" % (n, Path(path).name))
            self.durum_guncelle.emit("Stok Hazırlık: %d satır Excel'e aktarıldı" % n)
            QMessageBox.information(
                self, "Aktarım Tamamlandı",
                "%d satır aktarıldı.\n\n%s" % (n, path))
        except Exception as e:
            QMessageBox.critical(self, "Excel Hatası", "Dosya yazılamadı:\n%s" % e)

    # ── log ───────────────────────────────────────────────────────────────────

    def _log_yaz(self, mesaj: str):
        self._log.moveCursor(QTextCursor.End)
        self._log.insertPlainText(mesaj + "\n")
        self._log.moveCursor(QTextCursor.End)

    # ── ağaç verisi serileştirme (main thread'de çağrılır) ───────────────────

    def _agac_verisi_al(self) -> list:
        def _item_dict(item: QTreeWidgetItem, par_kod="", par_adi="", par_tip="") -> dict:
            uid  = item.data(_COL_TIP, Qt.UserRole)
            kod  = item.text(_COL_KOD).strip()
            adi  = item.text(_COL_ADI).strip()
            tip  = item.text(_COL_TIP)
            node = {
                "uid":        uid,
                "tip":        tip,
                "kod":        kod,
                "adi":        adi,
                "adi2":       item.text(_COL_ADI2).strip(),
                "miktar":     item.text(_COL_MIKTAR).strip() or "1",
                "birim":      item.text(_COL_BIRIM).strip() or "ADET",
                "parent_kod": par_kod,
                "parent_adi": par_adi,
                "parent_tip": par_tip,
                "girdi_id":   item.data(_COL_TIP, _ROLE_GIRDI_ID),
                "org_miktar": item.data(_COL_TIP, _ROLE_ORG_MIKTAR),
                "children":   [],
            }
            for i in range(item.childCount()):
                node["children"].append(
                    _item_dict(item.child(i), kod, adi, tip))
            return node

        return [
            _item_dict(self._agac.topLevelItem(i))
            for i in range(self._agac.topLevelItemCount())
        ]

    # ── Durum sütunu güncelleme (signal slot — main thread) ──────────────────

    def _durum_guncelle_slot(self, uid: int, metin: str, renk: QColor):
        item = self._uid_harita.get(uid)
        if item:
            item.setText(_COL_DURUM, metin)
            item.setBackground(_COL_DURUM, renk)

    # ── buton durumu ─────────────────────────────────────────────────────────

    def _butonlar_durum(self, aktif: bool):
        for btn in (self._kontrol_btn, self._canli_btn,
                    self._excel_btn, self._excel_aktar_btn, self._yukle_btn):
            btn.setEnabled(aktif)
        if aktif:
            self._kontrol_btn.setStyleSheet(_BTN_TURUNCU)
            self._canli_btn.setStyleSheet(_BTN_YESIL)
            self._yukle_btn.setStyleSheet(_BTN_MAVI)
        else:
            self._kontrol_btn.setStyleSheet(_BTN_PASIF)
            self._canli_btn.setStyleSheet(_BTN_PASIF)
            self._yukle_btn.setStyleSheet(_BTN_PASIF)

    # ── Reçete Yükle ─────────────────────────────────────────────────────────

    def _recete_yukle(self):
        """CEO ERP'deki mevcut reçeteyi ağaca yükler — seçim dialog'u ile."""
        # Dialog'u oluştur ve reçete listesini arka planda çek
        self._recete_sec_dialog = _ReceteSecDialog(self)

        self._liste_thread = _TumRecetelerThread()
        self._liste_thread.bitti.connect(self._recete_sec_dialog.doldur)
        self._liste_thread.hata.connect(
            lambda mesaj: QMessageBox.critical(
                self, "Bağlantı Hatası",
                "Reçete listesi alınamadı:\n\n" + mesaj))
        self._liste_thread.start()

        if self._recete_sec_dialog.exec_() != QDialog.Accepted:
            return
        kod = self._recete_sec_dialog.secilen_kod().strip().upper()

        # Mevcut ağaç doluysa uyar
        if self._agac.topLevelItemCount() > 0:
            cevap = QMessageBox.question(
                self, "Mevcut Ağaç",
                "Mevcut ağaç temizlenecek ve <b>%s</b> reçetesi yüklenecek.\n\n"
                "Devam edilsin mi?" % kod,
                QMessageBox.Yes | QMessageBox.No,
            )
            if cevap != QMessageBox.Yes:
                return

        self._temizle()
        self._log_yaz("[Yükle] %s reçetesi yükleniyor…" % kod)
        self._butonlar_durum(False)
        self.durum_guncelle.emit("Stok Hazırlık: %s reçetesi yükleniyor…" % kod)

        self._thread = _ReceteYukleThread(kod, self._birimler)
        self._thread.bitti.connect(self._recete_yukle_bitti)
        self._thread.hata.connect(self._recete_yukle_hata)
        self._thread.start()

    def _recete_yukle_bitti(self, sonuc):
        self._butonlar_durum(True)
        if sonuc is None:
            self._log_yaz("[Yükle] CEO ERP'de bu koda ait reçete bulunamadı.")
            QMessageBox.warning(
                self, "Reçete Yok",
                "Girilen stok koduna ait reçete CEO ERP'de bulunamadı.\n"
                "Stok kartı yok ya da reçetesi henüz oluşturulmamış olabilir.")
            self.durum_guncelle.emit("Stok Hazırlık: Reçete bulunamadı.")
            return

        self._items_yukle_olustur(sonuc, parent=None)
        self._agac.expandAll()
        self._bom_nodes = self._agac_verisi_al()
        self._bom_btn.setEnabled(bool(self._bom_nodes))

        n_girdi = len(self._yuklenen_girdi_ids)
        flat    = _dfs([sonuc])
        n_item  = len(flat)
        self._log_yaz(
            "[Yükle] Tamamlandı — %d öğe yüklendi (%d bileşen bağlantısı)" % (
                n_item, n_girdi))
        mesaj = "%s yüklendi: %d öğe" % (sonuc.get("kod", "?"), n_item)
        self._ozet_lbl.setText(mesaj)
        self.durum_guncelle.emit("Stok Hazırlık: " + mesaj)

    def _recete_yukle_hata(self, mesaj: str):
        self._butonlar_durum(True)
        self._log_yaz("[Yükle-HATA] " + mesaj)
        QMessageBox.critical(self, "Yükleme Hatası",
                             "Reçete yüklenirken hata oluştu:\n\n" + mesaj)
        self.durum_guncelle.emit("Stok Hazırlık: Yükleme hatası — " + mesaj[:60])

    def _items_yukle_olustur(self, node: dict,
                              parent: "QTreeWidgetItem | None" = None):
        """Yüklenen node dict'inden QTreeWidgetItem oluşturur (recursive)."""
        item = self._item_olustur(
            tip=node.get("tip", "Hammadde"),
            kod=node.get("kod", ""),
            adi=node.get("adi", ""),
            adi2=node.get("adi2", ""),
            miktar=str(node.get("miktar", "1")),
            birim=node.get("birim_adi", "ADET"),
        )
        girdi_id   = node.get("girdi_id")
        org_miktar = node.get("org_miktar")
        item.setData(_COL_TIP, _ROLE_GIRDI_ID,   girdi_id)
        item.setData(_COL_TIP, _ROLE_ORG_MIKTAR, org_miktar)
        if girdi_id is not None:
            self._yuklenen_girdi_ids.add(int(girdi_id))

        if parent is None:
            self._agac.addTopLevelItem(item)
        else:
            parent.addChild(item)

        for child_node in node.get("children", []):
            self._items_yukle_olustur(child_node, parent=item)

    # ── Kontrol Et ───────────────────────────────────────────────────────────

    def _kontrol_et(self):
        if self._agac.topLevelItemCount() == 0:
            QMessageBox.information(
                self, "Boş Ağaç", "Önce öğe ekleyin veya Excel'den yükleyin.")
            return
        self._log.clear()
        self._butonlar_durum(False)
        self.durum_guncelle.emit("Stok Hazırlık: Kontrol ediliyor…")

        self._thread = _KontrolThread(self._agac_verisi_al())
        self._thread.satir_guncelle.connect(self._durum_guncelle_slot)
        self._thread.log_mesaj.connect(self._log_yaz)
        self._thread.bitti.connect(self._kontrol_bitti)
        self._thread.hata.connect(self._hata_slot)
        self._thread.start()

    def _kontrol_bitti(self, ozet: dict):
        self._butonlar_durum(True)
        mesaj = "%d mevcut  ·  %d yeni açılacak" % (
            ozet.get("mevcut", 0), ozet.get("yeni", 0))
        self._ozet_lbl.setText(mesaj)
        self.durum_guncelle.emit("Stok Hazırlık: Kontrol tamamlandı — " + mesaj)

    # ── CANLI Aktar ──────────────────────────────────────────────────────────

    def _canli_aktar(self):
        if self._agac.topLevelItemCount() == 0:
            QMessageBox.information(
                self, "Boş Ağaç", "Önce öğe ekleyin veya Excel'den yükleyin.")
            return

        # Silinen girdi_id'leri hesapla (yüklenen ama artık ağaçta olmayan)
        mevcut_flat = _dfs(self._agac_verisi_al())
        gorunen_girdi_ids = {
            n["girdi_id"] for n, _ in mevcut_flat
            if n.get("girdi_id") is not None
        }
        deleted_girdi_ids = self._yuklenen_girdi_ids - gorunen_girdi_ids

        firma = _erp.FIRMA_GERCEK if _erp else "?"

        # Silme uyarısı
        if deleted_girdi_ids:
            cevap = QMessageBox.question(
                self, "Bileşen Silme Onayı",
                "Ağaçtan kaldırılan <b>%d bileşen</b> CEO ERP reçetesinden de "
                "silinecek.\nBu işlem geri alınamaz.\n\n"
                "Devam edilsin mi?" % len(deleted_girdi_ids),
                QMessageBox.Yes | QMessageBox.No,
            )
            if cevap != QMessageBox.Yes:
                return
        else:
            cevap = QMessageBox.question(
                self, "CANLI Aktarım Onayı",
                "Firma <b>%s</b>'e stok kartı + BOM aktarımı yapılacak.\n"
                "Sistemde olmayan kodlar açılacak, BOM bağlantıları yazılacak.\n"
                "Mevcut kayıtlara dokunulmayacak.\n\n"
                "Devam edilsin mi?" % firma,
                QMessageBox.Yes | QMessageBox.No,
            )
            if cevap != QMessageBox.Yes:
                return

        self._log.clear()
        self._butonlar_durum(False)
        self.durum_guncelle.emit("Stok Hazırlık: CANLI aktarım başladı…")

        self._thread = _AktarThread(
            self._agac_verisi_al(),
            live=True,
            birimler=self._birimler,
            deleted_girdi_ids=deleted_girdi_ids,
        )
        self._thread.satir_guncelle.connect(self._durum_guncelle_slot)
        self._thread.log_mesaj.connect(self._log_yaz)
        self._thread.bitti.connect(self._aktar_bitti)
        self._thread.hata.connect(self._hata_slot)
        self._thread.start()

    def _aktar_bitti(self, ozet: dict):
        self._butonlar_durum(True)
        mesaj = (
            "Stok: oluşturuldu=%d atlandı=%d hata=%d  ·  BOM=%d  "
            "güncellendi=%d  silindi=%d"
        ) % (
            ozet.get("olusturuldu", 0), ozet.get("atlandi", 0),
            ozet.get("hata", 0), ozet.get("bom", 0),
            ozet.get("guncellendi", 0), ozet.get("silindi", 0),
        )
        self._ozet_lbl.setText(mesaj)
        self.durum_guncelle.emit("Stok Hazırlık: Aktarım tamamlandı — " + mesaj)

    def _hata_slot(self, mesaj: str):
        self._butonlar_durum(True)
        self._log_yaz("HATA: " + mesaj)
        QMessageBox.critical(self, "Bağlantı / İşlem Hatası",
                             "ERP bağlantısı veya işlem sırasında hata:\n\n" + mesaj)
        self.durum_guncelle.emit("Stok Hazırlık: Hata — " + mesaj[:80])
