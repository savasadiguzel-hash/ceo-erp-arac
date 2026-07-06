"""Excel raporu üretme iş mantığı. UI'dan bağımsızdır."""
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from logic.maliyet import mamul_maliyet_hesapla
from db.sorgular import bom_listesi, stok_fiyatlari_toplu

# ── Stil sabitleri ────────────────────────────────────────────────────────────

_KENAR = Border(
    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),  bottom=Side(style="thin", color="DDDDDD"),
)
_ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SOL  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_SAG  = Alignment(horizontal="right",  vertical="center")

_FILL = {
    "bilgi":    PatternFill("solid", fgColor="E8EAF6"),
    "baslik":   PatternFill("solid", fgColor="1A237E"),
    "mamul":    PatternFill("solid", fgColor="3F51B5"),
    "bileseni": PatternFill("solid", fgColor="F5F5F5"),
    "iscilik":  PatternFill("solid", fgColor="FFF3E0"),
    "toplam":   PatternFill("solid", fgColor="E8F5E9"),
    "baglandi": PatternFill("solid", fgColor="E8F5E9"),
    "atlandi":  PatternFill("solid", fgColor="FFF8E1"),
}
_FONT = {
    "baslik":   XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11),
    "mamul":    XFont(bold=True, color="FFFFFF", name="Segoe UI", size=11),
    "bileseni": XFont(name="Segoe UI", size=10),
    "iscilik":  XFont(italic=True, name="Segoe UI", size=10, color="E65100"),
    "toplam":   XFont(bold=True, name="Segoe UI", size=11, color="1B5E20"),
    "bilgi":    XFont(name="Segoe UI", size=10, color="555555"),
    "veri":     XFont(name="Segoe UI", size=10),
}

# ── Sayı format sabitleri ─────────────────────────────────────────────────────
# Excel'in "Metin olarak saklanan sayı" uyarısını önlemek için tüm sayısal
# hücreler cast edilip bu formatlarla yazılır.

_FMT_PARA   = '#,##0.00 "₺"'   # para birimi (2 ondalık)
_FMT_BIRIM  = '#,##0.0000 "₺"' # birim maliyet (4 ondalık)
_FMT_MIKTAR = '#,##0.##'        # BOM miktarı (isteğe bağlı ondalık)
_FMT_ADET   = '#,##0'           # tam sayı (fatura sayısı vb.)

# ── Sütun tanımları ───────────────────────────────────────────────────────────

_MALIYET_SUTUNLAR = [
    ("Tip", 10), ("Mamül Kodu", 14), ("Mamül Adı", 28),
    ("Bileşen Kodu", 14), ("Bileşen Adı", 30), ("BOM Miktarı", 13),
    ("Birim", 10), ("Birim Maliyet ₺", 18), ("Satır Maliyeti ₺", 18),
    ("Hammadde Toplamı ₺", 20), ("İşçilik ₺", 16), ("Genel Toplam ₺", 20),
]

# (alan_adı, sayısal?, number_format)
_BAGLAMA_ALAN = [
    ("stok_kodu",     False, None),
    ("stok_adi",      False, None),
    ("birim",         False, None),
    ("guncel_miktar",  True, _FMT_MIKTAR),
    ("fatura_sayisi",  True, _FMT_ADET),
    ("birim_fiyat",    True, _FMT_BIRIM),
    ("ilk_fatura",    False, None),
    ("son_fatura",    False, None),
    ("tedarikci",     False, None),
    ("mamul_kodu",    False, None),
    ("mamul_adi",     False, None),
    ("islem",         False, None),
]

_BAGLAMA_GENISLIK = [14, 32, 10, 16, 14, 20, 14, 14, 30, 14, 30, 13]

# ── Yardımcı fonksiyonlar ─────────────────────────────────────────────────────

def _hucre(ws, row, col, val, fill_key, font_key, alignment=_ORTA, fmt=None):
    """Hücreyi yazar; val=None ise boş bırakır (metin-sayı uyarısı olmaz)."""
    h = ws.cell(row=row, column=col, value=val)
    h.fill      = _FILL[fill_key]
    h.font      = _FONT[font_key]
    h.alignment = alignment
    h.border    = _KENAR
    if fmt:
        h.number_format = fmt
    return h


def _para(v) -> float | None:
    """
    Herhangi bir para birimi değerini float'a dönüştürür.
    Türkçe formatlı string ("37.500,00 ₺") ve sayısal tip desteklenir.
    Dönüştürülemeyen değer için None döner (boş hücre).
    """
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        # "37.500,00 ₺" → "37500.00"
        temiz = v.replace("₺", "").replace("\xa0", "").strip()
        temiz = temiz.replace(".", "").replace(",", ".")
        try:
            return float(temiz)
        except ValueError:
            return None
    return None


def _adet(v) -> int | None:
    """Tam sayıya dönüştürür; dönüştürülemeyen için None."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _miktar(v) -> float | None:
    """BOM miktarı — int veya float, None'a düşmez."""
    if isinstance(v, bool) or v is None:
        return None
    return float(v)


# ── Maliyet Raporu ────────────────────────────────────────────────────────────

def maliyet_excel_kaydet(
    dosya: str, conn,
    secili: list[tuple[str, float]],   # [(mamul_kodu, iscilik_float), ...]
    metod: str, bas: str, bit: str, bas_g: str, bit_g: str,
    ilerleme_cb=None,   # callable(str) | None — her mamül başında çağrılır
) -> None:
    metod_ad = {"WA": "Ağırlıklı Ortalama", "FIFO": "FIFO", "LIFO": "LIFO"}[metod]
    bom = bom_listesi(conn)

    # Tüm BOM'daki yaprak bileşen kodlarını topla ve fiyat geçmişini TEK sorguda çek.
    # Bu cache'i mamul_maliyet_hesapla ile paylaşarak her bileşen için ayrı DB
    # sorgusu gönderilmesini engeller (N+1 → 1 sorgu).
    cache: dict = {}
    yaprak_kodlari = list({
        b["kod"]
        for mamul_veri in bom.values()
        for b in mamul_veri["bilesenleri"]
        if b["kod"] not in bom and b.get("tip") != "masraf"
    })
    if yaprak_kodlari:
        toplu = stok_fiyatlari_toplu(conn, yaprak_kodlari, bas, bit)
        for stok_kodu, fiyat_listesi in toplu.items():
            key = ("birim", stok_kodu, metod, bas, bit)
            if not fiyat_listesi:
                cache[key] = 0.0
            elif metod == "WA":
                top_t = sum(f["birim_fiyat"] * f["miktar"] for f in fiyat_listesi)
                top_m = sum(f["miktar"] for f in fiyat_listesi)
                cache[key] = top_t / top_m if top_m else 0.0
            elif metod == "FIFO":
                cache[key] = min(fiyat_listesi, key=lambda x: x["tarih"])["birim_fiyat"]
            elif metod == "LIFO":
                cache[key] = max(fiyat_listesi, key=lambda x: x["tarih"])["birim_fiyat"]
            else:
                cache[key] = 0.0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maliyet Raporu"

    # Bilgi satırı
    n_sutun = len(_MALIYET_SUTUNLAR)
    ws.merge_cells(f"A1:{get_column_letter(n_sutun)}1")
    h = ws.cell(row=1, column=1,
                value=f"CEO ERP — Maliyet Raporu  |  Yöntem: {metod_ad}  |  "
                      f"Dönem: {bas_g} – {bit_g}  |  "
                      f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    h.fill = _FILL["bilgi"]; h.font = _FONT["bilgi"]
    h.alignment = _SOL; h.border = _KENAR
    ws.row_dimensions[1].height = 24

    # Başlık satırı
    ws.row_dimensions[2].height = 28
    for col, (ad, gen) in enumerate(_MALIYET_SUTUNLAR, 1):
        _hucre(ws, 2, col, ad, "baslik", "baslik")
        ws.column_dimensions[get_column_letter(col)].width = gen

    satir = 3

    for mamul_kodu, iscilik_ham in secili:
        mamul = bom.get(mamul_kodu)
        if not mamul:
            continue
        if ilerleme_cb:
            ilerleme_cb(f"{mamul_kodu} — {mamul['ad']} hesaplanıyor...")

        bilesenleri, hammadde_top = mamul_maliyet_hesapla(
            conn, mamul_kodu, metod, bas, bit, _cache=cache, bom=bom
        )
        iscilik   = float(iscilik_ham)
        genel_top = hammadde_top + iscilik

        # ── Mamül başlık satırı ──────────────────────────────────────────────
        # Sütunlar: Tip | MamulKod | MamulAd | — | — | — | — | — | — | HammaddeTop | Iscilik | GenelTop
        ws.row_dimensions[satir].height = 22
        _m = "mamul"
        for col, (val, aln, fmt) in enumerate([
            ("MAMÜL",       _ORTA, None),
            (mamul_kodu,    _ORTA, None),
            (mamul["ad"],   _SOL,  None),
            (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            (round(hammadde_top, 2), _SAG, _FMT_PARA),
            (round(iscilik,      2), _SAG, _FMT_PARA),
            (round(genel_top,    2), _SAG, _FMT_PARA),
        ], 1):
            _hucre(ws, satir, col, val, _m, _m, aln, fmt)
        satir += 1

        # ── Bileşen satırları ────────────────────────────────────────────────
        for b in bilesenleri:
            ws.row_dimensions[satir].height = 18
            bom_miktar = _miktar(b["bom_miktar"])
            birim_mal  = round(float(b["birim_mal"]),  4)
            satir_top  = round(float(b["satir_top"]),  2)
            _b = "bileseni"
            for col, (val, aln, fmt) in enumerate([
                (b["tip"],   _ORTA, None),
                (mamul_kodu, _ORTA, None),
                (mamul["ad"],_SOL,  None),
                (b["bil_kod"],_ORTA, None),
                (b["bil_ad"], _SOL,  None),
                (bom_miktar,  _SAG,  _FMT_MIKTAR),
                (b["birim"],  _ORTA, None),
                (birim_mal,   _SAG,  _FMT_BIRIM),
                (satir_top,   _SAG,  _FMT_PARA),
                (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            ], 1):
                _hucre(ws, satir, col, val, _b, _b, aln, fmt)
            satir += 1

        # ── İşçilik satırı ───────────────────────────────────────────────────
        if iscilik > 0:
            ws.row_dimensions[satir].height = 18
            _i = "iscilik"
            for col, (val, aln, fmt) in enumerate([
                ("İŞÇİLİK",              _ORTA, None),
                (mamul_kodu,             _ORTA, None),
                (mamul["ad"],            _SOL,  None),
                ("—",                    _ORTA, None),
                ("Manuel İşçilik Tutarı",_SOL,  None),
                (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
                (round(iscilik, 2),      _SAG,  _FMT_PARA),
                (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            ], 1):
                _hucre(ws, satir, col, val, _i, _i, aln, fmt)
            satir += 1

        # ── Toplam satırı ────────────────────────────────────────────────────
        ws.row_dimensions[satir].height = 20
        _t = "toplam"
        for col, (val, aln, fmt) in enumerate([
            ("TOPLAM",      _ORTA, None),
            (mamul_kodu,    _ORTA, None),
            (mamul["ad"],   _SOL,  None),
            (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            (None, _ORTA, None), (None, _ORTA, None), (None, _ORTA, None),
            (round(hammadde_top, 2), _SAG, _FMT_PARA),
            (round(iscilik,      2), _SAG, _FMT_PARA),
            (round(genel_top,    2), _SAG, _FMT_PARA),
        ], 1):
            _hucre(ws, satir, col, val, _t, _t, aln, fmt)
        satir += 1

        # Mamüller arası boşluk satırı
        ws.row_dimensions[satir].height = 6
        satir += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(n_sutun)}2"
    ws.freeze_panes = "A3"
    wb.save(dosya)


# ── Kesişim Kümesi Raporu ─────────────────────────────────────────────────────

_KESISIM_SUTUNLAR = [
    ("Stok Kodu", 16), ("Stok Adı", 36), ("Stok Adı-2", 36),
    ("Birim", 10), ("Güncel Miktar", 16),
    ("Fatura Sayısı", 14), ("Birim Fiyat ₺", 20),
    ("İlk Fatura", 14), ("Son Fatura", 14), ("Tedarikçi", 34),
]

_KESISIM_ALAN = [
    ("stok_kodu",     False, None),
    ("stok_adi",      False, None),
    ("stok_adi2",     False, None),
    ("birim",         False, None),
    ("guncel_miktar",  True, _FMT_MIKTAR),
    ("fatura_sayisi",  True, _FMT_ADET),
    ("birim_fiyat",    True, _FMT_BIRIM),
    ("ilk_fatura",    False, None),
    ("son_fatura",    False, None),
    ("tedarikci",     False, None),
]

_DETAY_SUTUNLAR = [
    ("Stok Kodu", 16), ("Stok Adı", 36), ("İşlem Türü", 20),
    ("Belge No", 16), ("Tarih", 14), ("Miktar", 12),
    ("Birim Fiyat ₺", 20), ("Tutar ₺", 18), ("Tedarikçi", 34),
]

_DETAY_ALAN = [
    ("stok_kodu",   False, None),
    ("stok_adi",    False, None),
    ("islem_turu",  False, None),
    ("belge_no",    False, None),
    ("tarih",       False, None),
    ("miktar",       True, _FMT_MIKTAR),
    ("birim_fiyat",  True, _FMT_BIRIM),
    ("tutar",        True, _FMT_PARA),
    ("tedarikci",   False, None),
]

_YONTEM_AD = {"FIFO": "FIFO (İlk Alış)", "LIFO": "LIFO (Son Alış)",
               "WA": "Ağırlıklı Ortalama"}


def kesisim_excel_kaydet(dosya: str, stoklar: list[dict], detaylar: list[dict],
                          tarih_aralik: str = "", yontem: str = "WA") -> None:
    yontem_ad = _YONTEM_AD.get(yontem, yontem)
    wb = openpyxl.Workbook()

    # ── Sayfa 1: Özet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Kesişim Kümesi"

    n = len(_KESISIM_SUTUNLAR)
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    h = ws.cell(
        row=1, column=1,
        value=f"CEO ERP — Kesişim Kümesi  |  {tarih_aralik}  |  "
              f"{len(stoklar)} stok  |  Yöntem: {yontem_ad}  |  "
              f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    h.fill = _FILL["bilgi"]; h.font = _FONT["bilgi"]
    h.alignment = _SOL; h.border = _KENAR
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 28
    for col, (baslik, gen) in enumerate(_KESISIM_SUTUNLAR, 1):
        _hucre(ws, 2, col, baslik, "baslik", "baslik")
        ws.column_dimensions[get_column_letter(col)].width = gen

    for satir, s in enumerate(stoklar, 3):
        ws.row_dimensions[satir].height = 18
        for col, (key, sayisal, fmt) in enumerate(_KESISIM_ALAN, 1):
            ham = s.get(key)
            if sayisal:
                val = _miktar(ham) if fmt in (_FMT_MIKTAR, _FMT_BIRIM) else _adet(ham)
                aln = _SAG
            else:
                val = str(ham).strip() if ham is not None else ""
                aln = _SOL
            h = ws.cell(row=satir, column=col, value=val)
            h.fill = _FILL["bileseni"]; h.font = _FONT["veri"]
            h.alignment = aln; h.border = _KENAR
            if fmt and val is not None:
                h.number_format = fmt

    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"
    ws.freeze_panes = "A3"

    # ── Sayfa 2: Detay (fatura satırı bazında) ───────────────────────────────
    wd = wb.create_sheet("Fatura Detayları")
    nd = len(_DETAY_SUTUNLAR)
    wd.merge_cells(f"A1:{get_column_letter(nd)}1")
    hd = wd.cell(
        row=1, column=1,
        value=f"CEO ERP — Fatura Detayları  |  {tarih_aralik}  |  "
              f"{len(detaylar)} satır  |  "
              f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    hd.fill = _FILL["bilgi"]; hd.font = _FONT["bilgi"]
    hd.alignment = _SOL; hd.border = _KENAR
    wd.row_dimensions[1].height = 24

    wd.row_dimensions[2].height = 28
    for col, (baslik, gen) in enumerate(_DETAY_SUTUNLAR, 1):
        _hucre(wd, 2, col, baslik, "baslik", "baslik")
        wd.column_dimensions[get_column_letter(col)].width = gen

    for satir, d in enumerate(detaylar, 3):
        wd.row_dimensions[satir].height = 18
        for col, (key, sayisal, fmt) in enumerate(_DETAY_ALAN, 1):
            ham = d.get(key)
            if sayisal:
                val = _miktar(ham)
                aln = _SAG
            else:
                val = str(ham).strip() if ham is not None else ""
                aln = _SOL
            hc = wd.cell(row=satir, column=col, value=val)
            hc.fill = _FILL["bileseni"]; hc.font = _FONT["veri"]
            hc.alignment = aln; hc.border = _KENAR
            if fmt and val is not None:
                hc.number_format = fmt

    wd.auto_filter.ref = f"A2:{get_column_letter(nd)}2"
    wd.freeze_panes = "A3"

    wb.save(dosya)


# ── Bağlama Raporu ────────────────────────────────────────────────────────────

def baglama_excel_kaydet(dosya: str, sonuclar: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mamül Ağacı Raporu"

    ws.row_dimensions[1].height = 30
    for col, ((_, __, ___), gen) in enumerate(
        zip(_BAGLAMA_ALAN, _BAGLAMA_GENISLIK), 1
    ):
        _hucre(ws, 1, col, _BAGLAMA_ALAN[col - 1][0].replace("_", " ").title(),
               "baslik", "baslik")
        ws.column_dimensions[get_column_letter(col)].width = gen

    # Başlık etiketlerini tekrar doğru adlarla yaz
    basliklar = [
        "Stok Kodu", "Stok Adı", "Birim", "Güncel Miktar",
        "Fatura Sayısı", "Birim Fiyat", "İlk Fatura", "Son Fatura",
        "Tedarikçi", "Mamül Kodu", "Mamül Adı", "İşlem",
    ]
    for col, (baslik, gen) in enumerate(zip(basliklar, _BAGLAMA_GENISLIK), 1):
        _hucre(ws, 1, col, baslik, "baslik", "baslik")
        ws.column_dimensions[get_column_letter(col)].width = gen

    for satir, s in enumerate(sonuclar, 2):
        fill_key = "baglandi" if s.get("islem") == "Bağlandı" else "atlandi"
        ws.row_dimensions[satir].height = 18

        for col, (key, sayisal, fmt) in enumerate(_BAGLAMA_ALAN, 1):
            ham = s.get(key)

            if sayisal:
                if fmt == _FMT_ADET:
                    val = _adet(ham)
                elif fmt == _FMT_PARA:
                    val = _para(ham)
                else:  # _FMT_MIKTAR, _FMT_BIRIM
                    val = _miktar(ham)
                aln = _SAG
            else:
                val = str(ham).strip() if ham is not None else ""
                aln = _SOL

            h = ws.cell(row=satir, column=col, value=val)
            h.fill      = _FILL[fill_key]
            h.font      = _FONT["veri"]
            h.alignment = aln
            h.border    = _KENAR
            if fmt and val is not None:
                h.number_format = fmt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_BAGLAMA_ALAN))}1"
    ws.freeze_panes = "A2"
    wb.save(dosya)
