"""Boş İş Emri Excel şablonu — PDF ile birebir aynı görünüm, A4 baskı."""

import os, sys
import openpyxl
from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
from openpyxl.utils  import get_column_letter

SIRKET_ADI   = "Gempa Elektro Mekanik Mühendislik Ltd. Şti"
FORM_BASLIGI = "ÜRETİM İŞ EMRİ"
FORM_NO_PARTS = ["SRCF.06.01.01", "05.06.2026", "REV02"]

BOSH_SATIR = 10   # ürün tablosundaki boş satır sayısı

# ── Renkler (PDF ile aynı) ────────────────────────────────────────────────────
NAVY     = "1A237E"
NAVY_MID = "3949AB"
LABEL_BG = "F0F2FC"
BEYAZ    = "FFFFFF"
ALT_ROW  = "F6F7FD"
CIZGI    = "C5CAE9"
GRI      = "9E9E9E"

# ── Sütun genişlikleri ────────────────────────────────────────────────────────
# PDF sütun oranları: No=5.6% Kodu=23.3% Adı=60% Talep=11.1%
# Toplam 84 birim × oran = A(8) B(20) C(14) D(28) E(14)
# Header oranları: Logo=22% Başlık=56% FormNo=22%
#   → A:B=28 birim ≈ logo, C:D=42 ≈ başlık, E=14 ≈ formno
KOLON = {"A": 8, "B": 20, "C": 14, "D": 28, "E": 14}


# ── Stil yardımcıları ─────────────────────────────────────────────────────────

def _fill(renk: str) -> PatternFill:
    return PatternFill("solid", fgColor=renk)

def _f(bold=False, renk="000000", boy=9, italic=False) -> XFont:
    return XFont(name="Calibri", bold=bold, color=renk, size=boy, italic=italic)

def _al(yatay="left", dikey="center", sarma=False) -> Alignment:
    return Alignment(horizontal=yatay, vertical=dikey, wrap_text=sarma)

def _ince_cerceve(renk=CIZGI) -> Border:
    y = Side(style="thin", color=renk)
    return Border(left=y, right=y, top=y, bottom=y)

def _dis_cerceve(renk=NAVY, stil="medium") -> Border:
    y = Side(style=stil, color=renk)
    return Border(left=y, right=y, top=y, bottom=y)

def _alt_cizgi(renk=NAVY_MID) -> Border:
    return Border(bottom=Side(style="medium", color=renk))


def _tum_aralik_cercevele(ws, r1, c1, r2, c2):
    """Aralığın dış kenarlarını NAVY medium, iç çizgilerini CIZGI thin yapar."""
    ince = Side(style="thin",   color=CIZGI)
    dis  = Side(style="medium", color=NAVY)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = Border(
                left   = dis if c == c1 else ince,
                right  = dis if c == c2 else ince,
                top    = dis if r == r1 else ince,
                bottom = dis if r == r2 else ince,
            )


def _logo_yolu() -> str | None:
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dirs: list[str] = []
    if getattr(sys, "_MEIPASS", None):
        dirs.append(os.path.join(sys._MEIPASS, "assets"))
    dirs.append(os.path.join(base, "assets"))
    for d in dirs:
        for f in ("logo.png", "logo.jpeg", "logo.jpg"):
            p = os.path.join(d, f)
            if os.path.isfile(p):
                return p
    return None


# ── Ana fonksiyon ─────────────────────────────────────────────────────────────

def bos_is_emri_excel_olustur(dosya_yolu: str) -> None:
    """Boş, doldurulabilir İş Emri Excel şablonu oluşturur (PDF ile aynı stil)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İş Emri"

    # ── Sayfa ayarları ────────────────────────────────────────────────────────
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = ws.page_margins.right  = 0.59   # 15mm
    ws.page_margins.top    = ws.page_margins.bottom = 0.59
    ws.page_margins.header = ws.page_margins.footer = 0.0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Son satır ve baskı alanı (dinamik)
    VERI_ILK = 10                          # ilk veri satırı
    VERI_SON = VERI_ILK + BOSH_SATIR - 1  # son veri satırı (= 19)
    AYR1     = VERI_SON + 1               # ayırıcı (= 20)
    FTR_ETK  = VERI_SON + 2              # footer etiket (= 21)
    FTR_IMZ  = VERI_SON + 3              # footer imza   (= 22)
    ws.print_area = f"A1:E{FTR_IMZ}"

    # ── Sütun genişlikleri ────────────────────────────────────────────────────
    for harf, gen in KOLON.items():
        ws.column_dimensions[harf].width = gen

    # ── Satır yükseklikleri ───────────────────────────────────────────────────
    yukler = {1: 22, 2: 14, 3: 12, 4: 6, 5: 18, 6: 18, 7: 6, 8: 17, 9: 22}
    for s, y in yukler.items():
        ws.row_dimensions[s].height = y
    for r in range(VERI_ILK, VERI_SON + 1):
        ws.row_dimensions[r].height = 16
    ws.row_dimensions[AYR1].height    = 6
    ws.row_dimensions[FTR_ETK].height = 15
    ws.row_dimensions[FTR_IMZ].height = 20

    # ═══════════════════════════════════════════════════════════════════════════
    # BAŞLIK KUTUSU (Satır 1-3) — PDF ISO 3-sütun stili
    # ═══════════════════════════════════════════════════════════════════════════
    #   A:B merge  → Logo        (28 birim ≈ 33% — PDF'de %22)
    #   C:D satır başına → Başlık metinleri (42 birim ≈ 50%)
    #   E   satır başına → Form No          (14 birim ≈ 17%)

    # Sol: Logo  [A1:B3]
    ws.merge_cells("A1:B3")
    logo_h = ws["A1"]
    logo_h.fill      = _fill(BEYAZ)
    logo_h.alignment = _al("center", "center")

    # Orta satır 1: Form başlığı  [C1:D1]
    ws.merge_cells("C1:D1")
    c = ws["C1"]
    c.value     = FORM_BASLIGI
    c.font      = _f(bold=True, renk=NAVY, boy=13)
    c.fill      = _fill(BEYAZ)
    c.alignment = _al("center", "center")

    # Orta satır 2: Şirket adı  [C2:D2]
    ws.merge_cells("C2:D2")
    c = ws["C2"]
    c.value     = SIRKET_ADI
    c.font      = _f(bold=False, renk=NAVY_MID, boy=7)
    c.fill      = _fill(BEYAZ)
    c.alignment = _al("center", "center")

    # Orta satır 3: boş  [C3:D3]
    ws.merge_cells("C3:D3")
    ws["C3"].fill = _fill(BEYAZ)

    # Sağ: Form No (her satır ayrı, 3 kısım)
    for row_idx, parca in enumerate(FORM_NO_PARTS, start=1):
        c = ws.cell(row_idx, 5)   # E sütunu
        c.value     = parca
        c.font      = _f(bold=True, renk=NAVY, boy=7)
        c.fill      = _fill(BEYAZ)
        c.alignment = _al("center", "center")

    # Başlık kutusunun dış çerçevesi
    _tum_aralik_cercevele(ws, 1, 1, 3, 5)

    # Logo görseli
    logo_p = _logo_yolu()
    if logo_p:
        try:
            from openpyxl.drawing.image import Image as XImage
            img = XImage(logo_p)
            img.width  = 140   # piksel (logo.jpeg 728×340, oran 2.14)
            img.height = 65    # 140/2.14 ≈ 65px
            ws.add_image(img, "A1")
        except Exception:
            ws["A1"].value = "LOGO"

    # ═══════════════════════════════════════════════════════════════════════════
    # BİLGİ IZGARASI (Satır 5-6) — PDF'deki 2×2 grid
    # ═══════════════════════════════════════════════════════════════════════════
    # A(8)=etiket  B(20)=değer1  C(14)=etiket2  D:E(42)=değer2
    # PDF oranı   : LBL=14.4%   VAL=35.6%       LBL=14.4%  VAL=35.6%
    # Excel birim : A=8/84=9.5% B=20/84=23.8%  C=14/84=16.7% D+E=42/84=50%

    def _etiket(konum, metin):
        c = ws[konum]
        c.value     = metin
        c.font      = _f(bold=True, renk=NAVY_MID, boy=9)
        c.fill      = _fill(LABEL_BG)
        c.alignment = _al("right", "center")
        c.border    = _ince_cerceve()

    def _deger(konum):
        c = ws[konum]
        c.fill      = _fill(BEYAZ)
        c.font      = _f(boy=9)
        c.alignment = _al("left", "center")
        c.border    = _ince_cerceve()

    # Satır 5
    _etiket("A5", "Emir No:")
    _deger("B5")
    _etiket("C5", "Tarih:")
    ws.merge_cells("D5:E5")
    _deger("D5")
    ws["E5"].border = _ince_cerceve()   # birleşik hücrenin sağ kenarı

    # Satır 6
    _etiket("A6", "Durum:")
    _deger("B6")
    _etiket("C6", "Açıklama:")
    ws.merge_cells("D6:E6")
    _deger("D6")
    ws["E6"].border = _ince_cerceve()

    # Bilgi ızgarası dış çerçevesi
    _tum_aralik_cercevele(ws, 5, 1, 6, 5)

    # ═══════════════════════════════════════════════════════════════════════════
    # BÖLÜM BAŞLIĞI (Satır 8) — PDF'deki lacivert bar
    # ═══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A8:E8")
    b8 = ws["A8"]
    b8.value     = "  ÜRETİLECEK ÜRÜNLER"
    b8.font      = _f(bold=True, renk=BEYAZ, boy=10)
    b8.fill      = _fill(NAVY)
    b8.alignment = _al("left", "center")

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLO BAŞLIĞI (Satır 9) — PDF sütun oranları: 10:42:108:20
    # Excel: A(8)=No  B(20)=Kodu  C:D(42)=Adı  E(14)=Talep
    # ═══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("C9:D9")
    for konum, metin, hiza in [
        ("A9", "No",            "center"),
        ("B9", "Kodu",          "left"),
        ("C9", "Adı",           "left"),
        ("E9", "Talep Miktarı", "right"),
    ]:
        c = ws[konum]
        c.value     = metin
        c.font      = _f(bold=True, renk=NAVY_MID, boy=9)
        c.fill      = _fill(LABEL_BG)
        c.alignment = _al(hiza, "center", sarma=True)
        c.border    = _ince_cerceve()
    ws["D9"].fill   = _fill(LABEL_BG)   # birleşik Adı hücresi
    ws["D9"].border = _ince_cerceve()

    # ═══════════════════════════════════════════════════════════════════════════
    # VERİ SATIRLARI (Satır 10 – VERI_SON) — doldurulabilir boş satırlar
    # ═══════════════════════════════════════════════════════════════════════════
    for i in range(BOSH_SATIR):
        r   = VERI_ILK + i
        alt = (i % 2 == 1)
        zem = _fill(ALT_ROW) if alt else _fill(BEYAZ)

        # A: satır numarası (salt okunur görünüm)
        c_a = ws.cell(r, 1)
        c_a.value     = i + 1
        c_a.font      = _f(renk=GRI, boy=8)
        c_a.fill      = zem
        c_a.alignment = _al("center", "center")
        c_a.border    = _ince_cerceve()

        # B: Kodu
        c_b = ws.cell(r, 2)
        c_b.fill      = zem
        c_b.alignment = _al("left", "center")
        c_b.border    = _ince_cerceve()

        # C:D: Adı (birleşik)
        ws.merge_cells(f"C{r}:D{r}")
        c_c = ws.cell(r, 3)
        c_c.fill      = zem
        c_c.alignment = _al("left", "center")
        c_c.border    = _ince_cerceve()
        ws.cell(r, 4).border = _ince_cerceve()

        # E: Talep Miktarı
        c_e = ws.cell(r, 5)
        c_e.fill      = zem
        c_e.alignment = _al("right", "center")
        c_e.border    = _ince_cerceve()

    # Veri + başlık dış çerçevesi
    _tum_aralik_cercevele(ws, 9, 1, VERI_SON, 5)

    # ═══════════════════════════════════════════════════════════════════════════
    # FOOTER — PDF'deki Hazırlayan / Onaylayan / Teslim Alan
    # ═══════════════════════════════════════════════════════════════════════════
    # Etiketler
    ws.merge_cells(f"A{FTR_ETK}:B{FTR_ETK}")
    ws.merge_cells(f"D{FTR_ETK}:E{FTR_ETK}")
    for konum, metin in [
        (f"A{FTR_ETK}", "Hazırlayan:"),
        (f"C{FTR_ETK}", "Onaylayan:"),
        (f"D{FTR_ETK}", "Teslim Alan:"),
    ]:
        c = ws[konum]
        c.value     = metin
        c.font      = _f(bold=True, renk=NAVY_MID, boy=9)
        c.alignment = _al("left", "bottom")

    # İmza satırları — PDF'deki uzun çizgiler
    ws.merge_cells(f"A{FTR_IMZ}:B{FTR_IMZ}")
    ws.merge_cells(f"D{FTR_IMZ}:E{FTR_IMZ}")
    for konum in [f"A{FTR_IMZ}", f"C{FTR_IMZ}", f"D{FTR_IMZ}"]:
        c = ws[konum]
        c.border    = _alt_cizgi()
        c.alignment = _al("center", "bottom")

    wb.save(dosya_yolu)
