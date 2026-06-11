#!/usr/bin/env python3
"""
tools/mutabakat.py
Uretim Eksik Stok ciktisini CEO ERP stok ekstresiyle karsilastirir.
Kullanim: python tools/mutabakat.py [--ekstrem DOSYA] [--rapor DOSYA]
Cikis kodu: 0=tum satirlar tutuyor, 1=uyumsuz satir var, 2=INVARIANT IHLALI (arac CEO'dan fazla stok sayiyor → false negative riski)
"""
import argparse
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent

EKSTREM_VARSAYILAN = BASE / "referans" / "Stok_Kartt_Ekstresi_10062026183813.xlsx"
EKSTREM_FALLBACK   = BASE / "Stok Kartı Ekstresi_10062026183813.xlsx"
RAPOR_VARSAYILAN   = BASE / "muhasebe_eksik_raporu.xlsx"


def _safe(row, idx, default=None):
    return row[idx] if len(row) > idx else default


def _parse_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None


def load_ekstrem(path: Path) -> dict:
    """
    CEO ekstresi -> {stok_kodu: {'devir': float, 'txns': [(date, float)]}}
    Kart bakiyesi(t) = devir + sum(miktar for tarih<=t)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cards: dict = {}
    current: str | None = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # genel baslik satiri

        col0  = _safe(row, 0)
        col4  = _safe(row, 4)
        col11 = _safe(row, 11)
        col23 = _safe(row, 23)

        if col0 is not None:
            current = str(col0).strip()
            devir = float(col23) if col23 is not None else 0.0
            cards[current] = {"devir": devir, "txns": []}
            continue

        if current is None:
            continue

        txn_date = _parse_date(col4)
        if txn_date is None or col11 is None:
            continue  # baslik satiri veya toplam satiri

        try:
            amount = float(col11)
        except (TypeError, ValueError):
            continue

        cards[current]["txns"].append((txn_date, amount))

    wb.close()
    return cards


def ceo_bakiye(card: dict, hedef: date) -> float:
    return card["devir"] + sum(a for d, a in card["txns"] if d <= hedef)


def load_rapor(path: Path) -> list[tuple[date, str, float]]:
    """
    muhasebe_eksik_raporu.xlsx -> [(tarih, stok_kodu, erp_bakiye)]
    Baslik satiri atlanir; stok_kodu veya erp_bakiye bos olan satirlar atlanir.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # baslik
        stok  = _safe(row, 3)
        bakiye = _safe(row, 6)
        tarih_ham = _safe(row, 0)
        if stok is None or bakiye is None:
            continue
        txn_date = _parse_date(tarih_ham)
        if txn_date is None:
            continue
        try:
            bakiye_f = float(bakiye)
        except (TypeError, ValueError):
            continue
        rows.append((txn_date, str(stok).strip(), bakiye_f))
    wb.close()
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(
        description="CEO ERP ekstresi ile Muhasebe Eksik Raporu mutabakat"
    )
    parser.add_argument("--ekstrem", default=None,
                        help="CEO stok ekstresi xlsx (varsayilan: referans/ veya kok dizin)")
    parser.add_argument("--rapor", default=str(RAPOR_VARSAYILAN),
                        help="muhasebe_eksik_raporu.xlsx")
    parser.add_argument("--tolerans", type=float, default=0.01)
    args = parser.parse_args()

    # Ekstremi bul
    if args.ekstrem:
        ekstrem_path = Path(args.ekstrem)
    elif EKSTREM_VARSAYILAN.exists():
        ekstrem_path = EKSTREM_VARSAYILAN
    elif EKSTREM_FALLBACK.exists():
        ekstrem_path = EKSTREM_FALLBACK
    else:
        print(f"HATA: CEO ekstresi bulunamadi. --ekstrem ile belirtin.")
        return 2

    rapor_path = Path(args.rapor)
    if not rapor_path.exists():
        print(f"HATA: Rapor dosyasi bulunamadi: {rapor_path}")
        return 2

    ceo_data = load_ekstrem(ekstrem_path)
    rapor_satirlari = load_rapor(rapor_path)

    tutuyor = 0
    tutmuyor_sayisi = 0
    kapsam_disi = 0
    mismatches: dict[str, list] = defaultdict(list)
    # INVARIANT: arac_bakiye > ceo_bakiye olan satirlar → potansiyel false negative
    # Kural: araç eksiği >= CEO eksiği (araç daha fazla stok saymamali)
    invariant_ihlal: list[tuple] = []

    for tarih, stok, erp_b in rapor_satirlari:
        if stok not in ceo_data:
            kapsam_disi += 1
            continue
        ceo_b = ceo_bakiye(ceo_data[stok], tarih)
        fark = erp_b - ceo_b
        # Araç bakiyesi CEO bakiyesinden yüksekse → araç daha fazla stok sayıyor
        # → eksik raporlamama riski (false negative)
        if fark > args.tolerans:
            invariant_ihlal.append((tarih, stok, erp_b, ceo_b, fark))
        if abs(fark) <= args.tolerans:
            tutuyor += 1
        else:
            tutmuyor_sayisi += 1
            mismatches[stok].append((tarih, erp_b, ceo_b, fark))

    toplam = tutuyor + tutmuyor_sayisi
    oran = (tutuyor / toplam * 100) if toplam > 0 else 0.0

    print(f"TUTAN: {tutuyor}/{toplam} (%{oran:.0f}) | kapsam disi: {kapsam_disi} satir")

    if mismatches:
        print()
        print(f"Tutmayan kartlar ({len(mismatches)} kart, {tutmuyor_sayisi} satir):")
        print(f"{'Stok Kodu':<22} {'Satir':>5}  {'Fark':>12}  Not")
        print("-" * 60)
        for kod in sorted(mismatches):
            kayitlar = mismatches[kod]
            farklar = [k[3] for k in kayitlar]
            sabit = len({round(f, 2) for f in farklar}) == 1
            fark_goster = f"{farklar[0]:.4g}" if sabit else f"{min(farklar):.4g}..{max(farklar):.4g}"
            not_str = "SABIT" if sabit else "degisken"
            print(f"{kod:<22} {len(kayitlar):>5}  {fark_goster:>12}  {not_str}")

    # ── INVARIANT KONTROLÜ ─────────────────────────────────────────────────
    # araç_bakiye > CEO_bakiye => araç fazla stok sayıyor => false negative riski
    if invariant_ihlal:
        print()
        print("!" * 60)
        print("KRITIK UYARI: INVARIANT IHLALI — arac CEO'dan FAZLA stok sayiyor")
        print("Kural: arac_bakiye <= CEO_bakiye olmali (arac eksigi >= CEO eksigi)")
        print(f"Ihlal sayisi: {len(invariant_ihlal)} satir ({len({x[1] for x in invariant_ihlal})} farkli kart)")
        print()
        print(f"{'Stok Kodu':<22} {'Tarih':<12} {'ERP Bak':>10} {'CEO Bak':>10} {'Fark':>10}")
        print("-" * 70)
        gosterilen = 0
        for tarih, stok, erp_b, ceo_b, fark in sorted(invariant_ihlal, key=lambda x: -x[4]):
            print(f"{stok:<22} {str(tarih):<12} {erp_b:>10.2f} {ceo_b:>10.2f} {fark:>+10.2f}")
            gosterilen += 1
            if gosterilen >= 20:
                kalan = len(invariant_ihlal) - gosterilen
                if kalan > 0:
                    print(f"  ... ve {kalan} satir daha")
                break
        print("!" * 60)
        return 2  # invariant ihlali: CEO'dan yüksek bakiye

    return 1 if tutmuyor_sayisi > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
