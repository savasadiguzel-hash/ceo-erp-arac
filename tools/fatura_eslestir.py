#!/usr/bin/env python3
"""
tools/fatura_eslestir.py
E-fatura kalemlerini CEO ERP stok kartlarıyla fuzzy eşleştirme önerileri üretir.
Otomatik bağlama yok — tüm öneriler insan onayına sunulur.

Kullanım:
    python tools/fatura_eslestir.py
    python tools/fatura_eslestir.py --kalemler referans/faturalar/baska.json
    python tools/fatura_eslestir.py --stoklar referans/stok_listesi.json   # DB yerine dosya
    python tools/fatura_eslestir.py --cikti tools/benim_rapor.xlsx
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from rapidfuzz import fuzz
except ImportError:
    sys.exit("rapidfuzz gerekli:  pip install rapidfuzz")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent

ESIK_GUCLU         = 85
ESIK_GOZDEN_GECIR  = 60
ADAY_SAYISI        = 5


# ── Türkçe normalizasyon ───────────────────────────────────────────────────────

# Python'un .upper() İ/I karışıklığına düşmemesi için manuel tablo
_TO_UPPER = str.maketrans("ışğüçö", "İŞĞÜÇÖ")


def normalize(text: str) -> str:
    """Büyük harf, Türkçe karakter normalizasyonu, fazla boşluk/satır sonu temizliği."""
    text = (text or "").strip()
    text = re.sub(r"[\n\r\t]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = text.translate(_TO_UPPER).upper()
    return text


# ── Ölçü / kalite token çıkarımı ──────────────────────────────────────────────

# Çap: Ø72, Ø 72, Ø72MM — Ø sembolü zorunlu (yalın "72" sayı çakışmaması için)
_RE_CAP = re.compile(r"[ØOø]\s*(\d+(?:[.,]\d+)?)")

# Kesit: 25*70, 25x70, 25X70, 25×70
_RE_KESIT = re.compile(r"(\d+(?:[.,]\d+)?)\s*[*Xx×]\s*(\d+(?:[.,]\d+)?)")

# Paslanmaz çelik kalite kodu
_RE_KALITE = re.compile(r"\b(304|316|316L|202|410|430|1\.4301|1\.4307)\b")

# Malzeme formu
_RE_FORM = re.compile(
    r"\b(DOLU|LAMA|SAC|BORU|MİL|MIL|BANT|KÖŞE|KOSE|PROFİL|PROFIL|LEVHA|ÇUBUK|CUBUK)\b",
    re.IGNORECASE,
)


def olcu_tokenlari_cikar(text: str) -> set[str]:
    """Ölçü ve kalite tokenlarını normalize edilmiş metinden çıkarır."""
    norm = normalize(text)
    tokens: set[str] = set()

    for m in _RE_CAP.finditer(norm):
        tokens.add(f"CAP_{float(m.group(1).replace(',', '.')):.6g}")

    for m in _RE_KESIT.finditer(norm):
        a = float(m.group(1).replace(",", "."))
        b = float(m.group(2).replace(",", "."))
        lo, hi = min(a, b), max(a, b)
        tokens.add(f"KESIT_{lo:.6g}x{hi:.6g}")

    for m in _RE_KALITE.finditer(norm):
        tokens.add("KAL_" + m.group(1))

    for m in _RE_FORM.finditer(norm):
        tokens.add("FORM_" + m.group(1).upper())

    return tokens


# ── Kova ayırıcı (temiz / kirli) ──────────────────────────────────────────────

# "muhtelif", "çeşitli", İngilizce "misc" → her zaman kirli
_RE_MUHTELIF = re.compile(
    r"\b(muhtelif|çeşitli|cesitli|misc)\b",
    re.IGNORECASE | re.UNICODE,
)

# "işler" veya "işleri" — çoğul belirsiz iş tanımı (işçiliği DEĞİL)
_RE_ISLER = re.compile(r"\bişler[iı]?\b", re.IGNORECASE | re.UNICODE)


def _tutar_tutarsiz(kalem: dict) -> bool:
    """birim_fiyat × miktar tutardan %2'den fazla sapıyorsa True döner."""
    bp = kalem.get("birim_fiyat") or 0
    mk = kalem.get("miktar") or 0
    tu = kalem.get("tutar") or 0
    if bp <= 0 or mk <= 0 or tu <= 0:
        return False
    beklenen = bp * mk
    return abs(beklenen - tu) / tu > 0.02


def kirli_neden(kalem: dict) -> str | None:
    """Kalem kirli ise neden kirli olduğunu döner; temizse None döner."""
    aciklama = kalem.get("aciklama") or ""

    if _RE_MUHTELIF.search(aciklama):
        return "Belirsiz 'muhtelif/çeşitli' ifadesi"

    if _RE_ISLER.search(aciklama):
        return "Çoğul belirsiz iş tanımı ('işler/işleri')"

    if _tutar_tutarsiz(kalem):
        return "Birim fiyat × miktar ≠ tutar (>%2 sapma)"

    return None


def kova_ayir(kalemler: list[dict]) -> tuple[list[dict], list[dict]]:
    """(temiz_kalemler, kirli_kalemler) döner. Kirli kalemlere 'kirli_neden' eklenir."""
    temiz: list[dict] = []
    kirli: list[dict] = []
    for k in kalemler:
        neden = kirli_neden(k)
        if neden:
            kirli.append({**k, "kirli_neden": neden})
        else:
            temiz.append(k)
    return temiz, kirli


# ── Fuzzy eşleştirme ──────────────────────────────────────────────────────────

def _olcu_bonusu(kalem_tokens: set[str], stok_tokens: set[str]) -> float:
    """Eşleşen ölçü/kalite tokenlarına göre ek puan (max 15)."""
    if not kalem_tokens:
        return 0.0
    eslesme = kalem_tokens & stok_tokens
    return (len(eslesme) / len(kalem_tokens)) * 15.0


# CEO ERP operasyon adlarında geçen anahtar kelimeler
_OP_ANAHTAR = {"TORNA", "FREZE", "KAPLAMA", "LAZER", "KAYNAK", "BÜKME", "BÜKÜM", "TAŞLAMA", "TASIMA"}


def _kod_deseni_bonusu(kalem_norm: str, stok_kodu: str, stok_norm: str) -> float:
    """Operasyon stok kartı (:NN) ile kalem açıklaması aynı işi anıyorsa küçük bonus (max 3)."""
    if ":" not in stok_kodu:
        return 0.0
    for kw in _OP_ANAHTAR:
        if kw in stok_norm and kw in kalem_norm:
            return 3.0
    return 0.0


def olcu_eleme_tavani(kalem_tokens: set[str], stok_tokens: set[str]) -> float:
    """
    Ölçü/kalite çakışmasına göre üst skor tavanı döner.

    Kural:
      - Her iki tarafta da aynı tür ölçü var ama değer farklı → çakışma → 59.9 (EŞLEŞME YOK)
      - Kalemde ölçü var, aday kartta o tür ölçü hiç yok  → belirsiz  → 84.9 (GÖZDEN GEÇİR max)
      - Kalemde o türde ölçü yok → kısıtlama yok

    Kapsanan türler: CAP_ (çap), KESIT_ (kesit), KAL_ (kalite), FORM_ (malzeme formu)
    """
    if not kalem_tokens:
        return 100.0

    ESLESME_YOK  = 59.9
    GOZDEN_GECIR = 84.9
    tavan = 100.0

    for prefix in ("CAP_", "KESIT_", "KAL_", "FORM_"):
        k_grup = {t for t in kalem_tokens if t.startswith(prefix)}
        s_grup = {t for t in stok_tokens  if t.startswith(prefix)}

        if not k_grup:
            continue

        if s_grup:
            if not (k_grup & s_grup):
                tavan = min(tavan, ESLESME_YOK)
        else:
            tavan = min(tavan, GOZDEN_GECIR)

    return tavan


def adaylar_bul(
    kalem: dict,
    stok_kartlari: list[tuple[str, str]],
    n: int = ADAY_SAYISI,
) -> list[dict]:
    """Kalem için en iyi n aday stok kartı döner (skor + etiket ile)."""
    kalem_adi  = kalem.get("aciklama") or ""
    kalem_norm = normalize(kalem_adi)
    kalem_tok  = olcu_tokenlari_cikar(kalem_adi)

    skorlar: list[tuple[float, str, str]] = []
    for kodu, adi in stok_kartlari:
        stok_norm = normalize(adi)
        stok_tok  = olcu_tokenlari_cikar(adi)

        tavan     = olcu_eleme_tavani(kalem_tok, stok_tok)
        base      = fuzz.token_set_ratio(kalem_norm, stok_norm)
        bonus     = _olcu_bonusu(kalem_tok, stok_tok) + _kod_deseni_bonusu(kalem_norm, kodu, stok_norm)
        rank_skor = min(tavan, base + bonus)   # eleme tavanı uygulandı
        disp_skor = min(100.0, rank_skor)

        skorlar.append((rank_skor, disp_skor, kodu, adi))

    skorlar.sort(key=lambda x: -x[0])

    sonuc: list[dict] = []
    for rank_skor, disp_skor, kodu, adi in skorlar[:n]:
        if disp_skor >= ESIK_GUCLU:
            etiket = "GÜÇLÜ"
        elif disp_skor >= ESIK_GOZDEN_GECIR:
            etiket = "GÖZDEN GEÇİR"
        else:
            etiket = "EŞLEŞME YOK"
        sonuc.append({"kodu": kodu, "adi": adi, "skor": round(disp_skor, 1), "etiket": etiket})

    return sonuc


# ── DB: stok kartları ──────────────────────────────────────────────────────────

def stok_kartlari_db(conn) -> list[tuple[str, str]]:
    """Tüm aktif stok kartlarını [(Kodu, Adi), ...] döner. Sadece okuma."""
    from db.baglanti import cursor_ctx
    with cursor_ctx(conn) as cur:
        cur.execute("SELECT Kodu, Adi FROM StokKarti WHERE Aktif=1 ORDER BY Kodu")
        return [(str(r[0]), str(r[1])) for r in cur.fetchall()]


# ── Excel çıktısı ──────────────────────────────────────────────────────────────

def _dolu(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def excel_yaz(
    temiz_sonuclar: list[dict],
    kirli_kalemler: list[dict],
    cikti: Path,
) -> None:
    wb = openpyxl.Workbook()

    # ── Sayfa 1: Öneriler ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Öneriler"
    ws.freeze_panes = "A2"

    basliklar = [
        "Tedarikçi", "Fatura Kalemi", "Miktar", "Birim", "Birim Fiyat", "Tutar",
        "Stok Kodu", "Stok Adı", "Skor", "Değerlendirme",
    ]
    ws.append(basliklar)
    for c in range(1, len(basliklar) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _dolu("2E4057")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 22

    row = 2
    for blok in temiz_sonuclar:
        kalem   = blok["kalem"]
        adaylar = blok["adaylar"]

        # ── Fatura kalemi başlık bandı ──────────────────────────────────────
        ws.append([
            kalem.get("tedarikci", ""),
            kalem.get("aciklama", ""),
            kalem.get("miktar", ""),
            kalem.get("birim", ""),
            kalem.get("birim_fiyat", ""),
            kalem.get("tutar", ""),
            "", "", "", "",
        ])
        for c in range(1, 11):
            cell = ws.cell(row=row, column=c)
            cell.fill = _dolu("1F4E79")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        row += 1

        # ── Aday satırları ──────────────────────────────────────────────────
        if not adaylar:
            ws.append(["", "", "", "", "", "", "(Aday bulunamadı)", "", "", ""])
            row += 1
        else:
            _RENK = {"GÜÇLÜ": "C6EFCE", "GÖZDEN GEÇİR": "FFEB9C", "EŞLEŞME YOK": "FFC7CE"}
            for aday in adaylar:
                renk = _RENK.get(aday["etiket"], "DDEEFF")
                ws.append([
                    "", "", "", "", "", "",
                    aday["kodu"],
                    aday["adi"],
                    aday["skor"],
                    aday["etiket"],
                ])
                for c in range(7, 11):
                    ws.cell(row=row, column=c).fill = _dolu(renk)
                row += 1

        ws.append([])   # boş ayırıcı
        row += 1

    for i, w in enumerate([20, 45, 9, 7, 12, 12, 20, 45, 7, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sayfa 2: Elle Bağlanacak ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Elle Bağlanacak")
    ws2.freeze_panes = "A2"

    basliklar2 = ["Tedarikçi", "Kalem Açıklaması", "Miktar", "Birim", "Birim Fiyat", "Tutar", "Neden Kirli"]
    ws2.append(basliklar2)
    for c in range(1, len(basliklar2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _dolu("843534")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.row_dimensions[1].height = 22

    for k in kirli_kalemler:
        ws2.append([
            k.get("tedarikci", ""),
            k.get("aciklama", ""),
            k.get("miktar", ""),
            k.get("birim", ""),
            k.get("birim_fiyat", ""),
            k.get("tutar", ""),
            k.get("kirli_neden", ""),
        ])
        for c in range(1, 8):
            ws2.cell(row=ws2.max_row, column=c).fill = _dolu("FBE4D5")

    for i, w in enumerate([20, 50, 9, 7, 12, 12, 38], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(cikti)
    print(f"\nExcel kaydedildi: {cikti}")


# ── Dağıtım hesabı ────────────────────────────────────────────────────────────

def _yuvarlama_duzelt(kalemler: list[dict], beklenen: float) -> None:
    """Yuvarlama farkını son satıra ekler (in-place)."""
    toplam = sum(k["tutar"] for k in kalemler)
    fark = round(beklenen - toplam, 2)
    if fark != 0 and kalemler:
        kalemler[-1]["tutar"] = round(kalemler[-1]["tutar"] + fark, 2)


def dagitim_hesapla(
    toplam_tutar: float,
    dagitim_kalemleri: list[dict],
    yontem: str = "esit",
) -> list[dict]:
    """
    Tek tutar satırını birden çok stok koduna dağıtır.

    yontem:
      "esit"    — eşit bölme
      "agirlik" — her kalemdeki 'agirlik' alanına göre orantı
      "miktar"  — her kalemdeki 'miktar' alanına göre orantı
      "elle"    — 'tutar' zaten kullanıcı tarafından girilmiş; dokunma

    Dağıtılan tutarların toplamı beklenen'e eşit olur; yuvarlama farkı
    son satıra eklenir.
    """
    n = len(dagitim_kalemleri)
    if n == 0 or toplam_tutar <= 0:
        return dagitim_kalemleri

    if yontem == "elle":
        # Toplam dogrulama elle_dogrula() ile cagirici yapmalı
        return list(dagitim_kalemleri)

    if yontem in ("agirlik", "miktar"):
        alan = yontem
        degerler = [float(item.get(alan) or 0) for item in dagitim_kalemleri]
        for i, d in enumerate(degerler):
            if d < 0:
                raise ValueError(
                    f"Kalem {i+1} '{dagitim_kalemleri[i].get('stok_kodu','?')}': "
                    f"{alan}={d} — negatif deger kabul edilmez"
                )
            if d == 0:
                raise ValueError(
                    f"Kalem {i+1} '{dagitim_kalemleri[i].get('stok_kodu','?')}': "
                    f"{alan}=0 — sifir deger dagitimda kullanilamaz"
                )
        toplam_baz = sum(degerler)
        sonuclar = [
            {**item, "tutar": round(d / toplam_baz * toplam_tutar, 2)}
            for item, d in zip(dagitim_kalemleri, degerler)
        ]
        _yuvarlama_duzelt(sonuclar, toplam_tutar)
        return sonuclar

    # "esit" (varsayilan)
    pay = round(toplam_tutar / n, 2)
    sonuclar = [{**item, "tutar": pay} for item in dagitim_kalemleri]
    _yuvarlama_duzelt(sonuclar, toplam_tutar)
    return sonuclar


def elle_dogrula(toplam_tutar: float, kalemler: list[dict]) -> str | None:
    """
    'elle' yönteminde girilen tutarların toplamını fatura tutarıyla karşılaştırır.
    Eşitse None döner; fark varsa uyarı mesajı döner.
    """
    giren = round(sum(float(k.get("tutar") or 0) for k in kalemler), 2)
    fark = round(giren - toplam_tutar, 2)
    if abs(fark) < 0.005:
        return None
    isaret = "+" if fark > 0 else ""
    return (
        f"Elle girilen toplam {giren:,.2f} TL, fatura {toplam_tutar:,.2f} TL — "
        f"fark: {isaret}{fark:,.2f} TL"
    )


# ── Dağıtım test runner ───────────────────────────────────────────────────────

def _dagitim_test_calistir() -> None:
    """4 + 2 senaryoyla dagitim_hesapla ve elle_dogrula fonksiyonlarini dogrular."""

    TOPLAM = 2500.00
    KALEMLER: list[dict] = [
        {"stok_kodu": "OP-KYN:10", "aciklama": "MIG KAYNAK"},
        {"stok_kodu": "OP-KYN:20", "aciklama": "TIG KAYNAK"},
        {"stok_kodu": "OP-KYN:30", "aciklama": "LAZER KAYNAK"},
    ]
    SEP  = "-" * 58
    SEP2 = "=" * 58

    def _yazdir(baslik: str, sonuclar: list[dict], beklenen: float, baz_pay: float | None = None) -> None:
        print(f"\n{SEP}")
        print(f"Senaryo: {baslik}")
        print(SEP)
        n = len(sonuclar)
        for i, k in enumerate(sonuclar):
            tutar = k["tutar"]
            not_ = ""
            # Yuvarlama notu: son satir ve baz_pay verilmisse fark var mi?
            if baz_pay is not None and i == n - 1 and abs(tutar - baz_pay) >= 0.005:
                not_ = f"  <- yuvarlama farki ({tutar - baz_pay:+.2f} TL)"
            print(f"  {k.get('stok_kodu','?'):<12}  {k.get('aciklama',''):<18}  {tutar:>9,.2f} TL{not_}")
        toplam = round(sum(k["tutar"] for k in sonuclar), 2)
        fark   = round(toplam - beklenen, 2)
        kontrol = "[OK] esit" if abs(fark) < 0.005 else f"[!!] HATA  fark={fark:+,.2f} TL"
        print(f"  {'':12}  {'TOPLAM':18}  {toplam:>9,.2f} TL  {kontrol}")

    def _hata_testi(baslik: str, fn) -> None:
        print(f"\n{SEP}")
        print(f"Senaryo: {baslik}")
        print(SEP)
        try:
            fn()
            print("  [!!] HATA: ValueError bekleniyor ama gelmedi!")
        except ValueError as exc:
            print(f"  [OK] ValueError dogru yakalandi: {exc}")

    print(SEP2)
    print(f"TEST: Dagitim Hesabi  |  Fatura: {TOPLAM:,.2f} TL  |  {len(KALEMLER)} kalem")
    print(SEP2)

    # 1 ── Eşit böl
    baz = round(TOPLAM / len(KALEMLER), 2)
    s1 = dagitim_hesapla(TOPLAM, KALEMLER, "esit")
    _yazdir("ESIT BOL (2500 / 3)", s1, TOPLAM, baz_pay=baz)

    # 2 ── Ağırlığa göre (5/3/2 KG → 1250/750/500)
    k_ag = [
        {**KALEMLER[0], "agirlik": 5.0},
        {**KALEMLER[1], "agirlik": 3.0},
        {**KALEMLER[2], "agirlik": 2.0},
    ]
    s2 = dagitim_hesapla(TOPLAM, k_ag, "agirlik")
    _yazdir("AGIRLIGA GORE (5 / 3 / 2 KG)", s2, TOPLAM)

    # 3 ── Miktara göre (2/1/1 adet → 1250/625/625)
    k_mk = [
        {**KALEMLER[0], "miktar": 2},
        {**KALEMLER[1], "miktar": 1},
        {**KALEMLER[2], "miktar": 1},
    ]
    s3 = dagitim_hesapla(TOPLAM, k_mk, "miktar")
    _yazdir("MIKTARA GORE (2 / 1 / 1 adet)", s3, TOPLAM)

    # 4a ── Elle — doğru (1000+1000+500 = 2500)
    k_el = [
        {**KALEMLER[0], "tutar": 1000.0},
        {**KALEMLER[1], "tutar": 1000.0},
        {**KALEMLER[2], "tutar":  500.0},
    ]
    s4 = dagitim_hesapla(TOPLAM, k_el, "elle")
    _yazdir("ELLE — DOGRU (1000 / 1000 / 500 TL)", s4, TOPLAM)
    uyari = elle_dogrula(TOPLAM, s4)
    print(f"  Dogrulama: {'[OK] Toplam esit' if not uyari else '[!!] ' + uyari}")

    # 4b ── Elle — hatalı (1000+1000+600 = 2600 ≠ 2500)
    k_el_h = [
        {**KALEMLER[0], "tutar": 1000.0},
        {**KALEMLER[1], "tutar": 1000.0},
        {**KALEMLER[2], "tutar":  600.0},
    ]
    print(f"\n{SEP}")
    print("Senaryo: ELLE HATALI (1000 / 1000 / 600 TL  =>  toplam 2600)")
    print(SEP)
    for k in k_el_h:
        print(f"  {k.get('stok_kodu','?'):<12}  {k.get('aciklama',''):<18}  {k['tutar']:>9,.2f} TL")
    uyari_h = elle_dogrula(TOPLAM, k_el_h)
    print(f"  Dogrulama: [!!] UYARI: {uyari_h}")

    # 5 ── Negatif ağırlık → ValueError bekleniyor
    _hata_testi(
        "NEGATIF AGIRLIK (hata bekleniyor)",
        lambda: dagitim_hesapla(TOPLAM, [
            {**KALEMLER[0], "agirlik": -1.0},
            {**KALEMLER[1], "agirlik":  3.0},
            {**KALEMLER[2], "agirlik":  2.0},
        ], "agirlik"),
    )

    # 6 ── Sıfır miktar → ValueError bekleniyor
    _hata_testi(
        "SIFIR MIKTAR (hata bekleniyor)",
        lambda: dagitim_hesapla(TOPLAM, [
            {**KALEMLER[0], "miktar": 0},
            {**KALEMLER[1], "miktar": 1},
            {**KALEMLER[2], "miktar": 1},
        ], "miktar"),
    )

    print(f"\n{SEP2}")
    print("TESTLER TAMAMLANDI")
    print(SEP2)


# ── Operasyon / işçilik tespiti ───────────────────────────────────────────────

# Operasyon kök kelimeleri (normalize edilmiş metin üzerinde substring arama)
_OP_KOKLER = {
    "KAPLAMA", "KESİM", "BÜKÜM", "BÜKME", "KAYNAK", "TORNA",
    "BOYAMA", "İŞÇİL",   # İŞÇİLİK + İŞÇİLİĞİ vb. formları kapsar
    "LAZER", "FREZE", "TAŞLAMA", "MONTAJ", "İŞLEME",
}


def operasyon_mu(kalem: dict) -> bool:
    """
    Kalem operasyon/işçilik tipinde ise True döner.

    Hammadde kalemleri (KG, ölçü tokenı olan) False döner.
    Sinyal: açıklamada operasyon kökü VAR, VEYA birim Set/Adet + ölçüsüz.
    """
    acik_norm = normalize(kalem.get("aciklama") or "")
    for kok in _OP_KOKLER:
        if kok in acik_norm:
            return True
    birim = normalize(kalem.get("birim") or "")
    if birim in ("SET", "ADET") and not olcu_tokenlari_cikar(kalem.get("aciklama") or ""):
        return True
    return False


# ── Kopyala-yapıştır metin bloğu ──────────────────────────────────────────────

def metin_blogu_olustur(
    temiz_sonuclar: list[dict],
    kirli_kalemler: list[dict],
    fatura_meta: dict | None = None,
) -> str:
    """
    Eşleştirme sonuçlarını satınalmaya gönderilebilecek metin bloğuna çevirir.

    fatura_meta: {tedarikci: {"fatura_no": str, "tarih": str}, ...}  — opsiyonel dış geçersiz kılma.
    Eksik alanlar önce kalem dict'inden, sonra fatura_meta'dan alınır;
    hâlâ bulunamazsa "VERİ EKSİK" yazılır (sessiz boşluk bırakılmaz).

    Çıktı UTF-8 dosyaya yazılmak üzere tasarlanmıştır; Unicode semboller (⚠, ⚙, →) içerir.
    """
    from collections import defaultdict

    fatura_meta = fatura_meta or {}
    satirlar: list[str] = []
    kart_yok: list[dict] = []

    # ── Tedarikçiye göre grupla (giriş sırasını koru) ────────────────────────
    sira: list[str] = []
    gruplar: dict[str, list[tuple[str, dict]]] = defaultdict(list)

    for blok in temiz_sonuclar:
        ted = blok["kalem"].get("tedarikci") or "?"
        if ted not in sira:
            sira.append(ted)
        gruplar[ted].append(("temiz", blok))

    for k in kirli_kalemler:
        ted = k.get("tedarikci") or "?"
        if ted not in sira:
            sira.append(ted)
        gruplar[ted].append(("kirli", k))

    # ── Her fatura bloğu ─────────────────────────────────────────────────────
    for tedarikci in sira:
        meta = fatura_meta.get(tedarikci, {})
        # fatura_no / tarih: fatura_meta > kalem dict > "VERİ EKSİK"
        ilk_kalem = gruplar[tedarikci][0][1]
        ilk_kalem = ilk_kalem["kalem"] if "kalem" in ilk_kalem else ilk_kalem
        fno   = meta.get("fatura_no")  or ilk_kalem.get("fatura_no")  or "VERİ EKSİK"
        tarih = meta.get("tarih")      or ilk_kalem.get("tarih")      or "VERİ EKSİK"

        satirlar.append(f"FATURA: {tedarikci} — {fno} — {tarih}")

        toplam = sum(
            float((blok["kalem"] if tur == "temiz" else blok).get("tutar") or 0)
            for tur, blok in gruplar[tedarikci]
        )
        toplam_str = f"{toplam:,.2f}" if toplam > 0 else "VERİ EKSİK"
        satirlar.append(f"Toplam (KDV hariç): {toplam_str} TL")
        satirlar.append("")

        for sayac, (tur, blok) in enumerate(gruplar[tedarikci], 1):
            kalem   = blok["kalem"] if tur == "temiz" else blok
            adaylar = blok["adaylar"] if tur == "temiz" else []

            acik  = kalem.get("aciklama") or ""
            mik   = kalem.get("miktar")
            birim = kalem.get("birim") or ""
            tutar = kalem.get("tutar")

            mik_str   = f"{mik} {birim}".strip() if mik is not None else birim
            tutar_str = f"{tutar:,.2f}" if tutar else "VERİ EKSİK"

            satirlar.append(f"  {sayac}) {acik} — {mik_str} — {tutar_str} TL")

            if tur == "kirli":
                neden = kalem.get("kirli_neden", "")
                satirlar.append(f"     → ⚠ ELLE DAĞITIM GEREKİYOR ({neden})")

            elif operasyon_mu(kalem):
                # Operasyon/işçilik: somut kod dayatma; yalnızca fikir olarak listele
                olasiler = [
                    f"{a['kodu']} (skor={a['skor']})"
                    for a in adaylar[:3]
                    if a["skor"] >= ESIK_GOZDEN_GECIR
                ]
                satirlar.append(
                    "     → ⚙ İŞÇİLİK/OPERASYON — doğru mamul operasyonunu belirleyin"
                )
                olasi_str = ", ".join(olasiler) if olasiler else "(eşleşen operasyon kartı bulunamadı)"
                satirlar.append(f"        olası: {olasi_str}")

            else:
                # Hammadde: normal eşleştirme
                en_iyi = adaylar[0] if adaylar else None
                if en_iyi and en_iyi["etiket"] != "EŞLEŞME YOK":
                    stok_str = en_iyi["kodu"]
                    if len(adaylar) > 1 and adaylar[1]["skor"] >= ESIK_GOZDEN_GECIR:
                        stok_str += (
                            f"  (diğer aday: {adaylar[1]['kodu']}"
                            f"  skor={adaylar[1]['skor']})"
                        )
                    satirlar.append(f"     → Stok: {stok_str}   [{en_iyi['etiket']}]")
                else:
                    satirlar.append(
                        "     → Stok: ⚠ KART YOK — açılması gerekebilir   [EŞLEŞME YOK]"
                    )
                    kart_yok.append({"tedarikci": tedarikci, "aciklama": acik})

        satirlar.append("")   # faturalar arası boşluk

    # ── Açılacak kart adayları özeti ─────────────────────────────────────────
    if kart_yok:
        sep = "=" * 57
        satirlar.append(sep)
        satirlar.append("AÇILACAK KART ADAYLARI:")
        for item in kart_yok:
            satirlar.append(f'  ⚠  {item["tedarikci"]}: "{item["aciklama"]}"')
            satirlar.append(
                "     (DB'de eşleşen kart yok — yeni stok kartı açılması gerekebilir)"
            )
        satirlar.append(sep)

    return "\n".join(satirlar)


# ── Demo stok listesi (DB yoksa) ───────────────────────────────────────────────

_DEMO_STOKLAR: list[tuple[str, str]] = [
    ("HM-BRZ-001",   "KIZIL BRONZ DOLU Ø72 MM"),
    ("HM-BRZ-002",   "KIZIL BRONZ DOLU Ø60 MM"),
    ("HM-BRZ-003",   "KIZIL BRONZ BORU Ø80 MM"),
    ("HM-PAS-001",   "PASLANMAZ LAMA 304 KALİTE 25*70 MM"),
    ("HM-PAS-002",   "PASLANMAZ LAMA 304 KALİTE 30*60 MM"),
    ("HM-PAS-003",   "PASLANMAZ LAMA 316 KALİTE 25*70 MM"),
    ("HM-PAS-004",   "PASLANMAZ BORU 304 KALİTE Ø50 MM"),
    ("HM-PAS-005",   "PASLANMAZ MİL 304 KALİTE Ø30 MM"),
    ("HM-ALM-001",   "ALÜMİNYUM LEVHA 5 MM"),
    ("HM-CEL-001",   "ÇELİK BORU Ø40 MM"),
    ("OP-KAP:10",    "KAPLAMA:10 KROM KAPLAMA"),
    ("OP-KAP:20",    "KAPLAMA:20 NİKEL KAPLAMA"),
    ("OP-KAP:30",    "KAPLAMA:30 SERT KROM KAPLAMA"),
    ("OP-LAZ:10",    "LAZER:10 LAZER KESİM"),
    ("OP-LAZ:20",    "LAZER:20 LAZER KESİM VE BÜKÜM"),
    ("OP-TRN:10",    "TORNA:10 CNC TORNA"),
    ("OP-KYN:10",    "KAYNAK:10 MIG KAYNAK"),
]


# ── Ana akış ──────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="E-fatura kalem eşleştirme önerisi")
    ap.add_argument(
        "--kalemler", type=Path,
        default=BASE / "referans" / "faturalar" / "test_kalemleri.json",
        help="Fatura kalemleri JSON dosyası",
    )
    ap.add_argument(
        "--stoklar", type=Path, default=None,
        help="Stok listesi JSON [(kodu, adi), ...] — belirtilmezse DB'den çekilir",
    )
    ap.add_argument(
        "--cikti", type=Path,
        default=BASE / "tools" / "eslestirme_onerileri.xlsx",
    )
    ap.add_argument(
        "--metin-dosya", type=Path,
        default=BASE / "tools" / "satinalma_notu.txt",
        help="Kopyala-yapıştır metin bloğunun kaydedileceği dosya (UTF-8)",
    )
    ap.add_argument(
        "--fatura-meta", type=Path, default=None,
        help="Fatura no/tarih JSON: {tedarikci: {fatura_no, tarih}}",
    )
    ap.add_argument(
        "--test-dagitim", action="store_true",
        help="Dagitim hesabi test senaryolarini calistir ve cik",
    )
    args = ap.parse_args()

    if args.test_dagitim:
        _dagitim_test_calistir()
        return

    # ── Fatura kalemleri ──────────────────────────────────────────────────────
    if args.kalemler.exists():
        with open(args.kalemler, encoding="utf-8") as f:
            kalemler: list[dict] = json.load(f)
        print(f"Fatura kalemleri: {args.kalemler}  ({len(kalemler)} kalem)")
    else:
        sys.exit(f"Kalem dosyası bulunamadı: {args.kalemler}")

    # ── Stok kartları ─────────────────────────────────────────────────────────
    stok_kartlari: list[tuple[str, str]]

    if args.stoklar and args.stoklar.exists():
        with open(args.stoklar, encoding="utf-8") as f:
            raw = json.load(f)
        stok_kartlari = [(r[0], r[1]) for r in raw]
        print(f"Stok kartları (dosya): {args.stoklar}  ({len(stok_kartlari)} kart)")
    else:
        try:
            sys.path.insert(0, str(BASE))
            from config import DB_DEFAULTS
            from db.baglanti import get_connection
            conn = get_connection(**DB_DEFAULTS)
            stok_kartlari = stok_kartlari_db(conn)
            print(f"Stok kartları (DB): {len(stok_kartlari)} kart çekildi")
        except Exception as e:
            print(f"UYARI: DB bağlantısı kurulamadı ({e})")
            print(f"Demo mod — {len(_DEMO_STOKLAR)} örnek stok kullanılıyor.")
            stok_kartlari = _DEMO_STOKLAR

    # ── Kova ayırımı ─────────────────────────────────────────────────────────
    temiz, kirli = kova_ayir(kalemler)
    print(f"\nKova:  {len(temiz)} temiz  |  {len(kirli)} kirli (elle bağlanacak)")

    if kirli:
        print("\nKirli kalemler:")
        for k in kirli:
            print(f"  [{k.get('tedarikci', '')}] {k.get('aciklama', '')} -> {k['kirli_neden']}")

    # ── Fuzzy eşleştirme ─────────────────────────────────────────────────────
    print(f"\nEşleştirme:  {len(temiz)} kalem × {len(stok_kartlari)} stok kartı…")

    temiz_sonuclar: list[dict] = []
    for kalem in temiz:
        adaylar = adaylar_bul(kalem, stok_kartlari)
        temiz_sonuclar.append({"kalem": kalem, "adaylar": adaylar})

        acik    = kalem.get("aciklama", "")
        en_iyi  = adaylar[0] if adaylar else None
        if en_iyi:
            print(
                f"  [{kalem.get('tedarikci', ''):<8}]  {acik[:38]:<38}  -> "
                f"{en_iyi['etiket']:<15}  skor={en_iyi['skor']:5.1f}  "
                f"{en_iyi['kodu']}  {en_iyi['adi'][:30]}"
            )
        else:
            print(f"  [{kalem.get('tedarikci', ''):<8}]  {acik[:38]:<38}  -> (Aday bulunamadı)")

    # ── Fatura meta (opsiyonel: fatura_no / tarih) ───────────────────────────
    fatura_meta: dict = {}
    if args.fatura_meta and args.fatura_meta.exists():
        with open(args.fatura_meta, encoding="utf-8") as f:
            fatura_meta = json.load(f)

    # ── Excel yaz ─────────────────────────────────────────────────────────────
    args.cikti.parent.mkdir(parents=True, exist_ok=True)
    excel_yaz(temiz_sonuclar, kirli, args.cikti)

    # ── Metin bloğu üret ve kaydet ────────────────────────────────────────────
    metin = metin_blogu_olustur(temiz_sonuclar, kirli, fatura_meta)
    metin_dosya = args.metin_dosya
    metin_dosya.parent.mkdir(parents=True, exist_ok=True)
    metin_dosya.write_text(metin, encoding="utf-8")
    print(f"Metin raporu : {metin_dosya}")

    # Konsola da bas (encoding güvenli)
    print()
    try:
        print(metin)
    except UnicodeEncodeError:
        print(metin.encode(sys.stdout.encoding or "ascii", errors="replace").decode(
            sys.stdout.encoding or "ascii"
        ))

    # ── Özet ──────────────────────────────────────────────────────────────────
    guclu       = sum(1 for b in temiz_sonuclar if b["adaylar"] and b["adaylar"][0]["skor"] >= ESIK_GUCLU)
    gozden      = sum(1 for b in temiz_sonuclar if b["adaylar"] and ESIK_GOZDEN_GECIR <= b["adaylar"][0]["skor"] < ESIK_GUCLU)
    eslesme_yok = len(temiz_sonuclar) - guclu - gozden

    SEP = "-" * 60
    print(f"\n{SEP}")
    print(f"OZET   {len(kalemler)} kalem islendi")
    print(f"  Kirli  (Elle Baglanacak)   : {len(kirli)}")
    print(f"  Temiz  toplam              : {len(temiz)}")
    print(f"    GUCLU       (>={ESIK_GUCLU})        : {guclu}")
    print(f"    GOZDEN GECIR ({ESIK_GOZDEN_GECIR}-{ESIK_GUCLU - 1})      : {gozden}")
    print(f"    ESLEME YOK  (<{ESIK_GOZDEN_GECIR})        : {eslesme_yok}")
    print(SEP)


if __name__ == "__main__":
    main()
