#!/usr/bin/env python3
import openpyxl
wb = openpyxl.load_workbook("C:/yeni-erp/maliyet.xlsx")
ws = wb.active
print(f"Toplam satir: {ws.max_row}\n")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if not any(row):
        continue
    tip = str(row[0] or "")
    b   = str(row[1] or "")
    c   = str(row[2] or "")[:35]
    d   = str(row[3] or "")
    e   = str(row[4] or "")[:25]
    f   = row[5]
    h   = row[7]
    i   = row[8]
    print(f"{tip:<10} {b:<22} {c:<35} {d:<22} {e:<25} miktar={f}  birim_mal={h}  satir={i}")
