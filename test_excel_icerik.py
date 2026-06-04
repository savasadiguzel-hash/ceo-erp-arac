#!/usr/bin/env python3
"""Excel icerigini kontrol et - mamul ve bilesenleri"""
import openpyxl

wb = openpyxl.load_workbook("C:/yeni-erp/maliyet_2024_test.xlsx")
ws = wb.active

print("Excel icerigi (ilk 40 satir):\n")
print(f"{'Satir':<6} {'A (Tip)':<12} {'B (MamulKod)':<20} {'C (MamulAd)':<30} {'D (BilesenKod)':<20} {'I (SatirMal)':<15}")
print("-" * 110)

for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
    tip = str(row[0] or "")[:11]
    mamul_kod = str(row[1] or "")[:19]
    mamul_ad = str(row[2] or "")[:29]
    bilesen_kod = str(row[3] or "")[:19]
    satir_mal = row[8]
    satir_str = f"{satir_mal:>12,.2f}" if isinstance(satir_mal, (int, float)) else ""

    if tip:  # Bos satirlari atla
        print(f"{ws.min_row:<6} {tip:<12} {mamul_kod:<20} {mamul_ad:<30} {bilesen_kod:<20} {satir_str}")

# Satir sayisi
toplam = sum(1 for row in ws.iter_rows(min_row=3, values_only=True) if row[0])
print(f"\nToplam kayit satiri: {toplam}")
print(f"Satir tipleri:")
from collections import Counter
tipler = Counter(row[0] for row in ws.iter_rows(min_row=3, values_only=True) if row[0])
for tip, sayi in tipler.items():
    print(f"  {tip}: {sayi} adet")
