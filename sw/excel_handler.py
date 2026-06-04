"""excel_handler.py — konfig kod havuzu yonetimi (format-farkinda).

IKI FORMAT DESTEKLENIR (sekme adlarina bakilarak OTOMATIK algilanir):

  1) EV (eski) format  — "konfig.xlsx"
     5 sekme: Proje Kodu + Mekanik/Elektronik/Montaj/Optik
     Kategori sekmesi = kategori adi. Baslik 1. satir, veri 2'den.
     Sutunlar: A=Kod, B=Parca Adi, C=Proje, D=Ekleyen, E=Tarih, F=Geo Imza

  2) IS YERI format — "PROJE LISTESI VE KONFIGURASYON ... TAKIP LISTESI..xlsx"
     54 sekme. Kategori -> hedef sekme eslemesi:
        Mekanik   -> "200-Mekanik"            (baslik 2. satir, veri 3'ten;
                                                Kod=B, Konfig Adi=D, Proje=H)
        Montaj    -> "201-mekanik yari mamul" (baslik 1. satir, veri 2'den;
                                                Kod=A, Konfig Adi=D, Proje=H)
        Elektronik -> ATLA (kod verilmez)
        Optik      -> ATLA (kod verilmez)
     Kodlar "mevcut havuzdan" alinir: Konfig Adi (isim) hucresi BOS olan ILK
     satirin hazir kodu kullanilir; YENI KOD URETILMEZ.
     Geometrik Imza icin kopyaya YENI bir sutun eklenir.

Iki ana is:
  1. assign_codes() : her Part'a, kategorisinin hedef sekmesinde SIRADAKI bos
                      kodu atar. "Rezervasyonlu": ayni oturumda ayni kod iki
                      parcaya verilmez.
  2. write_back()    : atanan kodlari dosyaya yazar (once zaman damgali yedek).

Bir kodun "bos/musait" sayilma kurali: o satirin ISIM (Konfig Adi / Parca Adi)
hucresi bos.
"""
from __future__ import annotations

import getpass
import os
import shutil
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    raise SystemExit("HATA: 'openpyxl' eksik. Kurulum: pip install openpyxl")

from sw.models import Part, STATUS_EXISTING, STATUS_UNKNOWN, STATUS_PREVIOUSLY_CODED

# --- FORCE FORMATTING -------------------------------------------------------
# Is yeri Excel'inde havuzdaki bos hucreler bazen devasa font (72pt) ve 90 derece
# rotasyon tasiyor; yazdigimiz her hucreye bu STANDART bicimi ZORLA uygulayarak
# bozuk formati eziyoruz (talimat: Bicimlendirme Sifirlama).
std_font = Font(name="Calibri", size=11, bold=False)
std_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

# ---------------------------------------------------------------------------
# FORMAT PROFILLERI
# ---------------------------------------------------------------------------

# --- EV (eski) format ---
LEGACY_SHEETS = ["Mekanik", "Elektronik", "Montaj", "Optik"]
LEGACY_LAYOUT = {
    "header_row": 1, "code_col": 1, "name_col": 2, "proje_col": 3,
    "ekleyen_col": 4, "tarih_col": 5, "geo_col": 6,
    # musaitlik kontrolu: sadece isim sutunu (eski davranis korunur)
    "scan_cols": [2],
}

# Geriye donuk uyumluluk icin eski sabit isimler (disaridan import edilebilir)
COL_KOD, COL_PARCA_ADI, COL_PROJE = 1, 2, 3
COL_EKLEYEN, COL_TARIH, COL_GEO_SIG = 4, 5, 6
HEADER_ROW = 1
VALID_SHEETS = LEGACY_SHEETS

# --- IS YERI format ---
# Bu sekmenin varligi "is yeri formati" oldugunu gosterir.
WORKPLACE_MARKER_SHEET = "200-Mekanik"

# Kategori -> hedef sekme. None = kod verilmez (atla).
WORKPLACE_CATEGORY_SHEET = {
    "Mekanik":    "200-Mekanik",
    "Montaj":     "201-mekanik yarı mamül",
    "Elektronik": None,
    "Optik":      None,
}
# Hedef sekme -> yerlesim. geo_col yok: kopyaya dinamik olarak yeni sutun eklenir.
WORKPLACE_LAYOUTS = {
    # ekleyen_col=6 (Talep Eden), tarih_col=7 (Talep Tarihi), proje_col=8 (Proje Kodu).
    # scan_cols: bir satirin MUSAIT sayilmasi icin bu sutunlarin HEPSI bos olmali
    # (Yatay Rezervasyon Kontrolu). Sutun 1 (Sira No) ve 2 (Konfig No) HARIC tutulur —
    # bunlar havuzda onceden doldurulmus olabilir, dolu olmasi 'rezerve' demek degildir.
    "200-Mekanik":            {"header_row": 2, "code_col": 2, "name_col": 4, "proje_col": 8,
                               "ekleyen_col": 6, "tarih_col": 7, "scan_cols": [3, 4, 5, 6, 7, 8]},
    "201-mekanik yarı mamül": {"header_row": 1, "code_col": 1, "name_col": 4, "proje_col": 8,
                               "ekleyen_col": 6, "tarih_col": 7, "scan_cols": [3, 4, 5, 6, 7, 8]},
}
# Is yeri kopyasina eklenecek geometrik imza sutununun basligi.
GEO_HEADER = "Geometrik Imza"


def _is_workplace(sheetnames) -> bool:
    return WORKPLACE_MARKER_SHEET in sheetnames


def _detect_format(xlsx_path: str) -> str:
    """Dosyayi hizlica acip 'workplace' veya 'legacy' doner."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return "workplace" if _is_workplace(names) else "legacy"
    except Exception:
        return "legacy"


def _category_sheet(fmt: str, category: str) -> str | None:
    """Kategori icin hedef sekme adi. None ise kod verilmez (atlanir)."""
    if fmt == "workplace":
        # bilinen kategori degilse de atla (None)
        return WORKPLACE_CATEGORY_SHEET.get(category, None)
    # legacy: sekme adi = kategori adi (gecerliyse)
    return category if category in LEGACY_SHEETS else None


def _layout(fmt: str, sheet_name: str) -> dict:
    if fmt == "workplace":
        return WORKPLACE_LAYOUTS[sheet_name]
    return LEGACY_LAYOUT


def _target_sheets(fmt: str) -> list[str]:
    """Mevcut kayit tarama/temizleme icin hedef sekmeler."""
    if fmt == "workplace":
        return list(WORKPLACE_LAYOUTS.keys())
    return LEGACY_SHEETS


def _is_empty(value) -> bool:
    """Bir hucre 'bos' mu? None veya sadece bosluk ise bos sayilir."""
    return value is None or str(value).strip() == ""


def _write_cell(ws, row: int, col: int, value) -> None:
    """Bir hucreye deger yazar ve HEMEN ardindan standart bicimi zorla uygular
    (Force Formatting) — havuz hucrelerindeki bozuk (72pt/90 derece) formati ezer."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = std_font
    cell.alignment = std_align


def _geo_col(ws, layout: dict, header_row: int, create: bool = False) -> int | None:
    """Bu sekmedeki Geometrik Imza sutununun indeksini bulur.

    - Legacy: yerlesimde sabit 'geo_col' vardir, onu doner.
    - Workplace: baslik satirinda GEO_HEADER aranir. Yoksa ve create=True ise
      en sona yeni bir sutun olusturulup basligi yazilir. create=False ve yoksa
      None doner.
    """
    fixed = layout.get("geo_col")
    if fixed:
        return fixed
    # workplace: basliktan ara
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        if str(ws.cell(row=header_row, column=col).value or "").strip() == GEO_HEADER:
            return col
    if not create:
        return None
    new_col = max_col + 1
    ws.cell(row=header_row, column=new_col, value=GEO_HEADER)
    return new_col


# ---------------------------------------------------------------------------
# CALISMA KOPYASI (is yeri formatinda ana dosyaya dokunma)
# ---------------------------------------------------------------------------

def make_working_copy(xlsx_path: str) -> str:
    """Is yeri formatinda ana takip dosyasinin zaman damgali bir kopyasini
    olusturur ve KOPYANIN yolunu doner. Tum yazma islemleri bu kopyaya gider;
    ana dosyaya HIC dokunulmaz.

    Legacy (ev) formatinda orijinal yol degismeden doner (eski davranis korunur).
    """
    fmt = _detect_format(xlsx_path)
    if fmt != "workplace":
        return xlsx_path

    folder = os.path.dirname(os.path.abspath(xlsx_path))
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    copy_path = os.path.join(folder, f"TAKIP_calisma_{stamp}.xlsx")
    shutil.copy2(xlsx_path, copy_path)
    print("Calisma kopyasi olusturuldu (ana takip dosyasina dokunulmadi):")
    print(f"  {copy_path}")
    return copy_path


# ---------------------------------------------------------------------------
# MEVCUT KAYIT ESLEME (geometrik imza -> kod)
# ---------------------------------------------------------------------------

def build_existing_mapping(xlsx_path: str) -> dict[str, str]:
    """Dosyadan geometrik imza -> ERP kodu eslesmesini okur.

    NOT (is yeri formati): Geometrik imza sutunu yalnizca calisma kopyasinda
    olusturulup yazildigindan, ana dosyadan yeni bir kopya alindiginda bu
    veriler tasinmaz. Yani CALISMALAR ARASI mukerrer tespiti ancak kullanici
    calisma kopyasini ana dosyaya GERI BIRLESTIRDIKTEN sonra islerlik kazanir.
    """
    if not os.path.exists(xlsx_path):
        return {}

    existing: dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        fmt = "workplace" if _is_workplace(wb.sheetnames) else "legacy"
        for sheet_name in _target_sheets(fmt):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            layout = _layout(fmt, sheet_name)
            hr = layout["header_row"]
            geo_col = _geo_col(ws, layout, hr, create=False)
            if geo_col is None:
                continue  # bu sekmede henuz geo sutunu yok
            code_col = layout["code_col"]
            for row in range(hr + 1, (ws.max_row or 0) + 1):
                geo_val = ws.cell(row=row, column=geo_col).value
                kod_val = ws.cell(row=row, column=code_col).value
                if not _is_empty(geo_val) and not _is_empty(kod_val):
                    existing[str(geo_val).strip()] = str(kod_val).strip()
        wb.close()
    except Exception:
        pass
    return existing


# ---------------------------------------------------------------------------
# KOD ATAMA (rezervasyonlu)
# ---------------------------------------------------------------------------

def _cell_colored(cell) -> bool:
    """Hucrede arka plan boyamasi (fill) var mi? Bos/temasiz hucrede patternType None."""
    try:
        fill = cell.fill
        if fill is None:
            return False
        ptype = getattr(fill, "patternType", None)
        return ptype not in (None, "none")
    except Exception:
        return False


def _row_free(worksheet, row: int, layout: dict) -> bool:
    """Bir satir gercekten MUSAIT mi? (Yatay Rezervasyon Kontrolu)

    layout['scan_cols'] sutunlarinin HEPSI hem BOS hem de BOYANMAMIS olmalidir.
    Herhangi birinde veri VEYA boyama varsa satir REZERVE/dolu kabul edilir.
    Sira No (1) ve Konfig No (2) bu kontrole dahil DEGILDIR (scan_cols disindadir).
    """
    scan_cols = layout.get("scan_cols", [layout["name_col"]])
    for c in scan_cols:
        cell = worksheet.cell(row=row, column=c)
        if not _is_empty(cell.value):
            return False
        if _cell_colored(cell):
            return False
    return True


def find_next_free_code(worksheet, layout: dict, used_rows: set[int]) -> tuple[int, str] | None:
    """Gercekten MUSAIT olan, kodu hazir ILK satiri bulur (legacy/tek tek atama).

    used_rows: bu oturumda rezerve edilmis satirlar (ayni kod iki kez verilmez).
    Donus: (satir_no, kod) veya musait/kodlu satir kalmamissa None.
    """
    hr = layout["header_row"]
    code_col = layout["code_col"]
    for row in range(hr + 1, (worksheet.max_row or 0) + 1):
        if row in used_rows:
            continue
        if not _row_free(worksheet, row, layout):
            continue
        code = worksheet.cell(row=row, column=code_col).value
        if not _is_empty(code):
            return row, str(code).strip()
    return None


def _find_pool_block(worksheet, layout: dict, n: int) -> int | None:
    """Havuzda N+4 ardisik MUSAIT (free + kodu hazir) satirdan olusan ILK blogun
    baslangic satirini bulur. Veri satirlari start+2 .. start+2+N-1; basa/sona 2'ser
    tampon. Bulunamazsa None.
    """
    hr = layout["header_row"]
    code_col = layout["code_col"]
    need = n + 4
    start = None
    run = 0
    for row in range(hr + 1, (worksheet.max_row or 0) + 1):
        has_code = not _is_empty(worksheet.cell(row=row, column=code_col).value)
        if _row_free(worksheet, row, layout) and has_code:
            if start is None:
                start = row
            run += 1
            if run >= need:
                return start
        else:
            start = None
            run = 0
    return None


def _increment_code(code: str) -> str | None:
    """Kodun sonundaki sayisal grubu +1 artirir, genisligi (sifir dolgusu) korur.
    GMP-200-241816 -> GMP-200-241817 ; YMB-V00-R00-23030 -> YMB-V00-R00-23031.
    """
    import re
    m = re.search(r"(\d+)(\D*)$", code)
    if not m:
        return None
    num = m.group(1)
    inc = str(int(num) + 1).zfill(len(num))
    return code[:m.start(1)] + inc + m.group(2)


def _last_code_info(worksheet, layout: dict) -> tuple[int, str | None, int]:
    """Sekmedeki son kodlu satir no'sunu, son kodu ve en buyuk (sayisal) Sira No'yu doner."""
    hr = layout["header_row"]
    code_col = layout["code_col"]
    last_row = hr
    last_code = None
    last_sira = 0
    for row in range(hr + 1, (worksheet.max_row or 0) + 1):
        cv = worksheet.cell(row=row, column=code_col).value
        if not _is_empty(cv):
            last_row = row
            last_code = str(cv).strip()
        sv = worksheet.cell(row=row, column=1).value
        if isinstance(sv, (int, float)):
            last_sira = max(last_sira, int(sv))
        elif sv is not None and str(sv).strip().isdigit():
            last_sira = max(last_sira, int(str(sv).strip()))
    return last_row, last_code, last_sira


def _allocate_legacy(worksheet, layout: dict, sheet: str, sheet_parts: list[Part]) -> None:
    """Ev (legacy) formati: parca basina tek tek ilk musait kod (eski davranis)."""
    used: set[int] = set()
    for part in sheet_parts:
        result = find_next_free_code(worksheet, layout, used)
        if result is None:
            part.reason += f" | UYARI: '{sheet}' sekmesinde bos kod kalmadi"
            continue
        row, code = result
        used.add(row)
        part.erp_code   = code
        part._xlsx_row   = row    # type: ignore[attr-defined]
        part._xlsx_sheet = sheet  # type: ignore[attr-defined]
        part._xlsx_generated = False  # type: ignore[attr-defined]


def _allocate_workplace(worksheet, layout: dict, sheet: str, sheet_parts: list[Part]) -> None:
    """Is yeri formati: oturumun tum parcalarini BITISIK bir bloga, basa/sona
    2'ser bos satir birakarak yerlestirir. Havuzda boyle bir blok yoksa sayfa
    sonunda son koddan +1 artirarak yeni kodlar uretir (yine 2+2 tamponlu).
    """
    n = len(sheet_parts)
    code_col = layout["code_col"]

    # 1) Havuzda N+4 bitisik blok
    start = _find_pool_block(worksheet, layout, n)
    if start is not None:
        data_start = start + 2  # basta 2 tampon
        for i, part in enumerate(sheet_parts):
            row = data_start + i
            code = str(worksheet.cell(row=row, column=code_col).value).strip()
            part.erp_code   = code
            part._xlsx_row   = row     # type: ignore[attr-defined]
            part._xlsx_sheet = sheet   # type: ignore[attr-defined]
            part._xlsx_generated = False  # type: ignore[attr-defined]
        return

    # 2) Fallback: sayfa sonunda yeni kod uret
    last_row, last_code, last_sira = _last_code_info(worksheet, layout)
    if not last_code:
        for part in sheet_parts:
            part.reason += f" | UYARI: '{sheet}' yeni kod uretilemedi (referans kod yok)"
        return

    data_start = last_row + 3  # son veriden sonra 2 tampon, sonra blok
    cur = last_code
    for i, part in enumerate(sheet_parts):
        cur = _increment_code(cur)
        if cur is None:
            part.reason += f" | UYARI: kod artirilamadi ({last_code})"
            continue
        row = data_start + i
        part.erp_code   = cur
        part._xlsx_row   = row      # type: ignore[attr-defined]
        part._xlsx_sheet = sheet    # type: ignore[attr-defined]
        part._xlsx_generated = True  # type: ignore[attr-defined]
        # 200'de Sira No ayri sutun (1); 201'de kod zaten sutun 1, ayrica yazilmaz
        part._xlsx_sira = (last_sira + 1 + i) if code_col != 1 else None  # type: ignore[attr-defined]


def assign_codes(parts: list[Part], xlsx_path: str) -> list[Part]:
    """Her Part'a kategorisinin hedef sekmesinde kod atar.

    DOSYAYI DEGISTIRMEZ — sadece okur, Part nesnelerini doldurur. Gercek yazma
    write_back() ile yapilir.

    Is yeri formatinda hedefi olmayan kategoriler (Elektronik/Optik) STATUS_EXISTING
    olarak isaretlenir: kod verilmez, ama montaj kopyasinda orijinal adiyla yer
    alir (referans kopmaz). Kalan parcalar SEKME BAZLI, blok+tampon mantigiyla atanir.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"konfig dosyasi bulunamadi: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    fmt = "workplace" if _is_workplace(wb.sheetnames) else "legacy"

    # --- 1. GECIS: kategori -> hedef sekme cozumle, atlananları isaretle, grupla ---
    by_sheet: dict[str, list[Part]] = {}
    for part in parts:
        if part.status in (STATUS_EXISTING, STATUS_PREVIOUSLY_CODED) or part.category is None:
            continue
        # STEP / ice aktarilmis parca istisnasi: kod verilmez (guvenlik agi — main de
        # erken isaretler). Uzanti .stp/.step VEYA ad .stp/.step ile bitiyor VEYA
        # ozellik agacinda Imported (sw_reader -> part.is_imported).
        _nm = (part.original_name or "").strip().lower()
        if (part.extension in (".stp", ".step")
                or _nm.endswith((".stp", ".step"))
                or getattr(part, "is_imported", False)):
            part.status = STATUS_EXISTING
            part.reason += " | Imported/STEP parca — kod verilmez (orijinal adla kopyalanir)"
            continue
        category = part.category
        if not category:
            part.reason += " | UYARI: kategori yok, kod atanamadi"
            continue

        sheet = _category_sheet(fmt, category)
        if sheet is None:
            if fmt == "workplace" and category in ("Elektronik", "Optik"):
                part.status = STATUS_EXISTING
                part.reason += f" | {category}: kod verilmedi (atlandi), orijinal adla kopyalanir"
                continue
            elif fmt == "legacy":
                part.reason += f" | UYARI: gecersiz kategori '{category}', Mekanik'e donusturuldu"
                part.category = "Mekanik"
                part.status   = STATUS_UNKNOWN
                sheet = "Mekanik"
            else:
                part.reason += f" | UYARI: '{category}' icin hedef sekme yok, atlandi"
                continue

        if sheet not in wb.sheetnames:
            part.reason += f" | UYARI: '{sheet}' sekmesi dosyada yok"
            continue
        by_sheet.setdefault(sheet, []).append(part)

    # --- 2. GECIS: sekme bazli atama ---
    for sheet, sheet_parts in by_sheet.items():
        ws = wb[sheet]
        layout = _layout(fmt, sheet)
        if fmt == "legacy":
            _allocate_legacy(ws, layout, sheet, sheet_parts)
        else:
            _allocate_workplace(ws, layout, sheet, sheet_parts)

    wb.close()
    return parts


# ---------------------------------------------------------------------------
# YAZMA
# ---------------------------------------------------------------------------

def _backup(xlsx_path: str) -> str:
    """Dosyanin zaman damgali bir yedegini alir, yedek yolunu dondurur."""
    folder = os.path.dirname(os.path.abspath(xlsx_path))
    name = os.path.splitext(os.path.basename(xlsx_path))[0]
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = os.path.join(folder, f"{name}_yedek_{stamp}.xlsx")
    shutil.copy2(xlsx_path, backup_path)
    return backup_path


def write_back(parts: list[Part], xlsx_path: str, added_by: str | None = None,
               proje_kodu: str = "") -> str:
    """Atanan kodlari dosyaya yazar. Yazmadan ONCE yedek alir.

    Her kodlanan parca icin hedef sekmedeki satira:
      - Isim (Konfig Adi / Parca Adi) <- part.original_name
      - Proje Kodu                    <- proje_kodu (verilmisse)
      - Ekleyen/Talep Eden, Tarih     <- added_by, bugun (layout'ta sutun varsa)
      - Geometrik Imza                <- part.geometric_signature
    Is yeri formatinda YENI URETILEN satirlar icin ayrica Kod (ve 200'de Sira No)
    da yazilir; havuzdan alinanlarda kod zaten dosyada hazirdir.

    Donus: alinan yedegin dosya yolu.
    """
    if added_by is None:
        added_by = getpass.getuser()

    backup_path = _backup(xlsx_path)   # GUVENLIK AGI

    wb = openpyxl.load_workbook(xlsx_path)
    fmt = "workplace" if _is_workplace(wb.sheetnames) else "legacy"
    today = datetime.now()
    written = 0

    for part in parts:
        row = getattr(part, "_xlsx_row", None)
        sheet = getattr(part, "_xlsx_sheet", None) or part.category
        if row is None or not part.erp_code or not sheet:
            continue
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        layout = _layout(fmt, sheet)

        # YENI URETILEN satir: kodu (ve 200'de Sira No'yu) da yaz
        if getattr(part, "_xlsx_generated", False):
            _write_cell(ws, row, layout["code_col"], part.erp_code)
            sira = getattr(part, "_xlsx_sira", None)
            if sira is not None:
                _write_cell(ws, row, 1, sira)

        # Konfig Adi (isim) — Force Formatting ile yazilir
        _write_cell(ws, row, layout["name_col"], part.original_name)
        if proje_kodu and layout.get("proje_col"):
            _write_cell(ws, row, layout["proje_col"], proje_kodu)
        if layout.get("ekleyen_col"):
            _write_cell(ws, row, layout["ekleyen_col"], added_by)
        if layout.get("tarih_col"):
            _write_cell(ws, row, layout["tarih_col"], today)

        geo_col = _geo_col(ws, layout, layout["header_row"], create=True)
        if geo_col and part.geometric_signature:
            _write_cell(ws, row, geo_col, part.geometric_signature)

        written += 1

    wb.save(xlsx_path)
    wb.close()

    print(f"{written} kod dosyaya yazildi. Yedek: {backup_path}")
    return backup_path


def clear_old_record(part_name: str, xlsx_path: str) -> bool:
    """Belirtilen parca adini hedef sekmelerde tarayip eski kaydini bosaltir.

    ISIM sutununda part_name ile eslesen satirlarda Isim/Proje/Geo (ve legacy'de
    Ekleyen/Tarih) hucrelerini None yapar (satir tekrar kullanilabilir olur).
    Baslik satirlari korunur. En az bir satir bulunursa once yedek alinir.

    Donus: en az bir satir bosaltildiysa True.
    """
    if not os.path.exists(xlsx_path):
        return False

    wb = openpyxl.load_workbook(xlsx_path)
    fmt = "workplace" if _is_workplace(wb.sheetnames) else "legacy"
    hits: list[tuple[str, int]] = []

    for sheet_name in _target_sheets(fmt):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        layout = _layout(fmt, sheet_name)
        hr = layout["header_row"]
        name_col = layout["name_col"]
        for row in range(hr + 1, (ws.max_row or 0) + 1):
            cell_val = ws.cell(row=row, column=name_col).value
            if cell_val is not None and str(cell_val).strip() == part_name.strip():
                hits.append((sheet_name, row))

    if not hits:
        wb.close()
        return False

    backup_path = _backup(xlsx_path)

    for sheet_name, row in hits:
        ws = wb[sheet_name]
        layout = _layout(fmt, sheet_name)
        ws.cell(row=row, column=layout["name_col"]).value = None
        if layout.get("proje_col"):
            ws.cell(row=row, column=layout["proje_col"]).value = None
        if layout.get("ekleyen_col"):
            ws.cell(row=row, column=layout["ekleyen_col"]).value = None
        if layout.get("tarih_col"):
            ws.cell(row=row, column=layout["tarih_col"]).value = None
        geo_col = _geo_col(ws, layout, layout["header_row"], create=False)
        if geo_col:
            ws.cell(row=row, column=geo_col).value = None

    wb.save(xlsx_path)
    wb.close()
    print(f"      -> '{part_name}' eski kaydi silindi (yedek: {os.path.basename(backup_path)})")
    return True


def clear_old_records(part_names: list[str], xlsx_path: str) -> int:
    """clear_old_record'un TOPLU (performansli) surumu.

    Buyuk is yeri dosyasini her parca icin yeniden acmak yerine BIR KEZ acar,
    verilen tum parca adlarini tek taramada bulup bosaltir ve BIR KEZ kaydeder
    (tek yedek). 125 parca x (yukle+kaydet) yerine 1 yukle + 1 kaydet.

    Donus: bosaltilan satir sayisi.
    """
    if not os.path.exists(xlsx_path):
        return 0
    names = {str(n).strip() for n in part_names if n is not None and str(n).strip()}
    if not names:
        return 0

    wb = openpyxl.load_workbook(xlsx_path)
    fmt = "workplace" if _is_workplace(wb.sheetnames) else "legacy"
    hits: list[tuple[str, int]] = []

    for sheet_name in _target_sheets(fmt):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        layout = _layout(fmt, sheet_name)
        name_col = layout["name_col"]
        for row in range(layout["header_row"] + 1, (ws.max_row or 0) + 1):
            cell_val = ws.cell(row=row, column=name_col).value
            if cell_val is not None and str(cell_val).strip() in names:
                hits.append((sheet_name, row))

    if not hits:
        wb.close()
        return 0

    backup_path = _backup(xlsx_path)
    for sheet_name, row in hits:
        ws = wb[sheet_name]
        layout = _layout(fmt, sheet_name)
        ws.cell(row=row, column=layout["name_col"]).value = None
        if layout.get("proje_col"):
            ws.cell(row=row, column=layout["proje_col"]).value = None
        if layout.get("ekleyen_col"):
            ws.cell(row=row, column=layout["ekleyen_col"]).value = None
        if layout.get("tarih_col"):
            ws.cell(row=row, column=layout["tarih_col"]).value = None
        geo_col = _geo_col(ws, layout, layout["header_row"], create=False)
        if geo_col:
            ws.cell(row=row, column=geo_col).value = None

    wb.save(xlsx_path)
    wb.close()
    print(f"      -> {len(hits)} eski kayit bosaltildi (tek yedek: {os.path.basename(backup_path)})")
    return len(hits)


# --- Tek basina test bloku -------------------------------------------------
if __name__ == "__main__":
    import sys
    konfig = sys.argv[1] if len(sys.argv) > 1 else "konfig.xlsx"
    print("Format:", _detect_format(konfig))

    test_parts = [
        Part(r"C:\x\parca 1.SLDPRT", "parca 1", False, category="Mekanik"),
        Part(r"C:\x\parca 2.SLDPRT", "parca 2", False, category="Mekanik"),
        Part(r"C:\x\alt montaj 1.SLDASM", "alt montaj 1", True, category="Montaj"),
        Part(r"C:\x\lens 1.SLDPRT", "lens 1", False, category="Optik"),
    ]

    assign_codes(test_parts, konfig)
    print("ATAMA SONUCU (henuz dosyaya yazilmadi):")
    for p in test_parts:
        hedef = getattr(p, "_xlsx_sheet", "-")
        print(f"  {p.original_name:16} -> {p.erp_code}  [{p.category}] sekme={hedef} status={p.status}")
    print("\n(write_back cagrilmadigi icin dosya degismedi.)")
